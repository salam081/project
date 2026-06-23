from django.shortcuts import render,redirect,get_object_or_404
import calendar
from decimal import Decimal,DecimalException
from django.db import transaction
from django.http import HttpResponse

from django.db.models import Sum, F, DecimalField, OuterRef, Subquery
from django.db.models.functions import Coalesce

from django.db.models import Q, Sum, Count, Max, F, Min, OuterRef, Subquery, DecimalField ,Value
from django.db.models import F, ExpressionWrapper, DecimalField  
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Sum
from django.db.models.functions import ExtractYear, ExtractMonth
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import datetime, date, timedelta
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

from django.db.models import Q
from django.conf import settings
from django.contrib import messages
from django.db.models.functions import TruncMonth
from decimal import Decimal
from decimal import Decimal, InvalidOperation
from django.db import transaction
from decimal import Decimal
import openpyxl
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from accounts.decorators import group_required
from PurchasedItems.models import *
from django import forms
from projectfinance.models import *
from .models import *
from accounts.models import *
from consumable.models import *
from loan.models import *
from savings.models import *
from main.models import Withdrawal,UserActivity
from django.db.models.functions import ExtractMonth, ExtractYear
from accounts.utils import get_cooperative_withdrawal_stats, get_members_eligible_for_withdrawal
from accounts.views import *

# Create your views here.

@login_required
@group_required(['admin', 'staff'])
def payment_type_list(request):
    payment_types = PaymentType.objects.select_related().order_by("title")
    context = {"payment_types": payment_types}
    return render(request, "form_app/payment_type_list.html", context)

@login_required
@group_required(['admin'])
def payment_type_create(request):
    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        request_fee = request.POST.get("request_fee", 0)
        available = request.POST.get("available") == "on"

        if not title:
            messages.error(request, "Title is required.")
            return redirect("payment_type_create")

        PaymentType.objects.create(
            title=title,
            request_fee=request_fee,
            available=available,
        )
        messages.success(request, "Payment type created successfully.")
        return redirect("payment_type_list")

    return render(request, "form_app/payment_type_form.html")


@login_required
@group_required(['admin'])
def payment_type_edit(request, pk):
    payment_type = get_object_or_404(PaymentType, pk=pk)

    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        request_fee = request.POST.get("request_fee", 0)
        available = request.POST.get("available") == "on"

        if not title:
            messages.error(request, "Title is required.")
            return redirect("payment_type_edit", pk=pk)

        payment_type.title = title
        payment_type.request_fee = request_fee
        payment_type.available = available
        payment_type.save()

        messages.success(request, "Payment type updated successfully.")
        return redirect("payment_type_list")

    context = {"payment_type": payment_type}
    return render(request, "form_app/payment_type_form.html", context)


@login_required
@group_required(['admin'])
def payment_type_delete(request, pk):
    payment_type = get_object_or_404(PaymentType, pk=pk)

    if request.method == "POST":
        payment_type.delete()
        messages.success(request, "Payment type deleted successfully.")
        return redirect("payment_type_list")

    context = {"payment_type": payment_type}
    return render(request, "form_app/payment_type_confirm_delete.html", context)


# ── REQUEST FORM PAYMENT VIEWS ─────────────────────────────────────────────────

@login_required
@group_required(['admin', 'staff'])
def request_form_payment_list(request):
    payments = RequestFormPayment.objects.select_related(
        "payment_type", "member", "member__member", "created_by"
    ).order_by("-date_created")

    # filters
    status_filter = request.GET.get("status", "")
    payment_type_filter = request.GET.get("payment_type", "")
    search = request.GET.get("search", "").strip()
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    if status_filter:
        payments = payments.filter(status=status_filter)

    if payment_type_filter:
        payments = payments.filter(payment_type_id=payment_type_filter)

    if search:
        payments = payments.filter(
            models.Q(member__ippis__icontains=search) |
            models.Q(guest_ippis__icontains=search) |
            models.Q(guest_name__icontains=search)
        )
    if date_from:
        try:
            payments = payments.filter(date_created__date__gte=date.fromisoformat(date_from))
        except ValueError:
            messages.error(request, "Invalid 'Date From' format.")

    if date_to:
        try:
            payments = payments.filter(date_created__date__lte=date.fromisoformat(date_to))
        except ValueError:
            messages.error(request, "Invalid 'Date To' format.")
    
    
    # ── TOTALS ─────────────────────────────────────────────────────
    totals = payments.aggregate(
        total_amount=Sum("amount"),
        total_fee=Sum("payment_type__request_fee"),
    )
    
    total_amount = totals["total_amount"] or Decimal("0.00")
    total_fee = totals["total_fee"] or Decimal("0.00")
    
    payment_types = PaymentType.objects.filter(available=True)

    context = {
        "payments": payments,
        "payment_types": payment_types,
        "status_filter": status_filter,
        "payment_type_filter": payment_type_filter,
        "search": search,
        "date_from": date_from,
        "date_to": date_to,
        "total_amount": total_amount,
        "total_fee": total_fee,
    }
    return render(request, "form_app/request_form_payment_list.html", context)


@login_required
@group_required(['admin', 'staff'])
def request_form_payment_create(request):
    payment_types = PaymentType.objects.filter(available=True)
    member_info = None
    if request.method == "POST":

        # ── SEARCH MEMBER ──────────────────────────────────────────
        if "search_member" in request.POST:
            ippis = request.POST.get("ippis", "").strip()
            try:
                member = Member.objects.get(ippis=ippis)
                member_info = {
                    "id": member.id,
                    "name": f"{member.member.first_name} {member.member.last_name}",
                    "ippis": member.ippis,
                }
            except Member.DoesNotExist:
                messages.error(request, f"No member found with IPPIS {ippis}.")

        # ── SAVE MEMBER PAYMENT ────────────────────────────────────
        elif "save_member_payment" in request.POST:
            member_id = request.POST.get("member_id")
            payment_type_id = request.POST.get("payment_type")
            amount = request.POST.get("amount", "").strip()
            duration = request.POST.get("duration", "").strip()

            try:
                member = Member.objects.get(id=member_id)
                payment_type = PaymentType.objects.get(id=payment_type_id)

                amount = Decimal(amount) if amount else None

                if amount is not None and amount < payment_type.request_fee:
                    messages.error(request, f"Amount must be at least ₦{payment_type.request_fee}.")
                    return redirect("request_form_payment_create")

                RequestFormPayment.objects.create(
                    payment_type=payment_type,
                    member=member,
                    amount=amount,
                    duration=duration,
                    # status="paid",
                    status="used" if payment_type.title in ["Savings", "Target Savings", "Special Savings"] else "paid",
                    created_by=request.user,
                )
                messages.success(request, "Member payment recorded successfully.")
                return redirect("request_form_payment_list")

            except Member.DoesNotExist:
                messages.error(request, "Member not found.")
            except PaymentType.DoesNotExist:
                messages.error(request, "Payment type not found.")
            except Exception as e:
                messages.error(request, str(e))

        # ── SAVE GUEST PAYMENT ─────────────────────────────────────
        elif "save_guest_payment" in request.POST:
            guest_name = request.POST.get("guest_name", "").strip()
            guest_ippis = request.POST.get("guest_ippis", "").strip()
            guest_phone = request.POST.get("guest_phone", "").strip()
            payment_type_id = request.POST.get("payment_type")
            amount = request.POST.get("amount", "").strip()
            duration = request.POST.get("duration", "").strip()

            if not guest_name or not guest_ippis:
                messages.error(request, "Guest name and IPPIS are required.")
                return redirect("request_form_payment_create")

            try:
                payment_type = PaymentType.objects.get(id=payment_type_id)

                amount = Decimal(amount) if amount else None

                if amount is not None and amount < payment_type.request_fee:
                    messages.error(request, f"Amount must be at least ₦{payment_type.request_fee}.")
                    return redirect("request_form_payment_create")

                RequestFormPayment.objects.create(
                    payment_type=payment_type,
                    guest_name=guest_name,
                    guest_ippis=guest_ippis,
                    guest_phone=guest_phone,
                    amount=amount,
                    duration=duration,
                    status="paid",
                    created_by=request.user,
                )
                messages.success(request, "Guest payment recorded successfully.")
                return redirect("request_form_payment_list")

            except PaymentType.DoesNotExist:
                messages.error(request, "Payment type not found.")
            except Exception as e:
                messages.error(request, str(e))

    context = {
        "payment_types": payment_types,
        "member_info": member_info,
    }
    return render(request, "form_app/request_form_payment_create.html", context)

@login_required
@group_required(['admin', 'staff'])
def request_form_payment_detail(request, pk):
    payment = get_object_or_404(
        RequestFormPayment.objects.select_related(
            "payment_type", "member", "member__member", "created_by"), pk=pk,)
    context = {"payment": payment}
    return render(request, "form_app/request_form_payment_detail.html", context)


@login_required
@group_required(['admin'])
def request_form_payment_delete(request, pk):
    payment = get_object_or_404(RequestFormPayment, pk=pk)

    if request.method == "POST":
        payment.delete()
        messages.success(request, "Payment deleted successfully.")
        return redirect("request_form_payment_list")

    context = {"payment": payment}
    return render(request, "form_app/request_form_payment_confirm_delete.html", context)



@login_required
def request_form_payment_export_pdf(request):
    payments = RequestFormPayment.objects.select_related(
        "payment_type", "member", "member__member", "created_by"
    ).order_by("-date_created")

    # ── APPLY SAME FILTERS ─────────────────────────────────────────
    status_filter = request.GET.get("status", "")
    payment_type_filter = request.GET.get("payment_type", "")
    search = request.GET.get("search", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()

    if status_filter:
        payments = payments.filter(status=status_filter)

    if payment_type_filter:
        payments = payments.filter(payment_type_id=payment_type_filter)

    if search:
        payments = payments.filter(
            Q(member__ippis__icontains=search) |
            Q(guest_ippis__icontains=search) |
            Q(guest_name__icontains=search)
        )

    if date_from:
        try:
            payments = payments.filter(date_created__date__gte=date.fromisoformat(date_from))
        except ValueError:
            pass

    if date_to:
        try:
            payments = payments.filter(date_created__date__lte=date.fromisoformat(date_to))
        except ValueError:
            pass

    # ── RESPONSE ───────────────────────────────────────────────────
    filename = f"payments_{date.today().strftime('%Y%m%d')}.pdf"
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # ── TITLE ──────────────────────────────────────────────────────
    title_style = ParagraphStyle(
        "title", parent=styles["Heading1"],
        alignment=TA_CENTER, fontSize=14, spaceAfter=4
    )
    sub_style = ParagraphStyle(
        "sub", parent=styles["Normal"],
        alignment=TA_CENTER, fontSize=9,
        textColor=colors.grey, spaceAfter=10
    )

    elements.append(Paragraph("Request Form Payments Report", title_style))

    # subtitle with active filters info
    filter_parts = []
    if date_from:
        filter_parts.append(f"From: {date_from}")
    if date_to:
        filter_parts.append(f"To: {date_to}")
    if status_filter:
        filter_parts.append(f"Status: {status_filter.capitalize()}")
    if search:
        filter_parts.append(f"Search: {search}")

    subtitle = " | ".join(filter_parts) if filter_parts else "All Records"
    elements.append(Paragraph(subtitle, sub_style))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
        sub_style
    ))
    elements.append(Spacer(1, 6 * mm))

    # ── TABLE DATA ─────────────────────────────────────────────────
    headers = ["#", "Payment Type", "Member / Guest", "IPPIS", "Phone",
               "Amount (₦)", "Fee (₦)", "Status", "Recorded By", "Date"]

    data = [headers]
    total_amount = Decimal("0.00")
    total_fee = Decimal("0.00")

    for idx, payment in enumerate(payments, start=1):
        if payment.member:
            name = f"{payment.member.member.first_name} {payment.member.member.last_name}"
            ippis = payment.member.ippis
            phone = "—"
        else:
            name = payment.guest_name or "—"
            ippis = payment.guest_ippis or "—"
            phone = payment.guest_phone or "—"

        amount = payment.amount or Decimal("0.00")
        fee = payment.payment_type.request_fee or Decimal("0.00")
        total_amount += amount
        total_fee += fee

        recorded_by = ""
        if payment.created_by:
            recorded_by = payment.created_by.get_full_name() or payment.created_by.username

        data.append([
            str(idx),
            payment.payment_type.title,
            name,
            ippis,
            phone,
            f"{amount:,.2f}",
            f"{fee:,.2f}",
            payment.get_status_display(),
            recorded_by,
            payment.date_created.strftime("%d %b %Y"),
        ])

    # ── TOTALS ROW ─────────────────────────────────────────────────
    data.append([
        "", "", "", "", "TOTAL",
        f"{total_amount:,.2f}",
        f"{total_fee:,.2f}",
        "", "", ""
    ])

    # ── TABLE STYLE ────────────────────────────────────────────────
    col_widths = [8*mm, 30*mm, 40*mm, 25*mm, 22*mm,
                  22*mm, 20*mm, 15*mm, 30*mm, 22*mm]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # header
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#343a40")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 6),

        # body
        ("FONTNAME", (0, 1), (-1, -2), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -2), 7.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8f9fa")]),
        ("ALIGN", (0, 1), (-1, -1), "CENTER"),
        ("ALIGN", (2, 1), (2, -2), "LEFT"),  # name left aligned
        ("TOPPADDING", (0, 1), (-1, -2), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -2), 4),

        # totals row
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e9ecef")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 8),
        ("TOPPADDING", (0, -1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, -1), (-1, -1), 5),

        # grid
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dee2e6")),
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#343a40")),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.HexColor("#adb5bd")),
    ]))

    elements.append(table)

    # ── SUMMARY BELOW TABLE ────────────────────────────────────────
    elements.append(Spacer(1, 6 * mm))
    summary_style = ParagraphStyle(
        "summary", parent=styles["Normal"],
        fontSize=9, spaceAfter=3
    )
    elements.append(Paragraph(f"<b>Total Records:</b> {len(data) - 2}", summary_style))
    elements.append(Paragraph(f"<b>Total Amount Collected:</b> ₦{total_amount:,.2f}", summary_style))
    elements.append(Paragraph(f"<b>Total Request Fees:</b> ₦{total_fee:,.2f}", summary_style))

    doc.build(elements)
    return response


import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

@login_required
def request_form_payment_export_excel(request):
    payments = RequestFormPayment.objects.select_related(
        "payment_type", "member", "member__member", "created_by"
    ).order_by("-date_created")

    # ── APPLY SAME FILTERS AS LIST VIEW ───────────────────────────
    status_filter = request.GET.get("status", "")
    payment_type_filter = request.GET.get("payment_type", "")
    search = request.GET.get("search", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()

    if status_filter:
        payments = payments.filter(status=status_filter)

    if payment_type_filter:
        payments = payments.filter(payment_type_id=payment_type_filter)

    if search:
        payments = payments.filter(
            Q(member__ippis__icontains=search) |
            Q(guest_ippis__icontains=search) |
            Q(guest_name__icontains=search)
        )

    if date_from:
        try:
            payments = payments.filter(date_created__date__gte=date.fromisoformat(date_from))
        except ValueError:
            pass

    if date_to:
        try:
            payments = payments.filter(date_created__date__lte=date.fromisoformat(date_to))
        except ValueError:
            pass

    # ── BUILD WORKBOOK ─────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments"

    # styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="343a40", end_color="343a40", fill_type="solid")
    center = Alignment(horizontal="center")
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="e9ecef", end_color="e9ecef", fill_type="solid")

    # ── HEADER ROW ─────────────────────────────────────────────────
    headers = [
        "S/N","Member / Guest", "Payment Type",  "IPPIS",
        "Phone", "Amount (₦)", "Request Fee (₦)",
        "Status", "Recorded By", "Date"
    ]
    ws.append(headers)

    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # ── DATA ROWS ──────────────────────────────────────────────────
    total_amount = Decimal("0.00")
    total_fee = Decimal("0.00")

    for idx, payment in enumerate(payments, start=1):
        if payment.member:
            name = f"{payment.member.member.first_name} {payment.member.member.last_name}"
            ippis = payment.member.ippis
            phone = ""
        else:
            name = payment.guest_name or "—"
            ippis = payment.guest_ippis or "—"
            phone = payment.guest_phone or "—"

        amount = payment.amount or Decimal("0.00")
        fee = payment.payment_type.request_fee or Decimal("0.00")
        total_amount += amount
        total_fee += fee

        recorded_by = ""
        if payment.created_by:
            recorded_by = payment.created_by.get_full_name() or payment.created_by.username

        ws.append([
            idx,
            name,
            payment.payment_type.title,
            ippis,
            phone,
            float(amount),
            float(fee),
            payment.get_status_display(),
            recorded_by,
            payment.date_created.strftime("%d %b %Y %H:%M"),
        ])

    # ── TOTALS ROW ─────────────────────────────────────────────────
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=5).value = "TOTAL"
    ws.cell(row=total_row, column=6).value = float(total_amount)
    ws.cell(row=total_row, column=7).value = float(total_fee)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.font = total_font
        cell.fill = total_fill

    # ── COLUMN WIDTHS ──────────────────────────────────────────────
    column_widths = [5, 20, 25, 15, 15, 15, 18, 10, 20, 20]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # ── RESPONSE ───────────────────────────────────────────────────
    filename = f"payments_{date.today().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response