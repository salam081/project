from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth.decorators import login_required,user_passes_test
from accounts.decorators import group_required
from django.template.loader import get_template
from openpyxl.utils import get_column_letter
from django.views.decorators.http import require_http_methods
from django.utils.dateparse import parse_date
from django.template.loader import render_to_string
from datetime import date

from django.db.models.functions import TruncMonth
# from accounts.decorator import group_required
from django.db.models import Sum,Count, Q , Avg, Prefetch,F
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse
from django.http import JsonResponse
from xhtml2pdf import pisa
from io import BytesIO
from collections import defaultdict
from decimal import Decimal
import pandas as pd
from datetime import datetime
from .models import LoanRequest, LoanRepayback
import openpyxl
from django.core.paginator import Paginator,PageNotAnInteger,EmptyPage
from django.db.models.functions import ExtractYear
from django.db import transaction
import os
from django.conf import settings

from openpyxl import Workbook
from decimal import Decimal, ROUND_HALF_UP


from .forms import *
from .models import *
from accounts.models import *
from main.models import *
from member.models import *
from consumable.models import *


@login_required
@group_required(['admin','staff'])
def admin_loan_settings(request):
    try:
        settings = LoanSettings.objects.latest('id')
    except LoanSettings.DoesNotExist:
        settings = None
    
    if request.method == 'POST':
        form = LoanSettingsForm(request.POST, instance=settings)
        if form.is_valid():
            settings = form.save(commit=False)
            settings.created_by = request.user
            settings.save()
            messages.success(request, 'Settings updated successfully')
            return redirect('admin_loan_settings')
    else:
        form = LoanSettingsForm(instance=settings)
    # Loan types management
    loan_types = LoanType.objects.all()
    
    context = {'form': form,'settings': settings,'loan_types': loan_types,}
    return render(request, 'loan/settings.html', context)

@login_required
@group_required(['admin','staff'])
def add_loan_type(request):
    loan_types = LoanType.objects.all()

    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        max_amount = request.POST.get('max_amount') or None
        max_loan_term_months = request.POST.get('max_loan_term_months') or None
        loan_type_id = request.POST.get('loan_type_id')
        request_fee = request.POST.get('request_fee')
        action = request.POST.get('action')  # 'toggle' or 'edit'

        if loan_type_id:
            loan_type = get_object_or_404(LoanType, id=loan_type_id)

            if action == 'toggle':
                loan_type.available = not loan_type.available
                loan_type.save()
                messages.success(request, 'Loan type availability updated successfully.')
                return redirect('add_loan_type')

            elif action == 'edit':
                loan_type.name = name
                loan_type.description = description
                loan_type.request_fee = request_fee
                loan_type.max_amount = max_amount
                loan_type.max_loan_term_months = max_loan_term_months
                loan_type.save()
                messages.success(request, 'Loan type updated successfully.')
                return redirect('add_loan_type')

        else:
            LoanType.objects.create(name=name,description=description, request_fee=request_fee, max_amount=max_amount,
                max_loan_term_months=max_loan_term_months,available=True, created_by=request.user)
            messages.success(request, 'Loan type created successfully.')
            return redirect('add_loan_type')

    context = {'loan_types': loan_types}
    return render(request, 'loan/add_loan_type.html', context)

@login_required
@group_required(['admin','staff'])
def loan_request_fee_payment(request):
    member_info = None
    loan_types = LoanType.objects.filter(available=True)

    selected_loan_type_id = request.GET.get("loan_type")  # <-- new filter input

    if request.method == "POST":
        if "search_member" in request.POST:  # step 1: search
            ippis = request.POST.get("ippis")
            try:
                member = Member.objects.get(ippis=ippis)
            except Member.DoesNotExist:
                messages.error(request, f"No member found with IPPIS {ippis}.")
                member = None

            if member:
                latest_loanable = Loanable.objects.filter(member=member).order_by('-month').first()
                total_loanable = latest_loanable.total_amount if latest_loanable else Decimal("0.00")

                member_info = {
                    "id": member.id,
                    "name": f"{member.member.first_name} {member.member.last_name}",
                    "ippis": member.ippis,
                    "total_loanable": total_loanable,
                }

        elif "make_payment" in request.POST:  # step 2: make payment
            member_id = request.POST.get("member_id")
            loan_type_id = request.POST.get("loan_type")
            loan_amount = Decimal(request.POST.get("loan_amount") or "0.00")

            member = get_object_or_404(Member, id=member_id)
            loan_type = get_object_or_404(LoanType, id=loan_type_id)

            # Prevent duplicate fee payments
            if LoanRequestFee.objects.filter(member=member, loan_type=loan_type).exists():
                messages.warning(request, f"{member} has already paid the request fee for {loan_type.name}.")
                return redirect("loan_request_fee_payment")

            # Create the record
            LoanRequestFee.objects.create(
                member=member,
                loan_type=loan_type,
                form_fee=loan_type.request_fee,
                loan_amount=loan_amount,
                created_by=request.user
            )
            messages.success(request, f"Loan request fee of ₦{loan_type.request_fee} recorded for {member}.")
            return redirect("loan_request_fee_payment")

    # ==============================
    # Filter LoanRequestFee records
    # ==============================
    members = LoanRequestFee.objects.select_related('member', 'loan_type')

    if selected_loan_type_id:
        members = members.filter(loan_type_id=selected_loan_type_id)

    # Aggregates
    loan = members.aggregate(total=Sum('loan_amount'))['total'] or Decimal("0.00")
    fee = members.aggregate(total=Sum('form_fee'))['total'] or Decimal("0.00")
    loan_req_form = members.count()

    # Pagination
    page_number = request.GET.get('page')
    paginator = Paginator(members, 100)
    page_obj = paginator.get_page(page_number)

    context = {
        "fee": fee,
        "loan": loan,
        "loan_req_form": loan_req_form,
        "members": members,
        "page_obj": page_obj,
        "loan_types": loan_types,
        "selected_loan_type_id": selected_loan_type_id,
        "member_info": member_info,  
    }

    return render(request, "loan/loan_request_fee.html", context)

#  =========list of Pending Loans and others ==========
@login_required
@group_required(['admin','staff'])
def admin_loan_requests_list(request):
    requests_list = LoanRequest.objects.select_related('member', 'loan_type', 'bank_name').order_by('-date_created')
    # Filtering
    status_filter = request.GET.get('status')
    loan_type_filter = request.GET.get('loan_type')
    search_query = request.GET.get('search')
    
    if status_filter:
        requests_list = requests_list.filter(status=status_filter)
    
    if loan_type_filter:
        requests_list = requests_list.filter(loan_type_id=loan_type_filter)
    
    if search_query:
        requests_list = requests_list.filter(
            Q(member__member__first_name__icontains=search_query) |
            Q(member__member__last_name__icontains=search_query) |
            Q(member__ippis__icontains=search_query) 
            # Q(guarantor__member__member__first_name__icontains=search_query)
        )
    else:
         results_queryset = requests_list
    results_queryset = results_queryset.order_by('status')
    total_approved_amount = results_queryset.aggregate(total=Sum('approved_amount'))['total'] or 0
    totals_by_status = dict(
        results_queryset.values('status')
        .annotate(total=Sum('amount'))
        .values_list('status', 'total')
    )
    total_repaid = LoanRepayback.objects.filter(loan_request__in=results_queryset).aggregate(total=Sum('amount_paid'))['total'] or 0
    total_amont_loan_request = totals_by_status.get('approved', 0)
    total_pending = totals_by_status.get('pending', 0)

    # Totals by status
    totals_by_status = dict(
        results_queryset.values('status')
        .annotate(total=Sum('approved_amount'))
        .values_list('status', 'total')
    )
    # Pagination
    paginator = Paginator(requests_list, 100)
    page_number = request.GET.get('page')
    requests = paginator.get_page(page_number)
    # Filter options
    loan_types = LoanType.objects.all()
    status_choices = LoanRequest._meta.get_field('status').choices
    context = {
        'requests': requests,'loan_types': loan_types,
        'status_choices': status_choices,'current_status': status_filter,
        'current_loan_type': loan_type_filter,'search_query': search_query,
        'total_approved': total_amont_loan_request,'total_pending': total_pending,
        'total_repaid': total_repaid,'total_approved_amount': total_approved_amount,
    }
    return render(request, 'loan/requests_list.html', context)


# ========admin loan request details=========

@login_required
@group_required(['admin','staff'])
def loan_request_detail(request, id):
    loan_request = get_object_or_404(LoanRequest, id=id)
    repayments = loan_request.repaybacks.all().order_by('-repayment_date')
    monthly_payment = monthly_payment = loan_request.monthly_payment or 0
    # Calculate repayment summary
    total_paid = repayments.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
    balance = (loan_request.approved_amount or 0) - total_paid
    
    context = {'loan_request': loan_request, 'repayments': repayments,'total_paid': total_paid, 'balance': balance,
        'monthly_payment': monthly_payment,'title': f'Loan Request #{loan_request.id}'}
    return render(request, 'loan/loan_request_detail.html', context)

def is_admin(user):
    return user.is_staff

@login_required
@group_required(['admin'])
def approve_loan_request(request, id):
    loan_request = get_object_or_404(LoanRequest, id=id, status='pending')
    member = loan_request.member

    loanable = Loanable.objects.filter(member=member).aggregate(
        total=Sum('amount')
    )['total'] or Decimal("0.00")

    #  Only require guarantor acceptance for NON short-term loans
    if loan_request.loan_type and "short" not in loan_request.loan_type.name.lower():
        if not loan_request.guarantor_accepted:
            messages.error(request, "This loan cannot be approved because the guarantor has not accepted yet.")
            return redirect('admin_loan_requests')


    if request.method == "POST":
        approved_amount = request.POST.get('approved_amount')

        if not approved_amount:
            messages.error(request, "Please enter the approved loan amount.")
            return redirect('approve_loan_request', id=id)

        try:
            approved_amount = float(approved_amount)
            if approved_amount <= 0:
                messages.error(request, "Approved amount must be greater than zero.")
                return redirect('approve_loan_request', id=id)
            
             # ✅ NEW CHECK — matches your DB constraint
            if approved_amount > loan_request.amount:
                messages.error(
                    request,
                    "Approved amount cannot exceed requested amount."
                )
                return redirect('approve_loan_request', id=id)
            
            if (
                loan_request.loan_type 
                and loan_request.loan_type.max_amount is not None 
                and approved_amount > loan_request.loan_type.max_amount
                ):
                messages.error(
                    request,
                    f"Approved amount cannot exceed the maximum allowed: {loan_request.loan_type.max_amount}"
                )
                return redirect('approve_loan_request', id=id)
            # ✅ Save approval
            loan_request.approved_amount = approved_amount
            loan_request.approval_date = timezone.now().date()
            loan_request.status = 'approved'
            loan_request.approved_by = request.user
            loan_request.save()

            messages.success(
                request,
                f"Loan request ID {loan_request.id} has been approved for ₦{loan_request.approved_amount}."
            )
            return redirect('admin_loan_requests')

        except ValueError:
            messages.error(request, "Invalid approved amount.")
            return redirect('approve_loan_request', id=id)

    context = {
        'loan_request': loan_request,
        'loanable': loanable
    }
    return render(request, 'loan/approve_loan.html', context)

@login_required
@group_required(['admin','staff'])
def payslip_img_details(request, id):
    payslip_img = LoanRequest.objects.get(id=id)
    context = {'payslip_img': payslip_img}
    return render(request, 'loan/payslip_img_details.html', context)

#======edit loan request==========

@login_required
@group_required(['admin'])
def edit_requested_loan(request, id):
    loan_types = LoanType.objects.all()
    loanobj = LoanRequest.objects.get(id=id)

    # If user is staff or superuser, use the loan's member
    if request.user.is_staff or request.user.is_superuser:
        member = loanobj.member
    else:
        try:
            member = request.user.member
        except Member.DoesNotExist:
            messages.error(request, "You are not registered as a member.")
            return redirect('some_page')  

        if loanobj.member != member:
            messages.error(request, "You are not allowed to edit this request.")
            return redirect('requested_loan')

    if request.method == "POST":
        loan_type = request.POST['loan_type']
        amount = request.POST['amount']
        loan_term_months = request.POST['loan_term_months']

        LoanRequest.objects.filter(id=id).update( member=member, loan_type_id=loan_type, amount=amount,loan_term_months=loan_term_months, approved_amount=0,)
        return redirect('admin_loan_requests')

    context = {'loanobj': loanobj, 'loan_types': loan_types}
    return render(request, 'loan/edit_requested_loan.html', context)

#=======reject loan request========
@require_http_methods(["GET", "POST"])
def reject_loan_request(request, id):
    loan_request = LoanRequest.objects.filter(id=id).first()
    if not loan_request:
        messages.error(request, f"No LoanRequest with ID {id} found.")
        return redirect('requested_loan')

    if loan_request.status != 'pending':
        messages.warning(request, f"LoanRequest {id} is already {loan_request.status}. Cannot reject.")
        return redirect('requested_loan')

    if request.method == 'POST':
        reason = request.POST.get('rejection_reason')
        if not reason:
            messages.error(request, "Rejection reason is required.")
            return redirect('reject_loan_request', id=id)

        loan_request.status = 'rejected'
        loan_request.rejection_reason = reason
        loan_request.approval_date = timezone.now().date()
        loan_request.approved_by = request.user
        loan_request.save()

        messages.success(request, f"Loan request ID {loan_request.id} has been rejected with reason.")
        return redirect('admin_loan_requests')

    return render(request, 'loan/reject_loan_form.html', {'loan': loan_request})

# ======admin view rejected loan=========
def all_reject_loan(request):
    rejected = LoanRequest.objects.filter(status='rejected')
    return render(request,'loan/all_reject_loan.html',{'rejected':rejected})

# ======admin delete rejected loan=========
@login_required
@group_required(['admin'])
def delete_reject_loan(request,id):
    rejectObj = LoanRequest.objects.get(id=id)
    rejectObj.delete()
    return redirect('all_reject_loan')


#========list of loan request in a year===========
@login_required
@group_required(['admin','staff'])
def loan_years_list(request):
    # Get distinct year and loan_type combinations
    loans = LoanRequest.objects.annotate(year=ExtractYear('application_date')).values('year', 'loan_type__name').distinct().order_by('-year', 'loan_type__name')

    # Structure the data as {2025: ['LONG TERM LOAN'], 2024: ['SHORT TERM LOAN', ...]}
    year_to_loan_types = {}
    for loan in loans:
        year = loan['year']
        loan_type = loan['loan_type__name']
        year_to_loan_types.setdefault(year, []).append(loan_type)

    context = {'year_to_loan_types': year_to_loan_types,}
    return render(request, "loan/loan_years_list.html", context)

#========list of loan request in a year details===========

def loans_by_year(request, year, loan_type_filter):
    try:
        loan_type = get_object_or_404(LoanType, name__iexact=loan_type_filter)
    except LoanType.DoesNotExist:
        messages.error(request, 'Loan type does not exist.')
        return render(request, "loan/loans_by_year.html", {'loanobj': []})

    status_filter = request.GET.get('status')

    # Filter loans by type and year
    loanobj = LoanRequest.objects.filter(
        loan_type=loan_type, date_created__year=year
    )

    # Optional: Filter by status if given
    if status_filter:
        loanobj = loanobj.filter(status__iexact=status_filter)

    # Totals by status
    totals_by_status = dict(
        loanobj.values('status')
        .annotate(total=Sum('approved_amount'))
        .values_list('status', 'total')
    )

    context = {
        'year': year, 'loan_type': loan_type,'loanobj': loanobj,
        'totals_by_status': totals_by_status,'selected_status': status_filter,}

    # Handle PDF download
    if request.GET.get('download') == 'pdf':
        template_path = 'loan/loans_by_year_pdf.html'
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="loans_{loan_type.name}_{year}.pdf"'
        template = get_template(template_path)
        html = template.render(context)
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
        return response

    # Handle Excel download
    if request.GET.get('download') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Loan Data"

        headers = ['ID', 'First Name', 'Last Name', 'Other Name', 'Approved Amount', 'Account Number', 'Bank Name', 'Bank Code', 'Duration Month']
        ws.append(headers)

        for loan in loanobj:
            ws.append([
                loan.id,
                str(loan.member.member.first_name),
                str(loan.member.member.last_name),
                str(loan.member.member.other_name),
                loan.approved_amount if loan.approved_amount is not None else 'N/A',
                loan.account_number,
                str(loan.bank_name),
                str(loan.bank_code.name),
                loan.loan_term_months
            ])

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="loans_{loan_type.name}_{year}.xlsx"'
        wb.save(response)
        return response

    return render(request, "loan/loans_by_year.html", context)
# =========Loan Payment Section==========

# ====== make payment for one member =========

@login_required
@group_required(['admin'])
def add_payment(request):
    requests_list = []
    selected_user = None
    ippis = request.GET.get("ippis") or request.POST.get("ippis")

    total_paid = Decimal("0.00")
    remaining_balance = Decimal("0.00")

    if ippis:
        try:
            member_obj = Member.objects.filter(ippis=int(ippis)).first()
            if member_obj and member_obj.member:
                selected_user = member_obj.member
                requests_list = LoanRequest.objects.filter(
                    member__member=selected_user
                ).exclude(status__in=['Fullpaid', 'Declined'])

                # Calculate total paid & balance for the selected user
                for req in requests_list:
                    paid = LoanRepayback.objects.filter(loan_request=req).aggregate(
                        total=Sum('amount_paid')
                    )['total'] or Decimal("0.00")

                    approved = req.approved_amount or Decimal("0.00")  # use approved amount
                    req.total_paid = paid
                    req.remaining_balance = approved - paid

                    total_paid += paid
                    remaining_balance += req.remaining_balance

        except Exception as e:
            messages.error(request, f"Error fetching member: {e}")

    if request.method == "POST":
        amount_paid = request.POST.get("amount_paid")
        month = request.POST.get("repayment_date")
        request_id = request.POST.get("loan_request")
        comment = request.POST.get("comment")
        loan_receipt = request.FILES.get("loan_receipt")


        # Validate required fields
        if not (ippis and amount_paid and month and request_id):
            messages.error(request, "All fields are required.")
            return redirect(f"{request.path}?ippis={ippis}")

        try:
            amount_paid = Decimal(amount_paid)
            if amount_paid <= 0:
                raise ValueError("Amount must be positive")
            month_date = datetime.strptime(month, "%Y-%m-%d").date()
            # month_date = datetime.strptime(month, "%Y-%m").date()
            request_id = int(request_id)
        except (ValueError, TypeError) as e:
            messages.error(request, f"Invalid input: {e}")
            return redirect(f"{request.path}?ippis={ippis}")

        loan_request = LoanRequest.objects.filter(
            id=request_id, member__member=selected_user
        ).first()

        if not loan_request:
            messages.error(request, "Selected loan request not found.")
            return redirect(f"{request.path}?ippis={ippis}")


        total_paid = LoanRepayback.objects.filter(
            loan_request=loan_request
        ).aggregate(total=Sum('amount_paid'))['total'] or Decimal("0.00")

        approved_amount = loan_request.approved_amount or Decimal("0.00")
        remaining_balance = approved_amount - total_paid

        if amount_paid > remaining_balance:
            messages.error(request, "Payment exceeds remaining balance.")
            return redirect(f"{request.path}?ippis={ippis}")

        # After creating payment
        total_after_payment = total_paid + amount_paid
        if total_after_payment >= approved_amount:
            loan_request.status = 'Fullpaid'
            loan_request.save(update_fields=['status']) 

        # Create payment transaction
        with transaction.atomic():
            LoanRepayback.objects.create(
                loan_request=loan_request,amount_paid=amount_paid,
                repayment_date=month_date,comment=comment,loan_receipt=loan_receipt,
                balance_remaining=remaining_balance - amount_paid,
                created_by=request.user)

            # Update status if fully paid
            total_after_payment = total_paid + amount_paid
            if total_after_payment >= loan_request.amount:
                loan_request.status = 'Fullpaid'
                loan_request.save(update_fields=['status'])

        messages.success(request,f"Payment of ₦{amount_paid:,.2f} recorded for {selected_user.first_name} ({ippis}).")
        return redirect(f"{request.path}?ippis={ippis}")
    context = {"requests": requests_list,"selected_user": selected_user,"total_paid": total_paid,"remaining_balance": remaining_balance,}
    return render(request, "loan/add_payment.html",context)


@login_required
@group_required(['admin','staff'])
def get_loan_types_for_year(request):
    year = request.GET.get("year")
    if not year:
        return JsonResponse({"error": "Year not provided"}, status=400)

    loan_types = LoanRequest.objects.filter(
        application_date__year=year
    ).values_list("loan_type__name", flat=True).distinct().order_by("loan_type__name")

    return JsonResponse({"loan_types": list(loan_types)})


@login_required
@group_required(['admin'])
def upload_loan_repayment(request):
    # 1 — Group by LoanType
    available_loans = LoanRequest.objects.filter(status="approved").select_related(
        "member", "loan_type"
    )

    grouped_by_type = defaultdict(list)
    for req in available_loans:
        if req.balance > 0:  # Use the model method
            grouped_by_type[req.loan_type].append(req)

    grouped_list = sorted(grouped_by_type.items(), key=lambda x: x[0].name)

    # 2 — Handle upload
    if request.method == "POST":
        selected_type_id = request.POST.get("selected_type")
        repayment_date_str = request.POST.get("repayment_date")
        file = request.FILES.get("excel_file")

        if not selected_type_id or not repayment_date_str or not file:
            messages.error(request, "All fields are required.")
            return redirect("upload_loan_payment")

        try:
            selected_type = LoanType.objects.get(id=selected_type_id)
        except LoanType.DoesNotExist:
            messages.error(request, "Invalid loan type.")
            return redirect("upload_loan_payment")

        try:
            repayment_date = datetime.strptime(repayment_date_str, "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "Invalid repayment date format.")
            return redirect("upload_loan_payment")

        try:
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f"Error reading Excel file: {e}")
            return redirect("upload_loan_payment")

        required_cols = {"IPPIS", "Amount Paid"}
        if not required_cols.issubset(df.columns):
            messages.error(request, "Excel must contain 'IPPIS' and 'Amount Paid' columns.")
            return redirect("upload_loan_payment")
        
        # Map IPPIS to requests for the selected type
        type_requests = grouped_by_type.get(selected_type, [])
        ippis_map = {
            str(req.member.ippis).strip(): req
            for req in type_requests
            if req.member and req.member.ippis
        }

        repaybacks_to_create = []
        skipped = []
        uploaded = 0

        with transaction.atomic():
            for _, row in df.iterrows():
                # ✅ Force IPPIS to string and remove ".0" if Excel made it float
                ippis_raw = row["IPPIS"]
                ippis = str(int(ippis_raw)) if isinstance(ippis_raw, (int, float)) else str(ippis_raw).strip()

                try:
                    amount = Decimal(str(row["Amount Paid"])).quantize(Decimal("0.00"))
                except Exception:
                    skipped.append(f"{ippis} (invalid amount)")
                    continue

                req = ippis_map.get(ippis)
                if not req:
                    skipped.append(str(ippis))
                    continue

                # Skip duplicates
                if LoanRepayback.objects.filter(
                    loan_request=req,
                    repayment_date=repayment_date
                ).exists():
                    skipped.append(f"{ippis} (duplicate entry)")
                    continue

                # Calculate balance_remaining before bulk_create
                total_repaid_so_far = req.repaybacks.aggregate(total=Sum("amount_paid"))["total"] or Decimal(0)
                balance = req.approved_amount - (total_repaid_so_far + amount)

                repaybacks_to_create.append(
                    LoanRepayback(
                        loan_request=req,
                        amount_paid=amount,
                        repayment_date=repayment_date,
                        balance_remaining=balance,
                        created_by=request.user
                    )
                )
                uploaded += 1

            if repaybacks_to_create:
                LoanRepayback.objects.bulk_create(repaybacks_to_create)

                # ✅ Update statuses after creating repayments
                request_ids = {repay.loan_request_id for repay in repaybacks_to_create}
                for req in LoanRequest.objects.filter(id__in=request_ids):
                    if req.balance <= 0:
                        req.status = 'Fullpaid'
                        req.save()

            # ✅ Fix join issue (everything is string now)
            messages.success(request, f"{uploaded} payment(s) uploaded successfully.")
            if skipped:
                messages.warning(request, f"Skipped IPPIS: {', '.join(map(str, skipped))}")

            return redirect("upload_loan_payment")

    context = {"grouped_list": grouped_list}
    return render(request, "loan/upload_loan_payment.html", context)

@login_required
@group_required(['admin','staff'])
def admin_repayment_tracking(request):
    """Track all loan repayments"""
    repayments_list = LoanRepayback.objects.select_related(
        'loan_request__member', 'loan_request__loan_type'
    ).order_by('-repayment_date')
    
    # Filtering
    member_filter = request.GET.get('member')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if member_filter:
        repayments_list = repayments_list.filter(loan_request__member_id=member_filter)
    
    if date_from:
        repayments_list = repayments_list.filter(repayment_date__gte=date_from)
    
    if date_to:
        repayments_list = repayments_list.filter(repayment_date__lte=date_to)
    
    # Pagination
    paginator = Paginator(repayments_list, 100)
    page_number = request.GET.get('page')
    repayments = paginator.get_page(page_number)
    
    # Summary statistics
    total_repaid = repayments_list.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
    outstanding_loans = LoanRequest.objects.filter(status='approved').count()
    
    context = { 'repayments': repayments, 'total_repaid': total_repaid,'outstanding_loans': outstanding_loans,
        'members': Member.objects.all(),
        'current_member': member_filter, 'date_from': date_from,'date_to': date_to,
    }
    return render(request, 'loan/repayment_tracking.html', context)


#================= test ============================


# views.py - Complete implementation for your loan analytics
# --- HELPERS ---

def monthly_paybacks_summary():
    """Get monthly payback totals"""
    return (
        LoanRepayback.objects
        .annotate(month=TruncMonth("repayment_date"))
        .values("month")
        .annotate(
            total_payments=Sum("amount_paid"),
            number_of_payments=Count("id"),
            average_payment=Avg("amount_paid")  # ✅ more efficient than Sum/Count
        )
        .order_by("month")
    )


def total_payments_by_loan_type():
    """Get total payments by each loan type"""
    return (
        LoanRepayback.objects
        .values("loan_request__loan_type__name")
        .annotate(total_amount=Sum("amount_paid"))
        .order_by("-total_amount")
    )


def monthly_payments_by_loan_type():
    """Get monthly breakdown AND loan type totals combined"""
    return (
        LoanRepayback.objects
        .annotate(month=TruncMonth("repayment_date"))
        .values("month", "loan_request__loan_type__name")
        .annotate(
            total_amount=Sum("amount_paid"),
            payment_count=Count("id"),
            average_payment=Avg("amount_paid")
        )
        .order_by("month", "loan_request__loan_type__name")
    )


# --- MAIN VIEW FUNCTION ---
@login_required
@group_required(['admin'])
def loan_analytics_view(request):
    # Get the three main datasets
    monthly_payments = monthly_paybacks_summary()
    loan_type_totals = total_payments_by_loan_type()
    detailed_breakdown = monthly_payments_by_loan_type()

    # Pagination for detailed breakdown
    paginator = Paginator(detailed_breakdown, 10)  # 10 items per page
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)

    # Calculate summary statistics
    all_payments = LoanRepayback.objects.aggregate(
        total=Sum("amount_paid"),
        count=Count("id")
    )

    # Get current month total
    current_month = timezone.now().date().replace(day=1)
    current_month_payments = LoanRepayback.objects.filter(
        repayment_date__gte=current_month
    ).aggregate(total=Sum("amount_paid"))

    # Calculate percentages for loan types (for progress bars)
    total_all = all_payments["total"] or 0
    loan_type_list = []
    for item in loan_type_totals:
        percentage = (item["total_amount"] / total_all * 100) if total_all > 0 else 0
        loan_type_list.append({
            "loan_request__loan_type__name": item["loan_request__loan_type__name"],
            "total_amount": item["total_amount"],
            "percentage": percentage
        })

    # Prepare context for template
    context = {
        # Main data for charts/tables
        "monthly_payments": monthly_payments,
        "loan_type_totals": loan_type_list,  # Modified with percentages
        "detailed_breakdown": page_obj,      # Paginated data

        # Summary statistics for cards
        "total_all_payments": all_payments["total"] or 0,
        "total_transactions": all_payments["count"] or 0,
        "current_month_total": current_month_payments["total"] or 0,

        # Additional data
        "today": timezone.now().date(),

        # Pagination data
        "page_obj": page_obj,
        "is_paginated": page_obj.has_other_pages(),
    }

    return render(request, "loan/loan_analytics.html", context)



def export_loan_schedule(request, loan_id):
    loan = LoanRequest.objects.get(id=loan_id)

    amount = Decimal(loan.amount)
    months = loan.loan_term_months

    # Calculate base monthly
    monthly = (amount / months).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    payments = [monthly] * months

    # Fix last payment
    total = sum(payments)
    diff = amount - total
    payments[-1] += diff

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Schedule"

    # Header
    ws.append([
        "Month",
        "Payment Amount (₦)",
        "Cumulative Paid (₦)",
        "Balance Remaining (₦)"
    ])

    balance = amount
    cumulative = Decimal('0.00')

    for i, payment in enumerate(payments, start=1):
        cumulative += payment
        balance -= payment

        ws.append([
            i,
            float(payment),
            float(cumulative),
            float(balance)
        ])

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename=loan_{loan.id}_schedule.xlsx'

    wb.save(response)
    return response

