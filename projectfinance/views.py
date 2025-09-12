from django.shortcuts import render
from multiprocessing.sharedctypes import Value
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator,PageNotAnInteger,EmptyPage
from django.forms import DecimalField
from django.http import JsonResponse, HttpResponse
from django.db.models import F, Q, Sum,Prefetch, DecimalField, Value
from django.db.models.functions import Coalesce
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime
from django.contrib import messages
from django.db import transaction
from decimal import Decimal, InvalidOperation
from django.db.models import Count
from collections import defaultdict
from django.utils import timezone
import requests
from loan.models import *
from .models import *
from .forms import *
from accounts.models import *
from accounts.models import *
from accounts.models import *
from main.models import *

from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Sum, F, Q, DecimalField, Count
from django.db.models.functions import Coalesce
from decimal import Decimal
from datetime import datetime, date




# =============ProjectFinanceApplication===================
@login_required
def application_list_view(request):
    applications = ProjectFinanceApplication.objects.select_related('member__member').order_by('-created_at')
    context = {'applications': applications}
    return render(request, 'projectfinance/project_finance_application.html', context)


@login_required
def application_detail_view(request, application_id):
    application = get_object_or_404(ProjectFinanceApplication, id=application_id)

    if request.method == 'POST':
        action = request.POST.get('action')
        comments = request.POST.get('comments', '').strip()
        # Handle 'review' or 'approve' action
        if action == 'review_application':
            if application.status not in ['Reviewed', 'Rejected']:
                application.status = 'Reviewed'
                application.save()
                messages.success(request, "Application reviewed and approved successfully.")
            else:
                messages.info(request, "This application has already been reviewed or rejected.")
        
       # Handle 'reject' action
        elif action == 'reject_application':
            if application.status not in ['Reviewed', 'Rejected']:
                # Check if comments are provided for rejection
                if comments:
                    application.status = 'Rejected'
                    new_comment = f"[{request.user.username} - {timezone.now().strftime('%Y-%m-%d %H:%M')}] - REJECTED: {comments}"
                    if application.comments:
                        application.comments += f"\n\n{new_comment}"
                    else:
                        application.comments = new_comment
                    
                    application.save()
                    messages.success(request, "Application has been rejected.")
                else:
                    # Show a warning if no comment is provided
                    messages.warning(request, "Please provide a reason for rejection in the comments box.")
            else:
                messages.info(request, "This application has already been reviewed or rejected.")
        # Handle 'add_comment' action (independent of approval/rejection)
        elif action == 'add_comment':
            if comments:
                new_comment = f"[{request.user.username} - {timezone.now().strftime('%Y-%m-%d %H:%M')}]: {comments}"
                if application.comments:
                    application.comments += f"\n\n{new_comment}"
                else:
                    application.comments = new_comment
                
                application.save()
                messages.success(request, "Comment added successfully.")
            else:
                messages.warning(request, "Please provide a comment.")

        return redirect('application_detail', application_id=application_id)

    context = {'application': application}
    return render(request, 'projectfinance/application_detail.html', context)
   

@login_required
def admin_project_finance_requests_list(request):
    status_filter = request.GET.get('status', '')
    requests = ProjectFinanceRequest.objects.select_related('application__member__member' ).order_by('-created_at')
    # Apply filter only if a status is selected
    if status_filter:
        requests = requests.filter(status=status_filter)
    context = {'requests': requests,'selected_status': status_filter,}
    return render(request, 'projectfinance/project_finance_list_requests.html', context)

@login_required
def admin_approve_finance_request(request, id):
    finance_request = get_object_or_404(ProjectFinanceRequest, id=id)
    
    # Check if guarantor has approved
    if finance_request.guarantor_status != 'Approved':
        messages.error(request, "Cannot approve — guarantor has not approved this request.")
        return redirect('admin_project_finance_requests')

    if request.method == 'POST':
        markup_rate_str = request.POST.get('markup_rate', '0')
        try:
            markup_rate = Decimal(markup_rate_str)
        except (TypeError, ValueError, InvalidOperation):
            messages.error(request, "Invalid markup rate provided. Please enter a number.")
            return redirect('project_finance_approved', id=id)

        # Update the request with the new markup rate
        finance_request.markup_rate = markup_rate
        finance_request.status = 'Approved'
        finance_request.approved_by = request.user
        finance_request.save()

        messages.success(request, "Request approved with markup successfully.")
        return redirect('admin_project_finance_requests')

    return render(request, 'projectfinance/project_finance_approve.html', {'finance_request': finance_request})





# ==========================================
# MAIN REPORT GENERATION FUNCTION
# ==========================================
def generate_project_finance_report(start_date=None, end_date=None):
    # Base queryset with optional date filtering
    base_filter = Q()
    if start_date:
        base_filter &= Q(created_at__gte=start_date)
    if end_date:
        base_filter &= Q(created_at__lte=end_date)
    
    # 1. EXPENDITURE (Total amount disbursed to members)
    expenditure_data = ProjectFinanceRequest.objects.filter(
        base_filter,
        status__in=['Reviewed', 'Approved', 'FullyPaid']  # Only approved/disbursed funds
    ).aggregate(
        total_expenditure=Coalesce(Sum('requested_amount'), Decimal('0.00')),
        total_requests=Count('id')
    )
    
    # 2. INCOME (Total payments received from members)
    income_data = ProjectFinancePayment.objects.filter(
        request__created_at__gte=start_date if start_date else datetime.min,
        request__created_at__lte=end_date if end_date else datetime.now()
    ).aggregate(
        total_income=Coalesce(Sum('amount_paid'), Decimal('0.00')),
        total_payments=Count('id')
    )
    
    # 3. EXPECTED TOTAL INCOME (What we should receive in total)
    expected_income_data = ProjectFinanceRequest.objects.filter(
        base_filter,
        status__in=['Reviewed', 'Approved', 'FullyPaid'],
        total_repayment_amount__isnull=False
    ).aggregate(
        expected_total_income=Coalesce(Sum('total_repayment_amount'), Decimal('0.00'))
    )
    
    # 4. PROFIT ANALYSIS
    total_expenditure = expenditure_data['total_expenditure']
    total_income = income_data['total_income']
    expected_total_income = expected_income_data['expected_total_income']
    
    current_profit = total_income - total_expenditure
    expected_profit = expected_total_income - total_expenditure
    outstanding_amount = expected_total_income - total_income
    
    # 5. INDIVIDUAL MEMBER REQUESTS
    member_requests_data = []
    
    # Get all individual requests with member details
    requests_with_members = ProjectFinanceRequest.objects.filter(
        base_filter,
        status__in=['Reviewed', 'Approved', 'FullyPaid']
    ).select_related(
        'application__member__member'
    ).order_by(
        'application__member__member__first_name',
        'application__member__member__last_name',
        'created_at'
    )
    
    for request_obj in requests_with_members:
        member = request_obj.application.member.member
        member_name = f"{member.first_name} {member.last_name}"
        
        # Calculate income received for this specific request
        request_income = ProjectFinancePayment.objects.filter(
            request=request_obj
        ).aggregate(
            total=Coalesce(Sum('amount_paid'), Decimal('0.00'))
        )['total']
        
        # Calculate expected income for this request
        expected_income = request_obj.total_repayment_amount or Decimal('0.00')
        
        # Calculate profits for this request
        request_current_profit = request_income - request_obj.requested_amount
        request_expected_profit = expected_income - request_obj.requested_amount
        request_outstanding = expected_income - request_income
        
        # Get payment count for this request
        payment_count = ProjectFinancePayment.objects.filter(request=request_obj).count()
        
        # Calculate profit margin for this request
        profit_margin_current = (request_current_profit / request_obj.requested_amount * 100) if request_obj.requested_amount > 0 else 0
        profit_margin_expected = (request_expected_profit / request_obj.requested_amount * 100) if request_obj.requested_amount > 0 else 0
        
        member_requests_data.append({
            'request_id': request_obj.id,
            'member_id': request_obj.application.member.id,
            'member_name': member_name,
            'request_date': request_obj.created_at,
            'status': request_obj.status,
            'expenditure': request_obj.requested_amount,
            'income_received': request_income,
            'expected_income': expected_income,
            'current_profit': request_current_profit,
            'expected_profit': request_expected_profit,
            'outstanding_amount': request_outstanding,
            'profit_margin_current': profit_margin_current,
            'profit_margin_expected': profit_margin_expected,
            'payment_count': payment_count,
            'is_fully_paid': request_obj.status == 'FullyPaid',
            # Additional request details
            'interest_rate': getattr(request_obj, 'interest_rate', None),
            'duration_months': getattr(request_obj, 'duration_months', None),
            'application_id': request_obj.application.id
        })
    
    # 6. MEMBER SUMMARY (Aggregated view by member)
    member_summary = {}
    for request_data in member_requests_data:
        member_id = request_data['member_id']
        member_name = request_data['member_name']
        
        if member_id not in member_summary:
            member_summary[member_id] = {
                'member_id': member_id,
                'member_name': member_name,
                'total_expenditure': Decimal('0.00'),
                'total_income_received': Decimal('0.00'),
                'total_expected_income': Decimal('0.00'),
                'total_current_profit': Decimal('0.00'),
                'total_expected_profit': Decimal('0.00'),
                'total_outstanding': Decimal('0.00'),
                'request_count': 0,
                'active_requests': 0,
                'completed_requests': 0,
                'total_payments': 0
            }
        
        # Aggregate the data
        summary = member_summary[member_id]
        summary['total_expenditure'] += request_data['expenditure']
        summary['total_income_received'] += request_data['income_received']
        summary['total_expected_income'] += request_data['expected_income']
        summary['total_current_profit'] += request_data['current_profit']
        summary['total_expected_profit'] += request_data['expected_profit']
        summary['total_outstanding'] += request_data['outstanding_amount']
        summary['request_count'] += 1
        summary['total_payments'] += request_data['payment_count']
        
        if request_data['is_fully_paid']:
            summary['completed_requests'] += 1
        else:
            summary['active_requests'] += 1
    
    # Convert member_summary to list and sort by expected profit
    member_summary_list = list(member_summary.values())
    member_summary_list.sort(key=lambda x: x['total_expected_profit'], reverse=True)
    
    # 7. SUMMARY STATISTICS
    total_active_requests = ProjectFinanceRequest.objects.filter(
        base_filter,
        status__in=['Reviewed', 'Approved']
    ).count()
    
    total_completed_requests = ProjectFinanceRequest.objects.filter(
        base_filter,
        status='FullyPaid'
    ).count()
    
    # Compile final report
    report = {
        'summary': {
            'total_expenditure': total_expenditure,
            'total_income': total_income,
            'expected_total_income': expected_total_income,
            'current_profit': current_profit,
            'expected_profit': expected_profit,
            'outstanding_amount': outstanding_amount,
            'profit_margin_current': (current_profit / total_expenditure * 100) if total_expenditure > 0 else 0,
            'profit_margin_expected': (expected_profit / total_expenditure * 100) if total_expenditure > 0 else 0,
        },
        'statistics': {
            'total_requests': expenditure_data['total_requests'],
            'total_payments_made': income_data['total_payments'],
            'active_requests': total_active_requests,
            'completed_requests': total_completed_requests,
            'average_request_amount': total_expenditure / expenditure_data['total_requests'] if expenditure_data['total_requests'] > 0 else 0,
            'unique_members': len(member_summary_list)
        },
        'individual_requests': member_requests_data,  # Each request listed separately
        'member_summary': member_summary_list,       # Aggregated by member
        'generated_at': timezone.now(),
        'date_range': {
            'start_date': start_date,
            'end_date': end_date
        }
    }
    
    return report


# ==========================================
# DJANGO VIEWS
# ==========================================

@staff_member_required  # Only allow staff/admin users
def project_finance_report_view(request):
    """
    Renders the project finance report page with data based on selected filters.
    """
    context = {}
    
    # Get date parameters from request
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    view_type = request.GET.get('view', 'individual')  # 'individual' or 'summary'

    start_date = None
    end_date = None
    
    # Parse dates if provided
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            context['error'] = "Invalid start date format. Please use YYYY-MM-DD."
    
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            context['error'] = "Invalid end date format. Please use YYYY-MM-DD."

    # Always generate the report on a GET request
    try:
        report = generate_project_finance_report(start_date, end_date)
        context['report'] = report
        context['view_type'] = view_type
        context['success'] = True
    except Exception as e:
        context['error'] = f"Error generating report: {str(e)}"
    
    return render(request, 'projectfinance/project_finance_report.html', context)

@staff_member_required
def project_finance_report_api(request):
    """
    API endpoint to get report data as JSON
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    view_type = request.GET.get('view', 'individual')
    
    # Parse dates if provided
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Invalid start date format'}, status=400)
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Invalid end date format'}, status=400)
    
    try:
        report = generate_project_finance_report(start_date, end_date)
        
        # Convert Decimal objects to float for JSON serialization
        def decimal_to_float(obj):
            if isinstance(obj, Decimal):
                return float(obj)
            elif isinstance(obj, dict):
                return {k: decimal_to_float(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [decimal_to_float(v) for v in obj]
            return obj
        
        report = decimal_to_float(report)
        report['generated_at'] = report['generated_at'].isoformat()
        
        # Return only the requested view type if specified
        if view_type == 'individual':
            response_data = {
                'summary': report['summary'],
                'statistics': report['statistics'],
                'data': report['individual_requests'],
                'view_type': 'individual',
                'generated_at': report['generated_at'],
                'date_range': report['date_range']
            }
        elif view_type == 'summary':
            response_data = {
                'summary': report['summary'],
                'statistics': report['statistics'],
                'data': report['member_summary'],
                'view_type': 'summary',
                'generated_at': report['generated_at'],
                'date_range': report['date_range']
            }
        else:
            response_data = report  # Return full report
        
        return JsonResponse(response_data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ==========================================
# OPTIONAL: Helper function for template use
# ==========================================

def get_member_request_details(member_id, start_date=None, end_date=None):
    """
    Helper function to get detailed request information for a specific member
    """
    base_filter = Q(application__member__id=member_id)
    if start_date:
        base_filter &= Q(created_at__gte=start_date)
    if end_date:
        base_filter &= Q(created_at__lte=end_date)
    
    requests = ProjectFinanceRequest.objects.filter(
        base_filter,
        status__in=['Reviewed', 'Approved', 'FullyPaid']
    ).select_related(
        'application__member__member'
    ).prefetch_related(
        'projectfinancepayment_set'
    ).order_by('created_at')
    
    request_details = []
    for request_obj in requests:
        # Get all payments for this request
        payments = request_obj.projectfinancepayment_set.all()
        total_paid = sum(payment.amount_paid for payment in payments)
        
        request_details.append({
            'request': request_obj,
            'payments': payments,
            'total_paid': total_paid,
            'outstanding': (request_obj.total_repayment_amount or Decimal('0.00')) - total_paid,
            'profit_current': total_paid - request_obj.requested_amount,
            'profit_expected': (request_obj.total_repayment_amount or Decimal('0.00')) - request_obj.requested_amount
        })
    
    return request_details
#=======================================


def project_finance_report_excel(request):
    """
    Export ONLY 'Section 2 — Members’ Breakdown' to Excel.
    Columns: Member | Requests Amount | Income Received | Expected Income | Expected Profit | Outstanding | Status
    """
    # Query: each request with its member (optimized)
    requests_qs = (
        ProjectFinanceRequest.objects
        .select_related('application__member__member')
        .order_by(
            'application__member__member__first_name',
            'application__member__member__last_name',
            'created_at'
        )
    )

    # Prepare workbook/sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members Breakdown"

    # Header row
    headers = [
        "Member",
        "Requests Amount",
        "Income Received",
        "Expected Income",
        "Expected Profit",
        "Outstanding",
        "Status",
    ]
    ws.append(headers)
    for col_idx, header in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

    CURRENCY_FMT = '₦#,##0.00'

    # Populate rows
    for req in requests_qs:
        user = req.application.member.member  # follows your existing relation chain
        member_name = f"{getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')}".strip()

        requested_amount = req.requested_amount or Decimal('0.00')

        income_received = (
            ProjectFinancePayment.objects
            .filter(request=req)
            .aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
        )

        expected_income = req.total_repayment_amount if req.total_repayment_amount is not None else requested_amount
        expected_profit = expected_income - requested_amount
        profit_pct = (expected_profit / requested_amount * Decimal('100')) if requested_amount > 0 else Decimal('0')
        outstanding = expected_income - income_received

        # Normalize status (show "Fully Paid" if cleared, else use model status)
        status = 'Fully Paid' if outstanding <= 0 else (req.status or '')
        if status == 'FullyPaid':
            status = 'Fully Paid'

        # Append row (note: 'Expected Profit' includes % as text like "40000.0 (10.0%)")
        ws.append([
            member_name,
            float(requested_amount),
            float(income_received),
            float(expected_income),
            f"{float(expected_profit)} ({round(profit_pct, 1)}%)",
            float(outstanding),
            status,
        ])

        # Apply currency format to money columns on the just-added row
        last_row = ws.max_row
        for col_letter in ['B', 'C', 'D', 'F']:
            ws[f"{col_letter}{last_row}"].number_format = CURRENCY_FMT

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max_len + 3

    # Return file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="members_breakdown.xlsx"'
    wb.save(response)
    return response


@login_required
def upload_project_finance_repayment(request):
    if request.method == "POST":
        file = request.FILES.get("file")

        if not file:
            messages.error(request, "Please upload an Excel file.")
            return redirect('upload_project_finance_payment')

        try:
            # Read Excel file
            df = pd.read_excel(file)

            # Normalize column names
            df_columns = {col.strip().lower(): col.strip() for col in df.columns}
            required_columns_lower = ["ippis", "amount paid", "month"]

            # Validate required columns
            for col in required_columns_lower:
                if col not in df_columns:
                    found_columns = ", ".join(df.columns)
                    messages.error(request, f"Missing required column: '{col.title()}'. Found columns: {found_columns}")
                    return redirect('upload_project_finance_payment')

            # Get column names
            ippis_col = df_columns["ippis"]
            amount_paid_col = df_columns["amount paid"]
            month_col = df_columns["month"]

            successful_uploads = 0
            skipped_payments = []

            with transaction.atomic():
                for index, row in df.iterrows():
                    row_number = index + 1

                    # --- Validate IPPIS ---
                    ippis = str(row[ippis_col]).strip()
                    if not ippis or ippis.lower() in ['nan', 'none', '']:
                        skipped_payments.append({
                            "row": row_number,
                            "ippis": ippis or "Empty",
                            "reason": "IPPIS is empty or invalid"
                        })
                        continue

                    # --- Validate Amount Paid ---
                    try:
                        amount_str = str(row[amount_paid_col]).strip().replace('₦', '').replace(',', '').replace(' ', '')
                        amount_paid = Decimal(amount_str)

                        if amount_paid <= 0:
                            skipped_payments.append({
                                "row": row_number,
                                "ippis": ippis,
                                "reason": f"Invalid amount: {amount_paid}. Must be greater than 0"
                            })
                            continue
                    except (ValueError, InvalidOperation):
                        skipped_payments.append({
                            "row": row_number,
                            "ippis": ippis,
                            "reason": f"Invalid amount format: '{row[amount_paid_col]}'"
                        })
                        continue

                    # --- Validate Month ---
                    try:
                        month = pd.to_datetime(row[month_col]).date()
                    except (ValueError, TypeError):
                        skipped_payments.append({
                            "row": row_number,
                            "ippis": ippis,
                            "reason": f"Invalid date format: '{row[month_col]}'"
                        })
                        continue

                    # --- Find Active Finance Request ---
                    try:
                       request_obj = ProjectFinanceRequest.objects.select_related("application__member").annotate(
                        total_paid=Coalesce(
                            Sum('payments__amount_paid'),   # ✅ FIXED HERE
                            Decimal('0')
                        ),
                        remaining_balance=F('total_repayment_amount') - F('total_paid')
                    ).get(
                        application__member__ippis=ippis,
                        status__in=["Approved", "Pending"]
                    )

                    except ProjectFinanceRequest.DoesNotExist:
                        all_requests = ProjectFinanceRequest.objects.filter(application__member__ippis=ippis)
                        if all_requests.exists():
                            statuses = list(all_requests.values_list('status', flat=True))
                            skipped_payments.append({
                                "row": row_number,
                                "ippis": ippis,
                                "reason": f"No active request found. Existing statuses: {', '.join(statuses)}"
                            })
                        else:
                            skipped_payments.append({
                                "row": row_number,
                                "ippis": ippis,
                                "reason": "No finance request found for this IPPIS"
                            })
                        continue
                    except ProjectFinanceRequest.MultipleObjectsReturned:
                        skipped_payments.append({
                            "row": row_number,
                            "ippis": ippis,
                            "reason": "Multiple active requests found for this IPPIS"
                        })
                        continue

                    # --- Check for Existing Payment ---
                    if ProjectFinancePayment.objects.filter(
                        request=request_obj,
                        month__year=month.year,
                        month__month=month.month
                    ).exists():
                        skipped_payments.append({
                            "row": row_number,
                            "ippis": ippis,
                            "reason": f"Payment already exists for {month.strftime('%B %Y')}"
                        })
                        continue

                    # --- Check Fully Paid Status ---
                    if request_obj.remaining_balance <= 0:
                        skipped_payments.append({
                            "row": row_number,
                            "ippis": ippis,
                            "reason": f"Request fully paid (Total: ₦{request_obj.total_repayment_amount:,.2f}, Paid: ₦{request_obj.total_paid:,.2f})"
                        })
                        continue

                    # --- Check Overpayment ---
                    if amount_paid > request_obj.remaining_balance:
                        skipped_payments.append({
                            "row": row_number,
                            "ippis": ippis,
                            "reason": f"Payment ₦{amount_paid:,.2f} exceeds remaining balance ₦{request_obj.remaining_balance:,.2f}"
                        })
                        continue

                    # --- Save Payment ---
                    try:
                        ProjectFinancePayment.objects.create(
                            request=request_obj,
                            amount_paid=amount_paid,
                            month=month,
                        )

                        # Update request status if fully paid
                        new_total_paid = request_obj.total_paid + amount_paid
                        if new_total_paid >= request_obj.total_repayment_amount:
                            request_obj.status = "FullyPaid"
                            request_obj.save()

                        successful_uploads += 1

                    except Exception as save_error:
                        skipped_payments.append({
                            "row": row_number,
                            "ippis": ippis,
                            "reason": f"Database error: {str(save_error)}"
                        })
                        continue

                # --- Store Results in Session ---
                request.session['upload_results'] = {
                    'successful_uploads': successful_uploads,
                    'skipped_payments': skipped_payments,
                    'total_rows': len(df)
                }

                # --- Messages ---
                if successful_uploads > 0:
                    messages.success(request, f"Successfully uploaded {successful_uploads} payments!")
                if skipped_payments:
                    messages.warning(request, f"{len(skipped_payments)} payments were skipped.")
                if successful_uploads == 0 and skipped_payments:
                    messages.error(request, "No payments were uploaded. All rows had errors.")

        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")

        return redirect('upload_project_finance_payment')

    # --- GET Request ---
    upload_results = request.session.pop('upload_results', None)
    context = {'skipped_payments': upload_results['skipped_payments'] if upload_results else []}
    return render(request, "projectfinance/upload_project_finance_payment.html", context)

@login_required
def make_finance_payment(request, id):
    finance_request = get_object_or_404(ProjectFinanceRequest, id=id)

    if request.method == "POST":
        form = ProjectFinancePaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.request = finance_request
            payment.recorded_by = request.user

            # Calculate remaining balance
            total_paid = finance_request.payments.aggregate(
                total=Sum('amount_paid')
            )['total'] or Decimal('0.00')

            total_price = finance_request.total_repayment_amount or finance_request.requested_amount
            balance_remaining = total_price - total_paid

            # Block payment if balance is already cleared
            if balance_remaining <= 0:
                messages.error(request, "This request has been fully paid. No further payments are allowed.")
                return redirect("admin_project_finance_requests_list", id=finance_request.id)

            # Block duplicate payments for the same month
            existing_payment = finance_request.payments.filter(
                month__year=payment.month.year,
                month__month=payment.month.month
            ).exists()

            if existing_payment:
                messages.error(
                    request,f"A payment for {payment.month.strftime('%B %Y')} already exists.")
                return redirect("admin_project_finance_requests_list", id=finance_request.id)

            # Update remaining balance after this payment
            new_total_paid = total_paid + payment.amount_paid
            payment.balance_remaining = total_price - new_total_paid

            # Save payment
            payment.save()

            # Update finance request balance & status
            finance_request.update_balance_remaining()

            messages.success(
                request,f"Payment of ₦{payment.amount_paid} for {payment.month.strftime('%B %Y')} recorded successfully!")
            return redirect("make_finance_payment", id=finance_request.id)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = ProjectFinancePaymentForm()

    # Fetch all previous payments
    payments = finance_request.payments.all().order_by("-created_at")
    context = { "finance_request": finance_request, "form": form, "payments": payments,}
    return render(request, "projectfinance/make_payment.html", context)

# @login_required
# def make_finance_payment(request, id):
#     finance_request = get_object_or_404(ProjectFinanceRequest, id=id)
#     if request.method == "POST":
#         form = ProjectFinancePaymentForm(request.POST)
#         if form.is_valid():
#             payment = form.save(commit=False)
#             payment.request = finance_request
#             payment.recorded_by = request.user

#             # Calculate remaining balance
#             total_paid = finance_request.payments.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
#             new_total_paid = total_paid + payment.amount_paid
#             total_price = finance_request.total_repayment_amount or finance_request.requested_amount
#             payment.balance_remaining = total_price - new_total_paid

#             # Save payment
#             payment.save()

#             # Update finance request balance & status
#             finance_request.update_balance_remaining()

#             messages.success(request, f"Payment of ₦{payment.amount_paid} recorded successfully!")
#             return redirect("admin_project_finance_requests_list", id=finance_request.id)
#         else:
#             messages.error(request, " Please correct the errors below.")
#     else:
#         form = ProjectFinancePaymentForm()

#     # Fetch all previous payments
#     payments = finance_request.payments.all().order_by("-created_at")

#     context = {
#         "finance_request": finance_request,
#         "form": form,
#         "payments": payments,
#     }
#     return render(request, "projectfinance/make_payment.html", context)


#============================================


