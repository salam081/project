from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from accounts.decorators import group_required
from django.contrib import messages
from django.utils import timezone
from django.db import transaction
from datetime import datetime
from datetime import date
from django.db.models import Sum, Count, Q, F
from django.utils.dateparse import parse_date
from django.core.paginator import Paginator
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl
import datetime
from decimal import Decimal, InvalidOperation
from decimal import Decimal
from django.utils.timezone import localtime
from accounts.models import *
from .models import *
from form_app.models import *
import json
from django.http import JsonResponse





# Create your views here.

@login_required
@group_required(['admin'])
def ram_dashboard(request):
    # ── Budget Stats ──────────────────────────────────────────
    total_budgets = Budget.objects.count()
    approved_budgets = Budget.objects.filter(status="approved").count()
    pending_budgets = Budget.objects.filter(status="pending").count()
    total_budget_amount = Budget.objects.filter(status="approved").aggregate(
        total=Sum("total_amount")
    )["total"] or Decimal("0.00")

    # ── Request Stats ─────────────────────────────────────────
    total_requests = RamRequest.objects.count()
    pending_requests = RamRequest.objects.filter(status="pending").count()
    approved_requests = RamRequest.objects.filter(status="approved").count()
    rejected_requests = RamRequest.objects.filter(status="rejected").count()
    fully_paid_requests = RamRequest.objects.filter(status="fully Paid").count()

    # ── Financial Stats ───────────────────────────────────────
    total_cost_spent = RamRequestDetails.objects.filter(
        request__status__in=["approved", "fully Paid"]
    ).aggregate(
        total=Sum(F("cost_price") * F("quantity"))
    )["total"] or Decimal("0.00")

    total_selling_value = RamRequestDetails.objects.filter(
        request__status__in=["approved", "fully Paid"]
    ).aggregate(
        total=Sum(F("selling_price") * F("quantity"))
    )["total"] or Decimal("0.00")

    total_collected = Payment.objects.aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0.00")

    total_outstanding = total_selling_value - total_collected

    # ── Recent Requests ───────────────────────────────────────
    recent_requests = RamRequest.objects.select_related(
        "member__member", "budget", "approved_by"
    ).order_by("-date_requested")[:10]

    # ── Recent Payments ───────────────────────────────────────
    recent_payments = Payment.objects.select_related(
        "request__member__member", "received_by"
    ).order_by("-date_paid")[:10]

    # ── Monthly payments for chart (last 6 months) ────────────
   
    today = timezone.now()
    monthly_data = []
    monthly_labels = []

    for i in range(5, -1, -1):
        month_date = today - datetime.timedelta(days=30 * i)
        # month_date = today - timedelta(days=30 * i)
        month_total = Payment.objects.filter(
            date_paid__year=month_date.year,
            date_paid__month=month_date.month,
        ).aggregate(total=Sum("amount"))["total"] or 0
        monthly_data.append(float(month_total))
        monthly_labels.append(month_date.strftime("%b %Y"))

    context = {
        # Budget
        "total_budgets": total_budgets,
        "approved_budgets": approved_budgets,
        "pending_budgets": pending_budgets,
        "total_budget_amount": total_budget_amount,

        # Requests
        "total_requests": total_requests,
        "pending_requests": pending_requests,
        "approved_requests": approved_requests,
        "rejected_requests": rejected_requests,
        "fully_paid_requests": fully_paid_requests,

        # Financial
        "total_cost_spent": total_cost_spent,
        "total_selling_value": total_selling_value,
        "total_collected": total_collected,
        "total_outstanding": total_outstanding,

        # Tables
        "recent_requests": recent_requests,
        "recent_payments": recent_payments,

        # Chart
        "monthly_labels": json.dumps(monthly_labels),
        "monthly_data": json.dumps(monthly_data),
    }

    return render(request, 'ram_app/ram_dashboard.html', context)


# ─── BUDGET VIEWS ─────────────────────────────────────────────────────────────

@login_required
@group_required(['admin'])
def budget_list(request):
    budgets = Budget.objects.select_related("created_by", "approved_by").all()
    paginator = Paginator(budgets , 1) 
    page_number = request.GET.get("page")
    budgets = paginator.get_page(page_number)
    return render(request, "ram_app/budget_list.html", {"budgets": budgets})


@login_required
@group_required(['admin'])
def toggle_budget_status(request, pk):
    budget = get_object_or_404(Budget, pk=pk)

    # if activating this budget → deactivate others
    budget.is_active = not budget.is_active

    if budget.is_active:
        Budget.objects.exclude(pk=pk).update(is_active=False)

    budget.save()

    return JsonResponse({"success": True,"is_active": budget.is_active})

@login_required
@group_required(['admin'])
def budget_create(request):
    if request.method == "POST":
        name = request.POST.get("name")
        total_amount = request.POST.get("total_amount")
        is_active = request.POST.get("is_active") == "on"

        if not name or not total_amount:
            messages.error(request, "All fields are required.")
            return render(request, "ram_app/budget_form.html")

        # convert to proper decimal type
        try:
            total_amount = Decimal(total_amount)
        except:
            messages.error(request, "Invalid amount value.")
            return render(request, "ram_app/budget_form.html")

        # only one active budget allowed
        if is_active:
            Budget.objects.filter(is_active=True).update(is_active=False)

        Budget.objects.create(name=name,total_amount=total_amount,is_active=is_active,created_by=request.user,)

        messages.success(request, "Budget created successfully.")
        return redirect("budget_list")

    return render(request, "ram_app/budget_form.html")


@login_required
@group_required(['admin'])
def budget_update(request, pk):
    budget = get_object_or_404(Budget, pk=pk)

    if request.method == "POST":
        name = request.POST.get("name")
        total_amount = request.POST.get("total_amount")
        is_active = request.POST.get("is_active") == "on"

        if not name or not total_amount:
            messages.error(request, "All fields are required.")
            return render(request, "ram_app/budget_update_form.html", {"budget": budget})

        # convert safely
        try:
            total_amount = Decimal(total_amount)
        except:
            messages.error(request, "Invalid amount value.")
            return render(request, "ram_app/budget_update_form.html", {"budget": budget})

        # ensure only one active budget
        if is_active:
            Budget.objects.exclude(pk=pk).update(is_active=False)

        Budget.objects.filter(pk=pk).update(name=name,total_amount=total_amount,is_active=is_active,)

        messages.success(request, "Budget updated successfully.")
        return redirect("budget_list")

    return render(request,"ram_app/budget_update_form.html",{"budget": budget})

@login_required
@group_required(['admin'])
def budget_detail(request, pk):
    budget = get_object_or_404(Budget.objects.select_related("created_by", "approved_by"), pk=pk)
    requests_list  = budget.requests.select_related("member__member", "guarantor__member").prefetch_related("items")
    
    paginator = Paginator(requests_list , 30) 
    page_number = request.GET.get("page")
    requests = paginator.get_page(page_number)
    return render(request, "ram_app/budget_detail.html", {"budget": budget,"requests": requests,})


@login_required
@group_required(['admin'])
def budget_approve(request, pk):
    budget = get_object_or_404(Budget, pk=pk)

    if budget.status != "pending":
        messages.warning(request, "Budget has already been processed.")
        return redirect("budget_detail", pk=pk)

    budget.status = "approved"
    budget.is_active = True
    budget.approved_by = request.user
    budget.approved_date = timezone.now()
    budget.save()

    messages.success(request, f"Budget '{budget.name}' approved successfully.")
    return redirect("budget_detail", pk=pk)


@login_required
@group_required(['admin'])
def budget_reject(request, pk):
    budget = get_object_or_404(Budget, pk=pk)

    if budget.status != "pending":
        messages.warning(request, "Budget has already been processed.")
        return redirect("budget_detail", pk=pk)

    budget.status = "rejected"
    budget.is_active = False
    budget.approved_by = request.user
    budget.approved_date = timezone.now()
    budget.save()

    messages.error(request, f"Budget '{budget.name}' rejected.")
    return redirect("budget_detail", pk=pk)


# ─── RAM REQUEST VIEWS ────────────────────────────────────────────────────────

@login_required
@group_required(['admin', 'staff'])
def ram_request_list(request):
    # ── Base queryset ─────────────────────────────────────────
    if request.user.group.title == 'admin':
        ram_requests = RamRequest.objects.select_related(
            "budget", "member__member", "guarantor__member", "approved_by", "created_by"
        ).prefetch_related("items")
    else:
        member = get_object_or_404(Member, member=request.user)
        ram_requests = RamRequest.objects.filter(member=member).select_related(
            "budget", "guarantor__member", "approved_by"
        ).prefetch_related("items")

    # ── Filters ───────────────────────────────────────────────
    status = request.GET.get("status", "").strip()
    search = request.GET.get("search", "").strip()   # member name or guest name
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()
    budget_id = request.GET.get("budget", "").strip()

    if status:
        ram_requests = ram_requests.filter(status=status)

    if search:
        ram_requests = ram_requests.filter(
            Q(member__member__first_name__icontains=search) |
            Q(member__member__last_name__icontains=search) |
            Q(guest_name__icontains=search) |
            Q(member__ippis__icontains=search) |
            Q(guest_ippis__icontains=search)
        )

    if date_from:
        ram_requests = ram_requests.filter(date_requested__date__gte=date_from)

    if date_to:
        ram_requests = ram_requests.filter(date_requested__date__lte=date_to)

    if budget_id:
        ram_requests = ram_requests.filter(budget_id=budget_id)

    # ── Excel Export ──────────────────────────────────────────
    if request.GET.get("export") == "excel":
        return export_ram_requests_excel(ram_requests)

    # ── Pagination ────────────────────────────────────────────
    paginator = Paginator(ram_requests, 2)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # ── Budgets for filter dropdown ───────────────────────────
    budgets = Budget.objects.filter(status="approved")

    context = {
        "ram_requests": page_obj,
        "page_obj": page_obj,
        "budgets": budgets,
        # preserve filter values
        "filter_status": status,
        "filter_search": search,
        "filter_date_from": date_from,
        "filter_date_to": date_to,
        "filter_budget": budget_id,
        "total_count": ram_requests.count(),
    }

    return render(request, "ram_app/ram_request_list.html", context)


def export_ram_requests_excel(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RAM Requests"

    # ── Header style ──────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A2942", end_color="1A2942", fill_type="solid")
    center = Alignment(horizontal="center")

    headers = [
        "#", "Member / Guest", "IPPIS", "Budget", "Items","Duration (Months)",
        "Cost Price (₦)", "Selling Price (₦)", "Paid (₦)",
        "Balance (₦)", "Status", "Guarantor", "Date Requested",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # ── Data rows ─────────────────────────────────────────────
    for row_num, req in enumerate(queryset, 2):
        if req.member:
            name = req.member.member.get_full_name()
            ippis = req.member.ippis
        else:
            name = req.guest_name or "—"
            ippis = req.guest_ippis or "—"

        guarantor = req.guarantor.member.get_full_name() if req.guarantor else "—"
        items_count = req.items.count()
        
        durations = ", ".join(str(item.duration_months) for item in req.items.all())
        
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=name)
        ws.cell(row=row_num, column=3, value=str(ippis))
        ws.cell(row=row_num, column=4, value=req.budget.name)
        ws.cell(row=row_num, column=5, value=items_count)
        ws.cell(row=row_num, column=6, value=durations)
        ws.cell(row=row_num, column=7, value=float(req.total_cost_price))
        ws.cell(row=row_num, column=8, value=float(req.total_selling_price))
        ws.cell(row=row_num, column=9, value=float(req.total_paid))
        ws.cell(row=row_num, column=10, value=float(req.balance_remaining))
        ws.cell(row=row_num, column=11, value=req.get_status_display())
        ws.cell(row=row_num, column=12, value=guarantor)
        ws.cell(row=row_num, column=13, value=req.date_requested.strftime("%d %b %Y"))

        # Alternate row shading
        if row_num % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = PatternFill(
                    start_color="F0F4FF", end_color="F0F4FF", fill_type="solid"
                )

    # ── Column widths ─────────────────────────────────────────
    col_widths = [5, 25, 15, 25, 8, 18, 18, 15, 15, 15, 25, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # ── Response ──────────────────────────────────────────────
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ram_requests.xlsx"'
    wb.save(response)
    return response

@login_required
@group_required(['admin', 'staff'])
def ram_request_create(request):
    budgets = Budget.objects.filter(status="approved", is_active=True,total_amount__gt=0)
    selling_plans = Markup.objects.filter(is_active=True)

    context = {"budgets": budgets, "selling_plans": selling_plans}

    if request.method == "POST":
        budget_id = request.POST.get("budget")
        request_type = request.POST.get("request_type")
        note = request.POST.get("note", "")
        file_payslip = request.FILES.get("file_payslip")

        item_names = request.POST.getlist("item_name")
        quantities = request.POST.getlist("quantity")
        cost_prices = request.POST.getlist("cost_price")
        selling_plan_ids = request.POST.getlist("selling_plan")
        duration_months = request.POST.getlist("duration_months")

        if not budget_id or not item_names:
            messages.error(request, "Budget and at least one item are required.")
            return render(request, "ram_app/ram_request_form.html", context)

        budget = get_object_or_404(Budget, pk=budget_id, status="approved", is_active=True)

        member = None
        guarantor = None
        guest_name = None
        guest_phone = None
        guest_ippis = None

        if request_type == "member":
            member_ippis = request.POST.get("member_ippis")
            # guarantor_ippis = request.POST.get("guarantor_ippis")
           

            if not member_ippis:
                messages.error(request, "Member IPPIS is required.")
                return render(request, "ram_app/ram_request_form.html", context)

            member = Member.objects.filter(ippis=member_ippis).first()
            if not member:
                messages.error(request, f"No member found with IPPIS {member_ippis}.")
                return render(request, "ram_app/ram_request_form.html", context)
            
             # ── PAYMENT CHECK ──────────────────────
            with transaction.atomic():
                ram_payment_type = PaymentType.objects.get(title="Ram form")

                payment = RequestFormPayment.objects.filter(
                    member=member,
                    payment_type=ram_payment_type,
                    status="paid"
                ).select_for_update().first()
               

                if not payment:
                    messages.error(request, "You have not paid for this request form Fee.")
                    return render(request, "ram_app/ram_request_form.html", context)

                updated = RequestFormPayment.objects.filter(
                    id=payment.id,
                    status="paid"
                ).update(status="used")

                if not updated:
                    messages.error(request, "Payment already used.")
                    return render(request, "ram_app/ram_request_form.html", context)

            
            # if guarantor_ippis:
            #     guarantor = Member.objects.filter(ippis=guarantor_ippis).first()
            #     if not guarantor:
            #         messages.error(request, f"No member found with guarantor IPPIS {guarantor_ippis}.")
            #         return render(request, "ram_app/ram_request_form.html", context)

        elif request_type == "guest":
            guest_name = request.POST.get("guest_name", "").strip()
            guest_phone = request.POST.get("guest_phone", "").strip()
            guest_ippis = request.POST.get("guest_ippis", "").strip()
            # guarantor_ippis = request.POST.get("guarantor_ippis", "").strip()
           

            if not guest_name:
                messages.error(request, "Guest name is required.")
                return render(request, "ram_app/ram_request_form.html", context)
            
            
             # ── PAYMENT CHECK ──────────────────────
            with transaction.atomic():
                ram_payment_type = PaymentType.objects.get(title="Ram form")

                payment = RequestFormPayment.objects.filter(
                    member=member,
                    payment_type=ram_payment_type,
                    status="paid"
                ).select_for_update().first()
               

                if not payment:
                    messages.error(request, "Guest has not paid for this request form Fee.")
                    return render(request, "ram_app/ram_request_form.html", context)

                updated = RequestFormPayment.objects.filter(
                    id=payment.id,
                    status="paid"
                ).update(status="used")

                if not updated:
                    messages.error(request, "Payment already used.")
                    return render(request, "ram_app/ram_request_form.html", context)

            # if guarantor_ippis:
            #     guarantor = Member.objects.filter(ippis=guarantor_ippis).first()
            #     if not guarantor:
            #         messages.error(request, f"No member found with guarantor IPPIS {guarantor_ippis}.")
            #         return render(request, "ram_app/ram_request_form.html", context)
           
        else:
            messages.error(request, "Please select Member or Guest tab.")
            return render(request, "ram_app/ram_request_form.html", context)

        ram_request = RamRequest.objects.create(
            budget=budget,
            member=member,
            guest_name=guest_name,
            guest_phone=guest_phone,
            guest_ippis=guest_ippis if guest_ippis else None,
            file_payslip=file_payslip,
            # guarantor=guarantor,
            note=note,
            created_by=request.user,
        )

        for item_name, quantity, cost_price, plan_id, duration_month in zip(
            item_names, quantities, cost_prices, selling_plan_ids, duration_months
        ):
            if item_name and quantity and cost_price:
                RamRequestDetails.objects.create(
                    request=ram_request,
                    item_name=item_name,
                    quantity=int(quantity),
                    cost_price=cost_price,
                    duration_months=int(duration_month) if duration_month and str(duration_month).strip() else 1
                )

        messages.success(request, "RAM request submitted successfully. Awaiting admin approval.")
        return redirect("ram_request_list")

    return render(request, "ram_app/ram_request_form.html", context)


@login_required
@group_required(['admin', 'staff'])
def ram_request_detail(request, pk):
    ram_request = get_object_or_404(
        RamRequest.objects.select_related(
            "budget", "member__member", "guarantor__member", "approved_by", "created_by"
        ).prefetch_related("items", "payments__received_by"),
        pk=pk
    )
    return render(request, "ram_app/ram_request_detail.html", {"ram_request": ram_request})


@login_required
@group_required(['admin'])
def ram_request_approve(request, pk):
    if not request.user.is_staff:
        messages.error(request, "You are not authorized to perform this action.")
        return redirect("ram_request_list")

    ram_request = get_object_or_404(RamRequest, pk=pk)

    if ram_request.status != "pending":
        messages.warning(request, "Request has already been processed.")
        return redirect("ram_request_detail", pk=pk)

    budget = ram_request.budget
    if ram_request.total_cost_price > budget.remaining_amount:
        messages.error(request, f"Insufficient budget balance. Remaining: ₦{budget.remaining_amount}")
        return redirect("ram_request_detail", pk=pk)

    ram_request.status = "approved"
    ram_request.approved_by = request.user
    ram_request.approved_date = timezone.now()
    ram_request.save()

    messages.success(request, "RAM request approved successfully.")
    return redirect("ram_request_list")


@login_required
@group_required(['admin'])
def ram_request_reject(request, pk):
    if not request.user.is_staff:
        messages.error(request, "You are not authorized to perform this action.")
        return redirect("ram_request_list")

    ram_request = get_object_or_404(RamRequest, pk=pk)

    if ram_request.status != "pending":
        messages.warning(request, "Request has already been processed.")
        return redirect("ram_request_detail", pk=pk)

    ram_request.status = "rejected"
    ram_request.approved_by = request.user
    ram_request.approved_date = timezone.now()
    ram_request.save()

    messages.error(request, "RAM request rejected.")
    return redirect("ram_request_detail", pk=pk)


# ─── PAYMENT VIEWS ────────────────────────────────────────────────────────────

@login_required
@group_required(['admin'])
def payment_add(request, pk):
    ram_request = get_object_or_404(RamRequest, pk=pk, status="approved")

    if ram_request.is_fully_paid:
        messages.warning(request, "This request has already been fully paid.")
        return redirect("ram_request_detail", pk=pk)

    if request.method == "POST":
        amount = request.POST.get("amount")
        repayment_month = request.POST.get("repayment_mounth")
        note = request.POST.get("note", "")
        payment_receipt = request.FILES.get("payment_receipt")

        if not amount or not repayment_month:
            messages.error(request, "Amount and repayment month are required.")
            return render(request, "ram_app/payment_form.html", {"ram_request": ram_request})

        from decimal import Decimal
        amount = Decimal(amount)

        if amount <= 0:
            messages.error(request, "Amount must be greater than zero.")
            return render(request, "ram/payment_form.html", {"ram_request": ram_request})

        if amount > ram_request.balance_remaining:
            messages.error(request, f"Amount exceeds balance remaining. Balance: ₦{ram_request.balance_remaining}")
            return render(request, "ram/payment_form.html", {"ram_request": ram_request})
        
        # convert string -> date safely
        repayment_month = parse_date(repayment_month)

        if not repayment_month:
            messages.error(request, "Invalid repayment month format.")
            return render(request, "ram_app/payment_form.html", {"ram_request": ram_request})

        # check if payment already exists for this request + month
        exists = Payment.objects.filter(request=ram_request,repayment_mounth=repayment_month).exists()

        if exists:
            messages.error(request,"A payment for this month already exists for this request.")
            return render(request, "ram_app/payment_form.html", {"ram_request": ram_request})

        Payment.objects.create(
            request=ram_request,
            amount=amount,
            repayment_mounth=repayment_month,
            payment_receipt=payment_receipt,
            received_by=request.user,
            note=note,
        )

        # Update status to fully paid if balance is cleared
        ram_request.refresh_from_db()
        if ram_request.is_fully_paid:
            ram_request.status = "fully Paid"
            ram_request.save()
            messages.success(request, f"Payment of ₦{amount} recorded. Request is now FULLY PAID!")
        else:
            messages.success(request, f"Payment of ₦{amount} recorded successfully.")

        return redirect("ram_request_detail", pk=pk)

    return render(request, "ram_app/payment_form.html", {"ram_request": ram_request})


@login_required
@group_required(['admin', 'staff'])
def payment_list(request, pk):
    ram_request = get_object_or_404(RamRequest, pk=pk)
    payments = ram_request.payments.select_related("received_by").all()
    return render(request, "ram_app/payment_list.html", {"ram_request": ram_request,"payments": payments,})

@login_required
@group_required(['admin', 'staff'])    
def ram_payment_list(request):
    paymentObj = Payment.objects.select_related("request__member__member", "received_by").all()
    
    search = request.GET.get("search", "").strip()   # member name or guest name
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()
    
    if search:
        paymentObj = paymentObj.filter(
            Q(request__member__member__first_name__icontains=search) |
            Q(request__member__member__last_name__icontains=search) |
            Q(request__guest_name__icontains=search) |
            Q(request__member__ippis__icontains=search) |
            Q(request__guest_ippis__icontains=search)
        )
        
    if date_from:
        paymentObj = paymentObj.filter(date_paid__date__gte=date_from)

    if date_to:
        paymentObj = paymentObj.filter(date_paid__date__lte=date_to)
    
    
    context = {'paymentObj':paymentObj, "filter_search": search, "filter_date_from": date_from, "filter_date_to": date_to,}
    return render(request,'ram_app/ram_payment_list.html',context)    
    
@login_required
@group_required(['admin'])
def payment_upload(request):
    if not request.user.is_staff:
        messages.error(request, "You are not authorized to perform this action.")
        return redirect("ram_request_list")

    if request.method == "POST":

        # ── Download skipped report ───────────────────────────
        if request.POST.get("action") == "download_skipped":
            skipped_rows = request.session.get("skipped_rows", [])
            return export_skipped_excel(skipped_rows)

        # ── Get repayment month from form ─────────────────────
        repayment_month_str = request.POST.get("repayment_month", "").strip()
        excel_file = request.FILES.get("excel_file")

        if not repayment_month_str:
            messages.error(request, "Please select a repayment month.")
            return render(request, "ram_app/payment_upload.html")

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return render(request, "ram_app/payment_upload.html")

        if not excel_file.name.endswith((".xlsx", ".xls")):
            messages.error(request, "Invalid file format. Please upload .xlsx or .xls file.")
            return render(request, "ram_app/payment_upload.html")

        try:
            repayment_date = datetime.strptime(repayment_month_str + "-01", "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "Invalid repayment month format.")
            return render(request, "ram_app/payment_upload.html")

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        except Exception:
            messages.error(request, "Could not read the Excel file. Please check the format.")
            return render(request, "ram_app/payment_upload.html")

        success_count = 0
        skipped_rows = []

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        if not rows:
            messages.error(request, "The Excel file has no data rows.")
            return render(request, "ram_app/payment_upload.html")

        with transaction.atomic():
            for row_num, row in enumerate(rows, start=2):
                if not row or all(v is None for v in row):
                    continue

                # ── Parse columns: IPPIS | Amount ────────────
                try:
                    ippis = row[0]
                    amount = row[1]

                    if not ippis:
                        skipped_rows.append({
                            "row": row_num, "ippis": "—", "name": "—", "type": "—",
                            "amount": "—", "month": repayment_date.strftime("%b %Y"),
                            "reason": "Missing IPPIS"
                        })
                        continue

                    ippis = int(ippis)

                    if not amount:
                        skipped_rows.append({
                            "row": row_num, "ippis": ippis, "name": "—", "type": "—",
                            "amount": "—", "month": repayment_date.strftime("%b %Y"),
                            "reason": "Missing amount"
                        })
                        continue

                    amount = Decimal(str(amount)).quantize(Decimal("0.01"))
                    if amount <= 0:
                        skipped_rows.append({
                            "row": row_num, "ippis": ippis, "name": "—", "type": "—",
                            "amount": str(amount), "month": repayment_date.strftime("%b %Y"),
                            "reason": "Amount must be greater than zero"
                        })
                        continue

                except (ValueError, InvalidOperation, TypeError) as e:
                    skipped_rows.append({
                        "row": row_num, "ippis": row[0] if row else "—", "name": "—", "type": "—",
                        "amount": "—", "month": repayment_date.strftime("%b %Y"),
                        "reason": f"Invalid data: {str(e)}"
                    })
                    continue

                # ── Try member first, then guest ──────────────
                ram_request = None
                name = "—"
                requester_type = "—"

                # 1. Check if IPPIS belongs to a Member
                member = Member.objects.filter(ippis=ippis).first()
                if member:
                    name = member.member.get_full_name()
                    requester_type = "Member"
                    ram_request = RamRequest.objects.filter(
                        member=member,
                        status="approved"
                    ).order_by("-date_requested").first()

                    if not ram_request:
                        fully_paid = RamRequest.objects.filter(
                            member=member, status="fully Paid"
                        ).exists()
                        reason = "No active approved request — all requests are fully paid" if fully_paid else "No approved request found for this member"
                        skipped_rows.append({
                            "row": row_num, "ippis": ippis, "name": name, "type": requester_type,
                            "amount": str(amount), "month": repayment_date.strftime("%b %Y"),
                            "reason": reason
                        })
                        continue

                else:
                    # 2. Check if IPPIS belongs to a Guest request
                    ram_request = RamRequest.objects.filter(
                        guest_ippis=ippis,
                        status="approved"
                    ).order_by("-date_requested").first()

                    if ram_request:
                        name = ram_request.guest_name or f"Guest IPPIS {ippis}"
                        requester_type = "Guest"
                    else:
                        # Check if guest has a fully paid request
                        fully_paid_guest = RamRequest.objects.filter(
                            guest_ippis=ippis, status="fully Paid"
                        ).first()

                        if fully_paid_guest:
                            skipped_rows.append({
                                "row": row_num, "ippis": ippis,
                                "name": fully_paid_guest.guest_name or f"Guest IPPIS {ippis}",
                                "type": "Guest",
                                "amount": str(amount), "month": repayment_date.strftime("%b %Y"),
                                "reason": "No active approved request — all requests are fully paid"
                            })
                        else:
                            skipped_rows.append({
                                "row": row_num, "ippis": ippis, "name": "—", "type": "—",
                                "amount": str(amount), "month": repayment_date.strftime("%b %Y"),
                                "reason": f"IPPIS {ippis} not found as member or guest"
                            })
                        continue

                # ── Check duplicate ───────────────────────────
                duplicate = Payment.objects.filter(
                    request=ram_request,
                    repayment_mounth__year=repayment_date.year,
                    repayment_mounth__month=repayment_date.month,
                ).exists()

                if duplicate:
                    skipped_rows.append({
                        "row": row_num, "ippis": ippis, "name": name, "type": requester_type,
                        "amount": str(amount), "month": repayment_date.strftime("%b %Y"),
                        "reason": f"Duplicate — payment already exists for {repayment_date.strftime('%B %Y')}"
                    })
                    continue

                # ── Cap amount at balance ─────────────────────
                if amount > ram_request.balance_remaining:
                    amount = ram_request.balance_remaining

                # ── Create payment ────────────────────────────
                Payment.objects.create(
                    request=ram_request,
                    amount=amount,
                    repayment_mounth=repayment_date,
                    received_by=request.user,
                    note=f"Uploaded via Excel — {repayment_date.strftime('%B %Y')}",
                )
                success_count += 1

                # ── Auto update to fully paid ─────────────────
                ram_request.refresh_from_db()
                if ram_request.is_fully_paid:
                    ram_request.status = "fully Paid"
                    ram_request.save()

        # ── Save skipped to session ───────────────────────────
        request.session["skipped_rows"] = skipped_rows

        if success_count:
            messages.success(request, f"✅ {success_count} payment(s) uploaded for {repayment_date.strftime('%B %Y')}.")
        if skipped_rows:
            messages.warning(request, f"⚠️ {len(skipped_rows)} row(s) were skipped. See the table below.")

        context = {
            "skipped_rows": skipped_rows,
            "success_count": success_count,
        }
        return render(request, "ram_app/payment_upload.html", context)

    request.session.pop("skipped_rows", None)
    return render(request, "ram_app/payment_upload.html")


def export_skipped_excel(skipped_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Skipped Payments"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    center = Alignment(horizontal="center")

    headers = ["Row", "IPPIS", "Name", "Type", "Amount", "Month", "Reason"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_num, item in enumerate(skipped_rows, 2):
        ws.cell(row=row_num, column=1, value=item.get("row"))
        ws.cell(row=row_num, column=2, value=item.get("ippis"))
        ws.cell(row=row_num, column=3, value=item.get("name"))
        ws.cell(row=row_num, column=4, value=item.get("type"))
        ws.cell(row=row_num, column=5, value=item.get("amount"))
        ws.cell(row=row_num, column=6, value=item.get("month"))
        ws.cell(row=row_num, column=7, value=item.get("reason"))

        if row_num % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).fill = PatternFill(
                    start_color="FDECEA", end_color="FDECEA", fill_type="solid"
                )

    col_widths = [6, 12, 25, 10, 12, 12, 50]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="skipped_payments.xlsx"'
    wb.save(response)
    return response