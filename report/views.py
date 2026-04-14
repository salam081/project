import calendar
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils.dateparse import parse_date
from django.contrib.auth import get_user_model
from openpyxl.utils import get_column_letter
from django.utils import timezone as django_timezone
from django.db.models import Sum,Min, Max, Count, Count, Q, Avg , Min, Max, F, Case, When ,ExpressionWrapper, DecimalField, ExpressionWrapper
from django.db.models.functions import Coalesce
from django.db.models.functions import TruncMonth, TruncYear,TruncWeek
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from collections import defaultdict
from django.db.models.functions import ExtractYear
from django.db.models.functions import ExtractMonth
from django.db import transaction
from decimal import Decimal
from django.contrib import messages
import json
from django.utils import timezone
from datetime import datetime, time
import logging
from consumable.models import *
from accounts.models import User
import csv
from openpyxl import Workbook
from django.http import HttpResponse
from loan.models import *
from savings.models import *
from special_savings.models import *
from main.models import *
from PurchasedItems.models import *
from member.models import *
from projectfinance.models import *
from django.db.models.functions import Coalesce, ExtractYear
from django.core.paginator import Paginator
import logging
from datetime import datetime
from django.db.models import Q, Sum, F, DecimalField, ExpressionWrapper
from django.db.models.functions import Coalesce
import logging
from openpyxl import Workbook
from datetime import datetime, timedelta
from django.db.models.functions import TruncMonth, TruncWeek


User = get_user_model()
# from accounts.decorator import group_required

from datetime import date


@login_required
def all_income(request):
    """
    Calculates and displays all sources of income based on user-selected date filters.
    Requires the user to be logged in to view.
    """
    # Get filter parameters from GET request
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # Build Q objects for filtering
    savings_filter = Q()
    loanrepayback_filter = Q()
    consumable_payback_filter = Q()
    loan_fee_filter = Q()
    interest_filter = Q()
    item_purchase_filter = Q()
    total_consumable_form_fee_filter = Q()

    # Filter by date range
    if date_from:
        savings_filter &= Q(month__gte=date_from)
        loanrepayback_filter &= Q(repayment_date__gte=date_from)
        consumable_payback_filter &= Q(repayment_date__gte=date_from)
        loan_fee_filter &= Q(created_at__gte=date_from)
        interest_filter &= Q(month__gte=date_from)
        item_purchase_filter &= Q(date_created__gte=date_from)  # Use 'date_created' for filtering
        total_consumable_form_fee_filter &= Q(created_at__gte=date_from)

    if date_to:
        savings_filter &= Q(month__lte=date_to)
        loanrepayback_filter &= Q(repayment_date__lte=date_to)
        consumable_payback_filter &= Q(repayment_date__lte=date_to)
        loan_fee_filter &= Q(created_at__lte=date_to)
        interest_filter &= Q(month__lte=date_to)
        item_purchase_filter &= Q(date_created__lte=date_to)  # Use 'date_created' for filtering
        total_consumable_form_fee_filter &= Q(created_at__lte=date_to)

    # Aggregations with filters
    total_loans_fee = LoanRequestFee.objects.filter(loan_fee_filter).aggregate(total=Sum('form_fee'))['total'] or 0
    total_consumable_form_fee = ConsumableFormFee.objects.filter(total_consumable_form_fee_filter).aggregate(total=Sum('form_fee'))['total'] or 0
    total_savings = Savings.objects.filter(savings_filter).aggregate(total_savings=Sum('month_saving'))['total_savings'] or 0
    deducted_amount = Interest.objects.filter(interest_filter).aggregate(total_savings=Sum('amount_deducted'))['total_savings'] or 0
    payback_loans = LoanRepayback.objects.filter(loanrepayback_filter).aggregate(total=Sum('amount_paid'))['total'] or 0
    total_consumable_payback = PaybackConsumable.objects.filter(consumable_payback_filter).aggregate(total=Sum('amount_paid'))['total'] or 0

    # CORRECTED: Use 'expenditure_amount' instead of 'total_price' and filter by 'date_added'
    total_item_purchase = SellingPlan.objects.filter(item_purchase_filter).aggregate(total=Sum('profit'))['total'] or 0
    
    # Calculate total income
    income = sum([total_loans_fee, total_savings, deducted_amount, payback_loans, total_consumable_payback, total_consumable_form_fee, total_item_purchase])

    filters_applied = any([date_from, date_to])
    context = {
        'total_loans_fee': total_loans_fee,
        'total_consumable_form_fee': total_consumable_form_fee,
        'total_savings': total_savings,
        'payback_loans': payback_loans,
        'deducted_amount': deducted_amount,
        'total_consumable_payback': total_consumable_payback,
        'income': income,
        'date_from': date_from,
        'date_to': date_to,
        'filters_applied': filters_applied,
        'total_item_purchase': total_item_purchase
    }
    return render(request, "reports/all_income.html", context)

@login_required 
def summary_view(request):
    def get_monthly_totals(queryset, value_field):
        return (
            queryset
            .annotate(year=ExtractYear("month"), month_num=ExtractMonth("month"))
            .values("year", "month_num")
            .annotate(total=Sum(value_field, default=Decimal('0.00')))
            .order_by("-year", "-month_num")
        )

    def format_months(data):
        return [
            {
                "year": row["year"],
                "month_num": row["month_num"],
                "month": calendar.month_name[row["month_num"]],
                "total": Decimal(row["total"] or '0.00'),
            }
            for row in data
        ]

    # Use .all() or apply necessary filters to base querysets
    savings_monthly = format_months(get_monthly_totals(Savings.objects.all(), "month_saving"))
    interest_monthly = format_months(get_monthly_totals(Interest.objects.all(), "amount_deducted"))
    loanable_monthly = format_months(get_monthly_totals(Loanable.objects.all(), "amount"))
    investment_monthly = format_months(get_monthly_totals(Investment.objects.all(), "amount"))

    # Paginate the lists
    def paginate_data(data, page_size=12):
        paginator = Paginator(data, page_size)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        return page_obj

    savings_page = paginate_data(savings_monthly)
    interest_page = paginate_data(interest_monthly)
    loanable_page = paginate_data(loanable_monthly)
    investment_page = paginate_data(investment_monthly)

    # Total calculations (no pagination here, just sums)
    total_savings = Decimal(sum(item["total"] for item in savings_monthly))
    total_interest = Decimal(sum(item["total"] for item in interest_monthly))
    total_loanable = Decimal(sum(item["total"] for item in loanable_monthly))
    total_investment = Decimal(sum(item["total"] for item in investment_monthly))
    grand_total = total_savings + total_interest #+ total_loanable + total_investment
    # --- Existing Per-member totals ---
    member_savings = Member.objects.annotate(
        aggregated_savings=Sum('savings__month_saving', default=Decimal('0.00'))
    ).order_by('-aggregated_savings')
    
    # Optimized member interest fetching slightly
    member_interest_data = Member.objects.annotate(
        total_interest=Sum('interest__amount_deducted', default=Decimal('0.00'))
    )
    member_interest = {m.id: m.total_interest for m in member_interest_data}
    # print('member_interest_data', member_interest)

    context = {
        "savings_page": savings_page,
        "interest_page": interest_page,
        "loanable_page": loanable_page,
        "investment_page": investment_page,
        "total_savings": total_savings,
        "total_interest": total_interest,
        "total_loanable": total_loanable,
        "total_investment": total_investment,
        "grand_total": grand_total,
        "member_savings": member_savings,
        "member_interest": member_interest,
    }

    return render(request, "reports/summary.html", context)


def admin_loan_reports(request):
    # Default to current month
    month = request.GET.get('month', timezone.now().strftime('%Y-%m'))
    year, month_num = month.split('-')
    


    loan_type_id = request.GET.get('loan_type')

    monthly_requests = LoanRequest.objects.filter(
        application_date__year=year,
        # application_date__month=month_num
    )

    if loan_type_id:
        monthly_requests = monthly_requests.filter(loan_type_id=loan_type_id)

    monthly_approvals = monthly_requests.filter(status='approved')
    monthly_rejections = monthly_requests.filter(status='rejected')

    
    monthly_repayments = LoanRepayback.objects.filter(
        repayment_date__year=year,
        # repayment_date__month=month_num

    )
    if loan_type_id:
        monthly_repayments = monthly_repayments.filter(loan_request__loan_type_id=loan_type_id)

    loan_type_stats = LoanType.objects.annotate(
        total_requests=Count('loanrequest'),
        total_approved=Count('loanrequest', filter=Q(loanrequest__status='approved')),
        total_amount=Sum('loanrequest__approved_amount', filter=Q(loanrequest__status='approved'))
    )

    context = {
        'selected_month': month,
        'selected_loan_type': int(loan_type_id) if loan_type_id else None,
        'monthly_requests': monthly_requests.count(),
        'monthly_approvals': monthly_approvals.count(),
        'monthly_rejections': monthly_rejections.count(),
        'monthly_approved_amount': monthly_approvals.aggregate(Sum('approved_amount'))['approved_amount__sum'] or 0,
        'monthly_repayments': monthly_repayments.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
        'loan_type_stats': loan_type_stats,
        'loan_types': LoanType.objects.all(),  
    }
    return render(request, 'reports/reports.html', context)



@login_required
def loan_request_report(request):
    # Initialize query set with all loan requests
    loan_requests = LoanRequest.objects.all().order_by('-date_created')

    # Initialize filters dictionary
    filters = {
        'status': request.GET.get('status', 'all'),
        'date_from': request.GET.get('date_from'),
        'date_to': request.GET.get('date_to'),
        'month': request.GET.get('month'),
        'loan_type': request.GET.get('loan_type'),
    }

    # Apply filters
    if filters['status'] and filters['status'] != 'all':
        loan_requests = loan_requests.filter(status=filters['status'])

    if filters['date_from']:
        loan_requests = loan_requests.filter(application_date__gte=filters['date_from'])

    if filters['date_to']:
        loan_requests = loan_requests.filter(application_date__lte=filters['date_to'])

    if filters['month']:
        try:
            year, month = map(int, filters['month'].split('-'))
            loan_requests = loan_requests.filter(
                application_date__year=year,
                application_date__month=month
            )
        except ValueError:
            pass

    if filters['loan_type']:
        loan_requests = loan_requests.filter(loan_type__name=filters['loan_type'])

    # Add computed fields
    loan_requests = loan_requests.annotate(
        total_paid=Coalesce(Sum('repaybacks__amount_paid'), 0, output_field=DecimalField()),
        balance_value=ExpressionWrapper(
            F('approved_amount') - Coalesce(Sum('repaybacks__amount_paid'), 0),
            output_field=DecimalField()
        ),
        total_price=Coalesce(F('approved_amount'), F('amount'), output_field=DecimalField())
    )

    # Compute summary
    base_qs = LoanRequest.objects.all()
    if filters['status'] and filters['status'] != 'all':
        base_qs = base_qs.filter(status=filters['status'])
    if filters['date_from']:
        base_qs = base_qs.filter(application_date__gte=filters['date_from'])
    if filters['date_to']:
        base_qs = base_qs.filter(application_date__lte=filters['date_to'])
    if filters['month']:
        try:
            year, month = map(int, filters['month'].split('-'))
            base_qs = base_qs.filter(application_date__year=year, application_date__month=month)
        except ValueError:
            pass
    if filters['loan_type']:
        base_qs = base_qs.filter(loan_type__name=filters['loan_type'])

    total_value = base_qs.aggregate(
        total=Coalesce(Sum('approved_amount'), 0, output_field=DecimalField())
    )['total']

    total_paid = LoanRepayback.objects.filter(
        loan_request__in=base_qs
    ).aggregate(
        total=Coalesce(Sum('amount_paid'), 0, output_field=DecimalField())
    )['total']

    summary = {
        'total_requests': base_qs.count(),
        'total_value': total_value,
        'total_paid': total_paid,
        'total_balance': total_value - total_paid,
        'pending_count': base_qs.filter(status='pending').count(),
        'approved_count': base_qs.filter(status='approved').count(),
        'paid_count': base_qs.filter(status='Fullpaid').count(),
        'declined_count': base_qs.filter(status='rejected').count(),
    }

    # ✅ Excel export logic
    if request.GET.get('download') == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Loan Requests Report"

        # Header row
        headers = [
            "ID", "Member", "Loan Type", "Status", "Application Date",
            "Approved Amount", "Total Paid", "Balance", "Created By"
        ]
        ws.append(headers)

        # Data rows
        for loan in loan_requests:
            ws.append([
                loan.id,
                f"{loan.member.member.first_name} {loan.member.member.last_name}" if loan.member and loan.member.member else "",
                loan.loan_type.name if loan.loan_type else "",
                loan.status,
                loan.application_date.strftime("%Y-%m-%d") if loan.application_date else "",
                float(loan.approved_amount or 0),
                float(loan.total_paid or 0),
                float(loan.balance_value or 0),
                loan.created_by.username if loan.created_by else "",
            ])

        # Auto-adjust column width
        for i, col in enumerate(ws.columns, start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(i)].width = max(15, max_length + 2)

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = "loan_request_report.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

    # Pagination (for normal page view)
    page_number = request.GET.get('page')
    paginator = Paginator(loan_requests, 100)
    page_obj = paginator.get_page(page_number)

    status_choices = [('all', 'All Statuses')] + list(LoanRequest.status.field.choices)
    months = [d for d in LoanRequest.objects.dates('application_date', 'month', order='DESC')]
    loan_types_qs = LoanType.objects.all().order_by("name")

    context = {
        'requests': page_obj,
        'summary': summary,
        'filters': filters,
        'status_choices': status_choices,
        'months': months,
        'loan_types': loan_types_qs,
    }

    return render(request, 'reports/loan_request_report.html', context)

@login_required
def filtered_loan_repayments(request):
    # 1. Get filter options for dropdowns (unfiltered)
    years = (
        LoanRequest.objects.annotate(year=ExtractYear("application_date"))
        .values_list("year", flat=True)
        .distinct()
        .order_by("-year")
    )
    loan_types = (
        LoanType.objects.filter(available=True)
        .values_list("name", flat=True)
        .order_by("name")
    )
    loan_statuses = ['approved', 'Fullpaid']

    # 2. Get user's filter selections from the request
    selected_year = request.GET.get("year")
    selected_type = request.GET.get("loan_type")
    selected_status = request.GET.get("status")

    # 3. Build the master filter (Q object)
    # Start the filter with the base status condition
    filters = Q(status__in=['approved', 'Fullpaid'])
    
    if selected_year:
        filters &= Q(application_date__year=selected_year)
    if selected_type:
        filters &= Q(loan_type__name=selected_type)
    if selected_status:
        # Add the selected status filter to the Q object
        filters &= Q(status=selected_status)

    # 4. Create the base filtered queryset. This is the key.
    # It aggregates the total paid amount for each loan.
    base_queryset = (
        LoanRequest.objects.filter(filters)
        .annotate(
            total_paid_by_loan=Coalesce(Sum("repaybacks__amount_paid"), Decimal("0.00"), output_field=DecimalField())
        )
        .order_by("-application_date")
    )

    # 5. Calculate summary statistics from the filtered queryset
    summary_stats = base_queryset.aggregate(
        total_loans=Count('id'),
        total_approved_amount=Coalesce(Sum('approved_amount'), Decimal('0.00'), output_field=DecimalField()),
        total_amount_paid=Coalesce(Sum('total_paid_by_loan'), Decimal('0.00'), output_field=DecimalField()),
    )
    summary_stats['total_outstanding'] = summary_stats['total_approved_amount'] - summary_stats['total_amount_paid']

    # 6. Process loan data for the main table (for pagination)
    processed_loans = []
    for loan in base_queryset:
        approved_amount = loan.approved_amount or Decimal('0.00')
        balance_remaining = approved_amount - loan.total_paid_by_loan
        payment_percentage = (loan.total_paid_by_loan / approved_amount * 100) if approved_amount > 0 else 0
        
        if balance_remaining <= 0:
            payment_status, status_class = "Fully Paid", "success"
        elif loan.total_paid_by_loan > 0:
            payment_status, status_class = "Partial Payment", "warning"
        else:
            payment_status, status_class = "No Payment", "danger"
            
        processed_loans.append({
            'loan': loan,
            'total_paid': loan.total_paid_by_loan,
            'balance_remaining': balance_remaining,
            'payment_percentage': round(payment_percentage, 1),
            'payment_status': payment_status,
            'status_class': status_class,
        })

    # 7. Handle pagination
    paginator = Paginator(processed_loans, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 8. Pass all data to the template
    context = {
        'page_obj': page_obj,
        'years': years,
        'loan_types': loan_types,
        'loan_statuses': loan_statuses,
        'selected_year': selected_year,
        'selected_type': selected_type,
        'selected_status': selected_status,
        'summary_stats': summary_stats,
    }
    
    return render(request, "reports/filtered_loan_repayments.html", context)



from openpyxl.utils import get_column_letter
@login_required
def request_status_report(request):
    # --- 1. Get filters ---
    status_filter = request.GET.get('status', 'all')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    user_filter = request.GET.get('user')
    consumable_type_filter = request.GET.get('consumable_type')

    # --- 2. Base queryset ---
    queryset = ConsumableRequest.objects.select_related(
        'user', 'approved_by', 'consumable_type'
    ).prefetch_related('details__selling_item', 'repayments')

    # --- 3. Apply filters ---
    if status_filter != 'all':
        queryset = queryset.filter(status=status_filter)

    if date_from:
        try:
            date_from_parsed = datetime.strptime(date_from, '%Y-%m-%d').date()
            queryset = queryset.filter(date_created__date__gte=date_from_parsed)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_parsed = datetime.strptime(date_to, '%Y-%m-%d').date()
            queryset = queryset.filter(date_created__date__lte=date_to_parsed)
        except ValueError:
            pass

    if user_filter:
        try:
            user_id = int(user_filter)
            queryset = queryset.filter(user_id=user_id)
        except (ValueError, TypeError):
            pass

    if consumable_type_filter:
        try:
            consumable_type_id = int(consumable_type_filter)
            queryset = queryset.filter(consumable_type_id=consumable_type_id)
        except (ValueError, TypeError):
            pass

    queryset = queryset.order_by('-date_created')

    # --- 4. Build list with calculations ---
    requests_with_calculations = []
    for req in queryset:
        total_price = Decimal(req.calculate_total_price() or 0)
        total_paid = Decimal(req.total_paid or 0)
        balance = Decimal(req.balance or 0)
        items_count = req.details.count()

        requests_with_calculations.append({
            'id': req.id,
            'user': req.user,
            'date_created': req.date_created,
            'status': req.status,
            'total_price': total_price,
            'total_paid': total_paid,
            'balance': balance,
            'items_count': items_count,
            'approved_by': req.approved_by,
            'consumable_type': req.consumable_type,
        })

    # --- 5. Summary statistics ---
    total_requests = len(requests_with_calculations)
    pending_count = sum(1 for r in requests_with_calculations if r['status'] == 'Pending')
    approved_count = sum(1 for r in requests_with_calculations if r['status'] == 'Approved')
    itempicked_count = sum(1 for r in requests_with_calculations if r['status'] == 'Itempicked')
    paid_count = sum(1 for r in requests_with_calculations if r['status'] == 'FullyPaid')
    declined_count = sum(1 for r in requests_with_calculations if r['status'] == 'Declined')

    total_value = sum(r['total_price'] for r in requests_with_calculations)
    total_paid_sum = sum(r['total_paid'] for r in requests_with_calculations)
    total_balance = total_value - total_paid_sum

    summary = {
        'total_requests': total_requests,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'itempicked_count': itempicked_count,
        'paid_count': paid_count,
        'declined_count': declined_count,
        'total_value': total_value,
        'total_paid': total_paid_sum,
        'total_balance': total_balance,
    }

    # --- ✅ Excel Export ---
    if request.GET.get("download") == "excel":
        wb = Workbook()

        # Sheet 1: Consumable Requests
        ws1 = wb.active
        ws1.title = "Consumable Requests"
        ws1.append([
            "ID", "User", "Date Created", "Status", "Consumable Type",
            "Total Price", "Total Paid", "Balance", "Items Count", "Approved By"
        ])

        for item in requests_with_calculations:
            ws1.append([
                item['id'],
                item['user'].get_full_name() if item['user'] else "N/A",
                item['date_created'].strftime("%Y-%m-%d %H:%M"),
                item['status'],
                item['consumable_type'].name if item['consumable_type'] else "N/A",
                float(item['total_price']),
                float(item['total_paid']),
                float(item['balance']),
                item['items_count'],
                item['approved_by'].get_full_name() if item['approved_by'] else "N/A",
            ])

        for col in ws1.columns:
            max_length = max(len(str(c.value)) for c in col if c.value)
            ws1.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        # Sheet 2: Summary
        ws2 = wb.create_sheet(title="Summary Report")
        ws2.append(["Metric", "Value"])
        ws2.append(["Total Requests", total_requests])
        ws2.append(["Pending", pending_count])
        ws2.append(["Approved", approved_count])
        ws2.append(["Item Picked", itempicked_count])
        ws2.append(["Fully Paid", paid_count])
        ws2.append(["Declined", declined_count])
        ws2.append(["Total Value", float(total_value)])
        ws2.append(["Total Paid", float(total_paid_sum)])
        ws2.append(["Total Balance", float(total_balance)])

        for col in ws2.columns:
            max_length = max(len(str(c.value)) for c in col if c.value)
            ws2.column_dimensions[get_column_letter(col[0].column)].width = max_length + 4

        # Send file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="request_status_report.xlsx"'
        wb.save(response)
        return response

    # --- 6. Paginate and render ---
    consumable_types = ConsumableType.objects.filter(available=True).order_by('name')
    users_with_requests = User.objects.filter(consumablerequest__isnull=False).distinct().order_by('username')
    paginator = Paginator(requests_with_calculations, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'requests': page_obj,
        'summary': summary,
        'users_with_requests': users_with_requests,
        'consumable_types': consumable_types,
        'filters': {
            'status': status_filter,
            'date_from': date_from,
            'date_to': date_to,
            'user': user_filter,
            'consumable_type': consumable_type_filter,
        },
        'status_choices': [
            ('all', 'All Statuses'),
            ('Pending', 'Pending'),
            ('Approved', 'Approved'),
            ('Itempicked', 'Item Picked'),
            ('FullyPaid', 'Fully Paid'),
            ('Declined', 'Declined'),
        ]
    }

    return render(request, 'reports/consumable_request_status_report.html', context)




@login_required
def payment_analysis_report(request):
    """Detailed payment analysis and trends with enhanced filtering and performance"""

    # 1. --- FILTER INPUTS ---
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    user_id = request.GET.get('user_id')
    status_filter = request.GET.get('status', 'all')
    download = request.GET.get('download')

    parsed_date_from = None
    parsed_date_to = None

    try:
        if date_from:
            parsed_date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        if date_to:
            parsed_date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        if parsed_date_from and parsed_date_to and parsed_date_from > parsed_date_to:
            messages.error(request, "Start date cannot be after end date.")
            parsed_date_from = parsed_date_to = None
    except ValueError:
        messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
        parsed_date_from = parsed_date_to = None

    # 2. --- BASE QUERYSET ---
    queryset = PaybackConsumable.objects.select_related(
        'consumable_request__user',
        'consumable_request'
    )

    # Apply date filters
    if parsed_date_from:
        queryset = queryset.filter(repayment_date__gte=parsed_date_from)
    if parsed_date_to:
        queryset = queryset.filter(repayment_date__lte=parsed_date_to)
    if user_id:
        queryset = queryset.filter(consumable_request__user_id=user_id)

    # 3. --- PAYMENT STATS ---
    payment_stats = queryset.aggregate(
        total_payments=Sum('amount_paid'),
        avg_payment=Avg('amount_paid'),
        payment_count=Count('id'),
        min_payment=Min('amount_paid'),
        max_payment=Max('amount_paid')
    )
    payment_stats = {k: Decimal(v or 0) for k, v in payment_stats.items()}

    # 4. --- MONTHLY PAYMENTS ---
    monthly_payments = queryset.annotate(
        month=TruncMonth('repayment_date')
    ).values('month').annotate(
        total_paid=Sum('amount_paid'),
        payment_count=Count('id'),
        avg_payment=Avg('amount_paid')
    ).order_by('month')

    # 5. --- WEEKLY PAYMENTS ---
    three_months_ago = timezone.now().date() - timedelta(days=90)
    weekly_payments = queryset.filter(
        repayment_date__gte=three_months_ago
    ).annotate(
        week=TruncWeek('repayment_date')
    ).values('week').annotate(
        total_paid=Sum('amount_paid'),
        payment_count=Count('id')
    ).order_by('week')

    # 6. --- OUTSTANDING BALANCES ---
    outstanding_filter = Q(status__in=['Pending', 'Approved', 'Itempicked'])
    if status_filter != 'all':
        outstanding_filter &= Q(status=status_filter)

    outstanding_data = ConsumableRequest.objects.filter(outstanding_filter).annotate(
        total_price=Sum(F('details__quantity') * F('details__selling_item')),
        total_paid=Sum('repayments__amount_paid'),
    ).annotate(
        balance=F('total_price') - F('total_paid')
    ).filter(
        balance__gt=0
    ).select_related('user').order_by('-balance', '-date_created')

    total_outstanding = outstanding_data.aggregate(total=Sum('balance'))['total'] or 0

    outstanding_data_with_urgency = []
    for req in outstanding_data:
        days_outstanding = (timezone.now().date() - req.date_created.date()).days
        if days_outstanding > 90:
            urgency = 'critical'
        elif days_outstanding > 60:
            urgency = 'high'
        elif days_outstanding > 30:
            urgency = 'medium'
        else:
            urgency = 'low'

        payment_percentage = (req.total_paid / req.total_price * 100) if req.total_price else 0
        outstanding_data_with_urgency.append({
            'request': req,
            'total_price': req.total_price or Decimal(0),
            'total_paid': req.total_paid or Decimal(0),
            'balance': req.balance or Decimal(0),
            'days_outstanding': days_outstanding,
            'urgency': urgency,
            'payment_percentage': round(payment_percentage, 2),
        })

    outstanding_summary = {
        'critical': sum(1 for x in outstanding_data_with_urgency if x['urgency'] == 'critical'),
        'high': sum(1 for x in outstanding_data_with_urgency if x['urgency'] == 'high'),
        'medium': sum(1 for x in outstanding_data_with_urgency if x['urgency'] == 'medium'),
        'low': sum(1 for x in outstanding_data_with_urgency if x['urgency'] == 'low'),
    }

    top_users = queryset.values(
        'consumable_request__user__username',
        'consumable_request__user__first_name',
        'consumable_request__user__last_name'
    ).annotate(
        total_paid=Sum('amount_paid'),
        payment_count=Count('id')
    ).order_by('-total_paid')[:10]

    recent_payments = queryset.filter(
        repayment_date__gte=timezone.now().date() - timedelta(days=30)
    ).select_related('consumable_request__user').order_by('-repayment_date')[:10]

    month_list = ConsumableRequest.objects.dates('date_created', 'month', order='DESC')
    users_list = ConsumableRequest.objects.select_related('user').values(
        'user__id', 'user__username', 'user__first_name', 'user__last_name'
    ).distinct().order_by('user__username')

    # ✅ --- 13. HANDLE EXCEL DOWNLOAD ---
    if request.GET.get("download") == "excel":
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        wb = Workbook()

        # ======== Sheet 1: Monthly Trends ========
        ws1 = wb.active
        ws1.title = "Monthly Trends"

        ws1.append(["Month", "Total Paid", "Payments Count"])
        for item in monthly_payments:
            ws1.append([
                item["month"].strftime("%B %Y"),
                float(item["total_paid"]),
                item["payment_count"],
            ])

        # Auto-adjust column widths
        for column_cells in ws1.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws1.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        # ======== Sheet 2: Outstanding Payments ========
        ws2 = wb.create_sheet(title="Outstanding Payments")

        ws2.append(["ID", "Member", "Date Created", "Total Price", "Total Paid", "Balance"])

        for item in outstanding_data:
            member_name = (
                item.request.user.get_full_name()
                if getattr(item.request, "user", None)
                else f"{item.guest_name} ({item.guest_ippiss})"
            )
            ws2.append([
                f"#{item.request.id}",
                member_name,
                item.request.date_created.strftime("%Y-%m-%d"),
                float(item.total_price),
                float(item.total_paid),
                float(item.balance),
            ])

        # Auto-adjust column widths
        for column_cells in ws2.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws2.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        # ======== Send Excel file ========
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="payment_analysis.xlsx"'
        wb.save(response)
        return response



        # Data Rows
        for item in outstanding_data_with_urgency:
            ws.append([
                item['request'].id,
                item['request'].user.get_full_name() if item['request'].user else "N/A",
                item['request'].date_created.strftime("%Y-%m-%d"),
                float(item['total_price']),
                float(item['total_paid']),
                float(item['balance']),
                item['urgency'].capitalize(),
                item['days_outstanding']
            ])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = max_length + 3

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="payment_analysis_report.xlsx"'
        wb.save(response)
        return response

    # --- CONTEXT ---
    context = {
        'payment_stats': payment_stats,
        'monthly_payments': list(monthly_payments),
        'weekly_payments': list(weekly_payments),
        'outstanding_data': outstanding_data_with_urgency,
        'outstanding_summary': outstanding_summary,
        'top_users': top_users,
        'recent_payments': recent_payments,
        'months': month_list,
        'users_list': users_list,
        'total_outstanding': total_outstanding,
        'filters': {
            'date_from': date_from,
            'date_to': date_to,
            'user_id': user_id,
            'status': status_filter
        },
        'date_range_summary': {
            'start': parsed_date_from,
            'end': parsed_date_to,
            'days': (parsed_date_to - parsed_date_from).days if parsed_date_from and parsed_date_to else None
        }
    }

    return render(request, 'reports/consumable_payment_analysis_report.html', context)


@login_required
def user_spending_report(request):
    """Report showing spending patterns by user"""
    
    # Get date range filters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Base queryset
    queryset = User.objects.all()
    
    user_spending = []
    for user in queryset:
        user_requests = ConsumableRequest.objects.filter(user=user)
        
        # Apply date filters
        if date_from:
            user_requests = user_requests.filter(date_created__gte=date_from)
        if date_to:
            user_requests = user_requests.filter(date_created__lte=date_to)
        
        total_requested = sum(req.calculate_total_price() for req in user_requests)
        total_paid = sum(req.total_paid() for req in user_requests)
        outstanding_balance = total_requested - total_paid
        
        if total_requested > 0:  # Only include users with requests
            user_spending.append({
                'user': user,
                'total_requests': user_requests.count(),
                'total_requested': total_requested,
                'total_paid': total_paid,
                'outstanding_balance': outstanding_balance,
                'pending_requests': user_requests.filter(status='Pending').count(),
                'approved_requests': user_requests.filter(status='Approved').count(),
                'paid_requests': user_requests.filter(status='Fullpaid').count(),
                'payment_completion_rate': (total_paid / total_requested * 100) if total_requested > 0 else 0
            })
    
    # Sort by total requested (descending)
    user_spending.sort(key=lambda x: x['total_requested'], reverse=True)
    
    # Calculate totals
    totals = {
        'total_users': len(user_spending),
        'total_requested': sum(u['total_requested'] for u in user_spending),
        'total_paid': sum(u['total_paid'] for u in user_spending),
        'total_outstanding': sum(u['outstanding_balance'] for u in user_spending),
        'avg_request_value': sum(u['total_requested'] for u in user_spending) / len(user_spending) if user_spending else 0
    }
    
    context = {
        'user_spending': user_spending,
        'totals': totals,
        'filters': {
            'date_from': date_from,
            'date_to': date_to
        }
    }
    
    return render(request, 'reports/user_spending_report.html', context)


@login_required
def item_popularity_report(request):
    """Report showing most requested items"""
    
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Base queryset
    queryset = ConsumableRequestDetail.objects.select_related('item', 'request')
    
    # Apply date filters
    if date_from:
        queryset = queryset.filter(request__date_created__gte=date_from)
    if date_to:
        queryset = queryset.filter(request__date_created__lte=date_to)
    
    # Aggregate by item
    item_stats = queryset.values('item__title').annotate(
        total_quantity=Sum('quantity'),
        total_requests=Count('request', distinct=True),
        total_value=Sum('total_price'),
        avg_price=Avg('item_price')
    ).order_by('-total_quantity')
    
    context = {
        'item_stats': item_stats,
        'total_items': item_stats.count(),
        'filters': {
            'date_from': date_from,
            'date_to': date_to
        }
    }
    
    return render(request, 'reports/item_popularity_report.html', context)


@login_required
def approval_workflow_report(request):
    """Report on approval workflow and approver statistics"""
    
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Base queryset
    queryset = ConsumableRequest.objects.select_related('user', 'approved_by')
    
    # Apply date filters
    if date_from:
        queryset = queryset.filter(date_created__gte=date_from)
    if date_to:
        queryset = queryset.filter(date_created__lte=date_to)
    
    # Approval statistics
    approval_stats = {
        'total_requests': queryset.count(),
        'pending_requests': queryset.filter(status='Pending').count(),
        'approved_requests': queryset.filter(status='Approved').count(),
        'declined_requests': queryset.filter(status='Declined').count(),
        'paid_requests': queryset.filter(status='Fullpaid').count(),
    }
    
    # Calculate approval rates
    total_processed = approval_stats['approved_requests'] + approval_stats['declined_requests'] + approval_stats['paid_requests']
    approval_stats['approval_rate'] = (
        (approval_stats['approved_requests'] + approval_stats['paid_requests']) / total_processed * 100
    ) if total_processed > 0 else 0
    
    # Approver statistics
    

# Build approver stats in Python
    approver_dict = defaultdict(lambda: {'total_approved': 0, 'total_value_approved': 0})

    for req in queryset.filter(approved_by__isnull=False):
        username = req.approved_by.username
        approver_dict[username]['total_approved'] += 1
        approver_dict[username]['total_value_approved'] += req.calculate_total_price()

    # Convert to list and sort
    approver_stats = [
        {'approved_by__username': username, **stats}
        for username, stats in approver_dict.items()
    ]
    approver_stats.sort(key=lambda x: x['total_approved'], reverse=True)

    
    # Average approval time (for approved requests with approval_date)
    approved_requests = ConsumableRequestDetail.objects.filter(
        approval_date__isnull=False
    ).select_related('request')
    
    approval_times = []
    for detail in approved_requests:
        days_to_approve = (detail.approval_date - detail.request.date_created.date()).days
        approval_times.append(days_to_approve)
    
    avg_approval_time = sum(approval_times) / len(approval_times) if approval_times else 0
    
    context = { 'approval_stats': approval_stats, 'approver_stats': approver_stats,'avg_approval_time': avg_approval_time,
        'filters': {
            'date_from': date_from,
            'date_to': date_to
        }
    }
    
    return render(request, 'reports/consumable_approval_workflow_report.html', context)


@login_required
def export_report_csv(request):
    """Export reports to CSV format"""
    
    report_type = request.GET.get('type', 'requests')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report.csv"'
    
    writer = csv.writer(response)
    
    if report_type == 'requests':
        # Export all requests
        writer.writerow(['ID', 'User', 'Date Created', 'Status', 'Approved By', 'Total Price', 'Total Paid', 'Balance'])
        
        for req in ConsumableRequest.objects.select_related('user', 'approved_by'):
            writer.writerow([
                req.id,
                req.user.username,
                req.date_created.strftime('%Y-%m-%d'),
                req.status,
                req.approved_by.username if req.approved_by else '',
                req.calculate_total_price(),
                req.total_paid(),
                req.balance()
            ])
    
    elif report_type == 'payments':
        # Export all payments
        writer.writerow(['Request ID', 'User', 'Amount Paid', 'Payment Date', 'Balance Remaining'])
        
        for payment in PaybackConsumable.objects.select_related('consumable_request__user'):
            writer.writerow([
                payment.consumable_request.id,
                payment.consumable_request.user.username,
                payment.amount_paid,
                payment.repayment_date.strftime('%Y-%m-%d'),
                payment.balance_remaining or 0
            ])
    
    elif report_type == 'user_spending':
        # Export user spending summary
        writer.writerow(['User', 'Total Requests', 'Total Requested', 'Total Paid', 'Outstanding Balance', 'Completion Rate'])
        
        for user in User.objects.all():
            user_requests = ConsumableRequest.objects.filter(user=user)
            total_requested = sum(req.calculate_total_price() for req in user_requests)
            total_paid = sum(req.total_paid() for req in user_requests)
            
            if total_requested > 0:
                completion_rate = (total_paid / total_requested * 100) if total_requested > 0 else 0
                writer.writerow([
                    user.username,
                    user_requests.count(),
                    total_requested,
                    total_paid,
                    total_requested - total_paid,
                    f"{completion_rate:.1f}%"
                ])
    
    return response


@login_required
def report_api_data(request):
    """API endpoint for chart data"""
    
    chart_type = request.GET.get('chart', 'monthly_trends')
    
    if chart_type == 'monthly_trends':
        # Monthly request and payment trends
        end_date = timezone.now()
        start_date = end_date - timedelta(days=365)
        
        monthly_data = ConsumableRequest.objects.filter(
            date_created__range=[start_date, end_date]
        ).annotate(
            month=TruncMonth('date_created')
        ).values('month').annotate(
            request_count=Count('id'),
            total_value=Sum('details__total_price')
        ).order_by('month')
        
        return JsonResponse({
            'labels': [item['month'].strftime('%Y-%m') for item in monthly_data],
            'request_counts': [item['request_count'] for item in monthly_data],
            'total_values': [float(item['total_value'] or 0) for item in monthly_data]
        })
    
    elif chart_type == 'status_distribution':
        # Request status distribution
        status_data = ConsumableRequest.objects.values('status').annotate(
            count=Count('id')
        )
        
        return JsonResponse({
            'labels': [item['status'] for item in status_data],
            'data': [item['count'] for item in status_data]
        })
    
    elif chart_type == 'top_items':
        # Top 10 most requested items
        top_items = ConsumableRequestDetail.objects.values('item__title').annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity')[:10]
        
        return JsonResponse({
            'labels': [item['item__title'] for item in top_items],
            'data': [item['total_quantity'] for item in top_items]
        })
    
    return JsonResponse({'error': 'Invalid chart type'}, status=400)



# views.py - Fixed version (no model changes)

logger = logging.getLogger(__name__)

@login_required
def consolidated_report(request):
    start_month_str = request.GET.get('start_month', '').strip()
    end_month_str = request.GET.get('end_month', '').strip()

    parsed_date_from = None
    parsed_date_to = None

    def parse_month_str(value):
        """Handle both '2026-01' and '2026-01-01' formats, return first day of month"""
        value = value[:7]  # take only 'YYYY-MM' part regardless of format
        return datetime.strptime(value, '%Y-%m').date().replace(day=1)

    if start_month_str:
        try:
            parsed_date_from = parse_month_str(start_month_str)
        except (ValueError, TypeError):
            context = {
                'error': 'Invalid start month format.',
                'start_month': start_month_str,
                'end_month': end_month_str,
            }
            return render(request, "reports/consolidated_report.html", context)

    if end_month_str:
        try:
            end_parsed = parse_month_str(end_month_str)
            last_day = calendar.monthrange(end_parsed.year, end_parsed.month)[1]
            parsed_date_to = end_parsed.replace(day=last_day)
        except (ValueError, TypeError):
            context = {
                'error': 'Invalid end month format.',
                'start_month': start_month_str,
                'end_month': end_month_str,
            }
            return render(request, "reports/consolidated_report.html", context)

    if parsed_date_from and parsed_date_to and parsed_date_from > parsed_date_to:
        context = {
            'error': 'Start month cannot be later than end month.',
            'start_month': start_month_str,
            'end_month': end_month_str,
        }
        return render(request, "reports/consolidated_report.html", context)

    try:
       
        filters = {}
        if parsed_date_from:
            filters['date_from'] = django_timezone.make_aware(
                datetime.combine(parsed_date_from, time.min)
            )
        if parsed_date_to:
            filters['date_to'] = django_timezone.make_aware(
                datetime.combine(parsed_date_to, time.max)
            )

        print(f"DEBUG: parsed_date_from={parsed_date_from}, parsed_date_to={parsed_date_to}")
        print(f"DEBUG: filters={filters}")

        expenditure_data = calculate_total_expenditure(filters)
        income_data = calculate_total_income(filters)

        try:
            total_expenditure = sum(expenditure_data.values())
            total_income = sum(income_data.values())
        except (TypeError, ValueError):
            total_expenditure = Decimal('0')
            total_income = Decimal('0')

        net_position = total_income - total_expenditure
        filters_applied = bool(start_month_str or end_month_str)

        # ✅ Normalize display value to 'YYYY-MM' for template input
        display_start = start_month_str[:7] if start_month_str else ''
        display_end = end_month_str[:7] if end_month_str else ''

        context = {
            'total_expenditure': total_expenditure,
            'total_income': total_income,
            'net_position': net_position,
            'start_month': display_start,
            'end_month': display_end,
            'filters_applied': filters_applied,
            **expenditure_data,
            **income_data,
        }

        return render(request, "reports/consolidated_report.html", context)

    except Exception as e:
        import traceback
        traceback.print_exc()
        logger.error(f"Error generating consolidated report: {str(e)}", exc_info=True)
        context = {
            'error': f'An error occurred: {str(e)}',
            'start_month': start_month_str,
            'end_month': end_month_str,
        }
        return render(request, "reports/consolidated_report.html", context)


def calculate_total_expenditure(filters):
    date_from = filters.get('date_from')
    date_to = filters.get('date_to')
    print(f"DEBUG expenditure: date_from={date_from}, date_to={date_to}")

    purchase_filter = Q()
    finance_expenditure_filter = Q()
    loan_disbursement_filter = Q()
    target_savings_disbursement_filter = Q()
    special_savings_disbursement_filter = Q()

    if date_from:
        purchase_filter &= Q(date_added__gte=date_from)
        finance_expenditure_filter &= Q(created_at__gte=date_from)
        loan_disbursement_filter &= Q(date_created__gte=date_from)
        target_savings_disbursement_filter &= Q(requested_at__gte=date_from)
        special_savings_disbursement_filter &= Q(requested_at__gte=date_from)

    if date_to:
        purchase_filter &= Q(date_added__lte=date_to)
        finance_expenditure_filter &= Q(created_at__lte=date_to)
        loan_disbursement_filter &= Q(date_created__lte=date_to)
        target_savings_disbursement_filter &= Q(requested_at__lte=date_to)
        special_savings_disbursement_filter &= Q(requested_at__lte=date_to)

    def safe_sum(queryset, field):
        try:
            return queryset.aggregate(total=Sum(field))['total'] or Decimal('0')
        except Exception as e:
            logger.error(f"Error calculating {field}: {str(e)}")
            import traceback
            traceback.print_exc()
            return Decimal('0')

    try:
        staff_purchases = PurchasedItem.objects.filter(purchase_filter).aggregate(
            total=Sum(
                ExpressionWrapper(
                    F('unit_price') * F('quantity') + F('expenditure_amount'),
                    output_field=DecimalField()
                )
            )
        )['total'] or Decimal('0')
    except Exception as e:
        logger.error(f"Error calculating staff_purchases: {str(e)}")
        import traceback
        traceback.print_exc()
        staff_purchases = Decimal('0')

    member_finance_loans = safe_sum(
        ProjectFinanceRequest.objects.filter(
            finance_expenditure_filter,
            status__in=['Reviewed', 'Approved', 'FullyPaid']
        ),
        'requested_amount'
    )

    loan_disbursements = safe_sum(
        LoanRequest.objects.filter(
            loan_disbursement_filter,
            status__in=['approved', 'Fullpaid'],
            approved_amount__isnull=False
        ),
        'approved_amount'
    )
    target_savings_disbursements = safe_sum(
    TargetSavingsWithdrawal.objects.filter(
        target_savings_disbursement_filter,
        status='approved',
        amount__isnull=False
    ),
    'amount'
)
    special_savings_disbursements = safe_sum(
    SpecialSavingsWithdrawal.objects.filter(
        special_savings_disbursement_filter,
        status='approved',
        amount__isnull=False
    ),
    'amount'
)

    return {
        'staff_purchases': staff_purchases,
        'member_finance_loans': member_finance_loans,
        'loan_disbursements': loan_disbursements,
        'target_savings_disbursements': target_savings_disbursements,
        'special_savings_disbursements': special_savings_disbursements,
    }


def calculate_total_income(filters):
    date_from = filters.get('date_from')
    date_to = filters.get('date_to')

    try:
        def make_filter(date_field):
            q = Q()
            if date_from:
                q &= Q(**{f"{date_field}__gte": date_from})
            if date_to:
                q &= Q(**{f"{date_field}__lte": date_to})
            return q

        def safe_agg(model, q_filter, field):
            try:
                result = model.objects.filter(q_filter).aggregate(
                    total=Sum(field)
                )['total'] or Decimal('0')
                print(f"DEBUG {model.__name__}.{field} = {result}")
                return result
            except Exception as e:
                logger.error(f"Error aggregating {model.__name__}.{field}: {str(e)}")
                import traceback
                traceback.print_exc()
                return Decimal('0')

        # ✅ Savings — filter by 'month' (DateField)
        saving_income = safe_agg(
            Savings, make_filter('month'), 'month_saving'
        )
        # ✅ Interest — filter by 'month' (DateField)
        admin_fee_income = safe_agg(
            Interest, make_filter('month'), 'amount_deducted'
        )
        # ✅ These models need you to share their fields — using date_created for now
        saving_form_fee_income = safe_agg(
            SpecialSavingsTergetSavingsRequestForm, make_filter('date_created'), 'form_fee'
        )
        special_saving_income = safe_agg(
            SpecialSavings, make_filter('date_created'), 'month_savings'
        )
        target_saving_income = safe_agg(
            TargetSavings, make_filter('date_created'), 'month_savings'
        )
        consumable_payback_income = safe_agg(
            PaybackConsumable, make_filter('repayment_date'), 'amount_paid'
        )
        finance_payback_income = safe_agg(
            ProjectFinancePayment, make_filter('created_at'), 'amount_paid'
        )
        form_fee_income = safe_agg(
            ConsumableFormFee, make_filter('created_at'), 'form_fee'
        )
        loan_payback_income = safe_agg(
            LoanRepayback, make_filter('repayment_date'), 'amount_paid'
        )
        loan_form_fee_income = safe_agg(
            LoanRequestFee, make_filter('created_at'), 'form_fee'
        )

        return {
            'saving_income': saving_income,
            'saving_form_fee_income': saving_form_fee_income,
            'special_saving_income': special_saving_income,
            'target_saving_income': target_saving_income,
            'admin_fee_income': admin_fee_income,
            'consumable_payback_income': consumable_payback_income,
            'finance_payback_income': finance_payback_income,
            'form_fee_income': form_fee_income,
            'loan_payback_income': loan_payback_income,
            'loan_form_fee_income': loan_form_fee_income,
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        logger.error(f"Error in calculate_total_income: {str(e)}", exc_info=True)
        return {
            'saving_income': Decimal('0'),
            'saving_form_fee_income': Decimal('0'),
            'special_saving_income': Decimal('0'),
            'target_saving_income': Decimal('0'),
            'admin_fee_income': Decimal('0'),
            'consumable_payback_income': Decimal('0'),
            'finance_payback_income': Decimal('0'),
            'form_fee_income': Decimal('0'),
            'loan_payback_income': Decimal('0'),
            'loan_form_fee_income': Decimal('0'),
        }

# logger = logging.getLogger(__name__)

# @login_required
# def consolidated_report(request):
#     """Generate consolidated financial report with date filtering"""
#     date_from = request.GET.get('date_from')
#     date_to = request.GET.get('date_to')
    
#     parsed_date_from = None
#     parsed_date_to = None
    
#     # Parse start date
#     if date_from:
#         try:
#             parsed_date_from = parse_date(date_from)
#             if not parsed_date_from:
#                 raise ValueError("Invalid date format")
#         except (ValueError, TypeError):
#             context = {
#                 'error': 'Invalid start date format. Please use YYYY-MM-DD format.',
#                 'date_from': date_from,
#                 'date_to': date_to,
#             }
#             return render(request, "reports/consolidated_report.html", context)
    
#     # Parse end date
#     if date_to:
#         try:
#             parsed_date_to = parse_date(date_to)
#             if not parsed_date_to:
#                 raise ValueError("Invalid date format")
#         except (ValueError, TypeError):
#             context = {
#                 'error': 'Invalid end date format. Please use YYYY-MM-DD format.',
#                 'date_from': date_from,
#                 'date_to': date_to,
#             }
#             return render(request, "reports/consolidated_report.html", context)

#     # Check if start date > end date
#     if parsed_date_from and parsed_date_to and parsed_date_from > parsed_date_to:
#         context = {
#             'error': 'Start date cannot be later than end date',
#             'date_from': date_from,
#             'date_to': date_to,
#         }
#         return render(request, "reports/consolidated_report.html", context)

#     try:
#         filters = {}
#         if parsed_date_from:
#             filters['date_from'] = timezone.make_aware(datetime.combine(parsed_date_from, time.min))
#         if parsed_date_to:
#             filters['date_to'] = timezone.make_aware(datetime.combine(parsed_date_to, time.max))

#         # Calculate totals
#         expenditure_data = calculate_total_expenditure(filters)
#         income_data = calculate_total_income(filters)
        
#         # Handle potential errors
#         try:
#             total_expenditure = sum(expenditure_data.values())
#             total_income = sum(income_data.values())
#         except (TypeError, ValueError):
#             total_expenditure = Decimal('0')
#             total_income = Decimal('0')
        
#         net_position = total_income - total_expenditure
#         filters_applied = bool(date_from or date_to)

#         context = {
#             'total_expenditure': total_expenditure,
#             'total_income': total_income,
#             'net_position': net_position,
#             'date_from': date_from,
#             'date_to': date_to,
#             'filters_applied': filters_applied,
#             **expenditure_data,
#             **income_data,
#         }
        
#         return render(request, "reports/consolidated_report.html", context)

#     except Exception as e:
#         logger.error(f"Error generating consolidated report: {str(e)}", exc_info=True)
#         context = {
#             'error': 'An error occurred while generating the report. Please try again.',
#             'date_from': date_from,
#             'date_to': date_to,
#         }
#         return render(request, "reports/consolidated_report.html", context)


# def calculate_total_expenditure(filters):
#     """Calculate total expenditure with proper error handling (consumable expenditure removed)"""
#     date_from = filters.get('date_from')
#     date_to = filters.get('date_to')

#     # Build Q objects for filtering
#     purchase_filter = Q()
#     finance_expenditure_filter = Q()
#     loan_disbursement_filter = Q()

#     if date_from:
#         purchase_filter &= Q(date_added__gte=date_from)
#         finance_expenditure_filter &= Q(created_at__gte=date_from)
#         loan_disbursement_filter &= Q(date_created__gte=date_from)

#     if date_to:
#         purchase_filter &= Q(date_added__lte=date_to)
#         finance_expenditure_filter &= Q(created_at__lte=date_to)
#         loan_disbursement_filter &= Q(date_created__lte=date_to)

#     def safe_sum(queryset, field):
#         try:
#             return queryset.aggregate(total=Sum(field))['total'] or Decimal('0')
#         except Exception as e:
#             logger.error(f"Error calculating {field}: {str(e)}")
#             return Decimal('0')

#     # 1. Staff purchases
#     staff_purchases = safe_sum(
#         PurchasedItem.objects.filter(purchase_filter),
#         F('unit_price') * F('quantity') + F('expenditure_amount')
#     )

#     # 2. Project finance disbursed
#     member_finance_loans = safe_sum(
#         ProjectFinanceRequest.objects.filter(
#             finance_expenditure_filter,
#             status__in=['Reviewed', 'Approved', 'FullyPaid']
#         ),
#         'requested_amount'
#     )

#     # 3. Member loan disbursements
#     loan_disbursements = safe_sum(
#         LoanRequest.objects.filter(
#             loan_disbursement_filter,
#             status__in=['approved', 'Fullpaid'],
#             approved_amount__isnull=False
#         ),
#         'approved_amount'
#     )

#     return {
#         'staff_purchases': staff_purchases,
#         'member_finance_loans': member_finance_loans,
#         'loan_disbursements': loan_disbursements,
#     }

# def calculate_total_income(filters):
#     """Calculate total income with proper filtering and error handling"""
#     date_from = filters.get('date_from')
#     date_to = filters.get('date_to')
    
#     try:
#         # Build Q objects for filtering
#         admin_fee_filter = Q()
#         saving_filter = Q()
#         saving_form_fee_filter = Q()
#         special_saving_filter = Q()
#         target_saving_filter = Q()
#         member_payback_filter = Q()
#         member_finance_payback_filter = Q()
#         member_fees_filter = Q()
#         loan_payback_filter = Q()
#         loan_fee_filter = Q()
        
#         if date_from:
#             admin_fee_filter &= Q(date_deducted__gte=date_from)
#             saving_form_fee_filter &= Q(date_created__gte=date_from)
#             saving_filter &= Q(date_created__gte=date_from)
#             special_saving_filter &= Q(date_created__gte=date_from)
#             target_saving_filter &= Q(date_created__gte=date_from)
#             member_payback_filter &= Q(repayment_date__gte=date_from)
#             member_finance_payback_filter &= Q(amount_paid__gte=date_from)
#             member_fees_filter &= Q(created_at__gte=date_from)
#             loan_payback_filter &= Q(repayment_date__gte=date_from)
#             loan_fee_filter &= Q(created_at__gte=date_from)

#         if date_to:
#             admin_fee_filter &= Q(date_deducted__lte=date_to)
#             saving_filter &= Q(date_created__lte=date_to)
#             saving_form_fee_filter &= Q(date_created__lte=date_to)
#             special_saving_filter &= Q(date_created__lte=date_to)
#             target_saving_filter &= Q(date_created__lte=date_to)
#             member_payback_filter &= Q(repayment_date__lte=date_to)
#             member_finance_payback_filter &= Q(created_at__lte=date_to)
#             member_fees_filter &= Q(created_at__lte=date_to)
#             loan_payback_filter &= Q(repayment_date__lte=date_to)
#             loan_fee_filter &= Q(created_at__lte=date_to)

#         # 1. Income from saving items
#         try:
#             saving_income = Savings.objects.filter(
#                 saving_filter
#             ).aggregate(
#                 total=Sum('month_saving')
#             )['total'] or Decimal('0')
#         except Exception as e:
#             logger.error(f"Error calculating saving income: {str(e)}")
#             saving_income = Decimal('0')
            
#         # 1. Income from saving items
#         try:
#             saving_form_fee_income = SpecialSavingsTergetSavingsRequestForm.objects.filter(
#                 saving_form_fee_filter
#             ).aggregate(
#                 total=Sum('form_fee')
#             )['total'] or Decimal('0')
#         except Exception as e:
#             logger.error(f"Error calculating saving form fee income: {str(e)}")
#             saving_form_fee_income = Decimal('0')
            
#         # 2. Income from special saving items
#         try:
#             special_saving_income = SpecialSavings.objects.filter(
#                 special_saving_filter
#             ).aggregate(
#                 total=Sum('month_savings')
#             )['total'] or Decimal('0')
#         except Exception as e:
#             logger.error(f"Error calculating special saving income: {str(e)}")
#             special_saving_income = Decimal('0')
            
            
#         # 3. Income from target saving items
#         try:
#             target_saving_income = TargetSavings.objects.filter(
#                 target_saving_filter
#             ).aggregate(
#                 total=Sum('month_savings')
#             )['total'] or Decimal('0')
#         except Exception as e:
#             logger.error(f"Error calculating target saving income: {str(e)}")
#             target_saving_income = Decimal('0')

#         # 4. Income from Admin fee items
#         try:
#             admin_fee_income = Interest.objects.filter(
#                 admin_fee_filter
#             ).aggregate(
#                 total=Sum('amount_deducted')
#             )['total'] or Decimal('0')
#         except Exception as e:
#             logger.error(f"Error calculating admin fee income: {str(e)}")
#             admin_fee_income = Decimal('0')

#         # 5. Member repayments for consumables
#         try:
#             consumable_payback_income = PaybackConsumable.objects.filter(
#                 member_payback_filter
#             ).aggregate(
#                 total=Sum('amount_paid')
#             )['total'] or Decimal('0')
#         except Exception as e:
#             logger.error(f"Error calculating consumable payback income: {str(e)}")
#             consumable_payback_income = Decimal('0')

#         # 6. Member repayments for project finance
#         try:
#             finance_payback_income = ProjectFinancePayment.objects.filter(
#                 member_finance_payback_filter
#             ).aggregate(
#                 total=Sum('amount_paid')
#             )['total'] or Decimal('0')
#         except Exception as e:
#             logger.error(f"Error calculating finance payback income: {str(e)}")
#             finance_payback_income = Decimal('0')
#             print('finance_payback_income',finance_payback_income)
#         # 7. Income from consumable form fees
#         try:
#             form_fee_income = ConsumableFormFee.objects.filter(
#                 member_fees_filter
#             ).aggregate(
#                 total=Sum('form_fee')
#             )['total'] or Decimal('0')
#         except Exception as e:
#             logger.error(f"Error calculating form fee income: {str(e)}")
#             form_fee_income = Decimal('0')
            
#         # 8. Member repayments for loans
#         try:
#             loan_payback_income = LoanRepayback.objects.filter(
#                 loan_payback_filter
#             ).aggregate(
#                 total=Sum('amount_paid')
#             )['total'] or Decimal('0')
#         except Exception as e:
#             logger.error(f"Error calculating loan payback income: {str(e)}")
#             loan_payback_income = Decimal('0')

#         # 9. Income from loan form fees
#         try:
#             loan_form_fee_income = LoanRequestFee.objects.filter(
#                 loan_fee_filter
#             ).aggregate(
#                 total=Sum('form_fee')
#             )['total'] or Decimal('0')
#         except Exception as e:
#             logger.error(f"Error calculating loan form fee income: {str(e)}")
#             loan_form_fee_income = Decimal('0')

#         return {
#             'saving_income': saving_income,
#             'saving_form_fee_income': saving_form_fee_income,
#             'special_saving_income':special_saving_income,
#             'target_saving_income':target_saving_income,
#             'admin_fee_income': admin_fee_income,
#             'consumable_payback_income': consumable_payback_income,
#             'finance_payback_income': finance_payback_income,
#             'form_fee_income': form_fee_income,
#             'loan_payback_income': loan_payback_income,
#             'loan_form_fee_income': loan_form_fee_income,
#         }
        
#     except Exception as e:
#         logger.error(f"Error in calculate_total_income: {str(e)}", exc_info=True)
#         return {
#             'saving_income': Decimal('0'),
#             'saving_form_fee_income': Decimal('0'),
#             'special_saving_income': Decimal('0'),
#             'target_saving_income': Decimal('0'),
#             'admin_fee_income': Decimal('0'),
#             'consumable_payback_income': Decimal('0'),
#             'finance_payback_income': Decimal('0'),
#             'form_fee_income': Decimal('0'),
#             'loan_payback_income': Decimal('0'),
#             'loan_form_fee_income': Decimal('0'),
#         }





#======================= Loan part ====================



@login_required
def loan_payment_tracking(request):
    all_loans = LoanRequest.objects.filter(
        status__in=['approved', 'Fullpaid'],
        approved_amount__isnull=False
    )

    # Filter options
    years = all_loans.annotate(year=ExtractYear("application_date")).values_list("year", flat=True).distinct().order_by("-year")
    loan_types = LoanType.objects.filter(available=True).values_list("name", flat=True).order_by("name")
    loan_statuses = [('approved', 'Approved'), ('Fullpaid', 'Fully Paid')]

    # Get filters
    selected_year = request.GET.get("year", "").strip()
    selected_type = request.GET.get("loan_type", "").strip()
    selected_status = request.GET.get("status", "").strip()
    selected_member = request.GET.get("member", "").strip()

    # Apply filters
    queryset = all_loans
    if selected_year.isdigit():
        queryset = queryset.filter(application_date__year=int(selected_year))
    if selected_type:
        queryset = queryset.filter(loan_type__name=selected_type)
    if selected_status in ['approved', 'Fullpaid']:
        queryset = queryset.filter(status=selected_status)
    if selected_member:
        queryset = queryset.filter(
            Q(member__member__first_name__icontains=selected_member) |
            Q(member__member__last_name__icontains=selected_member) |
            Q(member__ippis__icontains=selected_member)
        )

    # Optimize queryset
    base_queryset = queryset.select_related(
        'member__member', 'loan_type', 'guarantor__member', 'created_by', 'approved_by'
    ).prefetch_related('repaybacks').order_by('-application_date')

    # Annotate each loan with total_paid
    annotated_loans = base_queryset.annotate(
        total_paid=Coalesce(Sum('repaybacks__amount_paid'), Decimal('0.00'))
    )

    # Summary stats
    total_approved = annotated_loans.aggregate(
        total=Coalesce(Sum('approved_amount'), Decimal('0.00'))
    )['total']
    total_paid = annotated_loans.aggregate(
        total=Coalesce(Sum('total_paid'), Decimal('0.00'))
    )['total']
    total_outstanding = total_approved - total_paid

    summary_stats = {
        'total_loans': annotated_loans.count(),
        'total_approved_amount': total_approved,
        'total_amount_paid': total_paid,
        'total_outstanding': total_outstanding
    }

    # Process loans for display
    processed_loans = []
    for loan in annotated_loans:
        approved_amount = loan.approved_amount or Decimal('0.00')
        total_paid = loan.total_paid or Decimal('0.00')
        balance_remaining = approved_amount - total_paid
        payment_percentage = (total_paid / approved_amount * 100) if approved_amount > 0 else 0

        if balance_remaining <= 0:
            payment_status, status_class = "Fully Paid", "success"
        elif total_paid > 0:
            payment_status, status_class = "Partial Payment", "warning"
        else:
            payment_status, status_class = "No Payment", "danger"

        last_payment = loan.repaybacks.order_by('-repayment_date', '-id').first()

        processed_loans.append({
            'loan': loan,
            'approved_amount': approved_amount,
            'total_paid': total_paid,
            'balance_remaining': balance_remaining,
            'payment_percentage': round(float(payment_percentage), 1),
            'payment_status': payment_status,
            'status_class': status_class,
            'payment_count': loan.repaybacks.count(),
            'last_payment': last_payment,
        })

    # ✅ Excel Export Logic
    if request.GET.get("download") == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Outstanding Payments"

        # Header row
        headers = ["ID", "Member", "Loan Type", "Date Created", "Approved Amount", "Total Paid", "Balance"]
        ws.append(headers)

        # Data rows
        for item in processed_loans:
            loan = item['loan']
            ws.append([
                loan.id,
                f"{loan.member.member.first_name} {loan.member.member.last_name}",
                loan.loan_type.name if loan.loan_type else "",
                loan.application_date.strftime("%Y-%m-%d") if loan.application_date else "",
                float(item['approved_amount']),
                float(item['total_paid']),
                float(item['balance_remaining']),
            ])

        # Adjust column widths
        for i, col in enumerate(ws.columns, start=1):
            max_length = 0
            column = get_column_letter(i)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 3

        # Return Excel response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="loan_outstanding_report.xlsx"'
        wb.save(response)
        return response

    # Recent payments
    recent_payments = LoanRepayback.objects.select_related(
        'loan_request__member__member', 'loan_request__loan_type'
    ).filter(loan_request__in=base_queryset).order_by('-repayment_date', '-id')[:20]

    # Pagination
    paginator = Paginator(processed_loans, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Final context
    context = {
        'page_obj': page_obj,
        'years': list(years),
        'loan_types': list(loan_types),
        'loan_statuses': loan_statuses,
        'selected_year': selected_year,
        'selected_type': selected_type,
        'selected_status': selected_status,
        'selected_member': selected_member,
        'summary_stats': summary_stats,
        'recent_payments': recent_payments,
    }

    return render(request, "reports/loan_payment_tracking.html", context)

@login_required
def loan_payment_detail(request, loan_id):
    """Loan payment detail view"""
    loan = get_object_or_404(
        LoanRequest.objects.select_related(
            'member__member', 'loan_type', 'guarantor__member', 
            'created_by', 'approved_by'
        ),
        id=loan_id,
        status__in=['approved', 'Fullpaid']
    )
    
    # Get all payments for this loan
    payments = loan.repaybacks.all().order_by('repayment_date', 'id')
    
    # Calculate running totals
    running_total = Decimal('0.00')
    payment_history = []
    
    for payment in payments:
        running_total += payment.amount_paid
        approved_amount = loan.approved_amount or Decimal('0.00')
        remaining_balance = approved_amount - running_total
        
        payment_history.append({
            'payment': payment, 
            'running_total': running_total, 
            'remaining_balance': remaining_balance
        })

    # Calculate loan metrics
    approved_amount = loan.approved_amount or Decimal('0.00')
    total_paid = Decimal(str(loan.total_repaid)) if loan.total_repaid else Decimal('0.00')
    balance_remaining = approved_amount - total_paid
    
    if approved_amount > 0:
        payment_percentage = (total_paid / approved_amount * 100)
    else:
        payment_percentage = 0
    
    context = {
        'loan': loan, 
        'payment_history': payment_history,
        'approved_amount': approved_amount,
        'total_paid': total_paid,
        'balance_remaining': balance_remaining,
        'payment_percentage': round(float(payment_percentage), 1),
        'expected_monthly': loan.monthly_payment or Decimal('0.00'),
    }
    
    return render(request, "reports/loan_payment_detail.html", context)



