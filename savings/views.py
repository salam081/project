from urllib import request
from django.shortcuts import render
from django.http import HttpResponse
from django.utils.timezone import now
import pandas as pd
from decimal import Decimal, InvalidOperation
from decimal import Decimal,DecimalException
from django.db import transaction
from datetime import datetime
from django.core.paginator import Paginator
from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth,Coalesce

from django.shortcuts import render
from .models import Savings
import openpyxl
import io
import os
import pandas as pd
import calendar
from django.shortcuts import render, redirect,get_object_or_404
from django.shortcuts import render, get_object_or_404, redirect

from django.db import transaction
from django.core.paginator import Paginator
from django.db.models import Q, Prefetch,Value
from decimal import Decimal
from django.db import transaction
from datetime import timedelta
from django.db.models import Q, Sum,DecimalField

from django.utils import timezone
from django.conf import settings
from django.contrib import messages
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.db.models import Sum,F
from django.db.models.functions import ExtractMonth, ExtractYear
from django.utils.dateparse import parse_date
import pandas as pd
from decimal import Decimal
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.db import transaction
from .models import *
from accounts.models import *
from accounts.decorators import *

@login_required
@group_required(['admin','staff'])
def search_member_for_savings(request):
    groups = UserGroup.objects.all().order_by('title')
    results = []
    search_term = request.GET.get('search_term', '').strip()
    if search_term:
        results = Member.objects.select_related('member').filter(
            Q(member__first_name__icontains=search_term) |
            Q(member__last_name__icontains=search_term) |
            Q(ippis__icontains=search_term) |
            Q(id__icontains=search_term)
        ).order_by('member__first_name', 'member__last_name')

        paginator = Paginator(results, 100)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
    else:
        page_obj = []

    context = {
        'results': page_obj,
        'search_term': search_term,
        'groups':groups
    }
    return render(request, 'saving/search_member.html', context)

def filter_requests(datefrom, dateto):
    filtered_requests = Savings.objects.all().order_by('-date_created')
    
    if datefrom:
        filtered_requests = filtered_requests.filter(month__gte=datefrom)
    if dateto:
        filtered_requests = filtered_requests.filter(month__lte=dateto)
    return filtered_requests


@login_required
@group_required(['admin','staff'])
def add_member_savings(request):
    if request.method == "POST":
        ippis = request.POST.get("ippis")
        month_str = request.POST.get("month")
        month_saving_str = request.POST.get("month_saving")

        # Validate IPPIS and inputs
        if not ippis:
            messages.error(request, "⚠️ Please provide the member IPPIS.")
            return redirect("add_member_savings")

        if not month_str or not month_saving_str:
            messages.error(request, "⚠️ Please provide both the month and the saving amount.")
            return redirect("add_member_savings")

        try:
            # Find member by IPPIS
            member = Member.objects.get(ippis=ippis)
        except Member.DoesNotExist:
            messages.error(request, f"No member found with IPPIS {ippis}.")
            return redirect("add_member_savings")

        try:
            # Parse date and amount
            month = timezone.datetime.strptime(month_str, "%Y-%m-%d").date()
            month_saving = Decimal(month_saving_str)

            # Prevent duplicate entries
            if Savings.objects.filter(member=member, month__year=month.year, month__month=month.month).exists():
                messages.warning(
                    request,
                    f"Savings for **{member.member}** in {month.strftime('%B %Y')} already exist."
                )
                return redirect("add_member_savings")

            # Get global subscription fee
            try:
                global_interest = InterestAmount.objects.latest("date_created").amount
            except InterestAmount.DoesNotExist:
                messages.error(request, "No Subscription amount has been set. Please set one first.")
                return redirect("add_member_savings")

            # Check validity
            if month_saving <= global_interest:
                messages.error(
                    request,
                    f"Savings must be greater than the subscription amount (₦{global_interest:,.2f})."
                )
                return redirect("add_member_savings")

            # Deduct subscription
            amount_after_interest = month_saving - global_interest
            half_amount = amount_after_interest / 2

            # Create savings record
            Savings.objects.create(
                member=member,
                month=month,
                month_saving=amount_after_interest,
                original_amount=month_saving
            )

            # Deduct subscription once
            Interest.objects.create(member=member, month=month, amount_deducted=global_interest)

            # Update Loanable
            if not Loanable.objects.filter(member=member, month=month).exists():
                current_loanable_total = Loanable.objects.filter(member=member).aggregate(
                    total=Sum("amount")
                )["total"] or Decimal("0.00")
                Loanable.objects.create(
                    member=member,
                    month=month,
                    amount=half_amount,
                    total_amount=current_loanable_total + half_amount
                )

            # Update Investment
            if not Investment.objects.filter(member=member, month=month).exists():
                current_investment_total = Investment.objects.filter(member=member).aggregate(
                    total=Sum("amount")
                )["total"] or Decimal("0.00")
                Investment.objects.create(
                    member=member,
                    month=month,
                    amount=half_amount,
                    total_amount=current_investment_total + half_amount
                )

            # Success message
            messages.success(
                request,
                f"Savings of ₦{month_saving:,.2f} added for **{member.member}** "
                f"({month.strftime('%B %Y')}). Subscription deducted: ₦{global_interest:,.2f}. "
                f"Loanable: ₦{half_amount:,.2f}, Investment: ₦{half_amount:,.2f}."
            )

            return redirect("add_member_savings")

        except (ValueError, DecimalException):
            messages.error(request, "Invalid date format or saving amount.")
        except Exception as e:
            messages.error(request, f"Unexpected error: {e}")

    return render(request, "saving/add_member_savings.html")




@login_required
@group_required(['admin'])
def process_member_savings(request, id):
    member = get_object_or_404(Member, id=id)

    if request.method == "POST":
        month_str = request.POST.get("month")
        month_saving_str = request.POST.get("month_saving")

        # Validate inputs
        if not month_str or not month_saving_str:
            messages.error(request, "⚠️ Please provide both the month and the saving amount.")
            return redirect("add_individual_savings", id=id)

        try:
            # Parse month and amount
            month = timezone.datetime.strptime(month_str, "%Y-%m-%d").date()
            month_saving = Decimal(month_saving_str)
            

            if Savings.objects.filter(member=member,month__year=month.year,month__month=month.month).exists():
                messages.warning( request,f" Savings for **{member.member}** in {month.strftime('%B %Y')} already exist.")
                return redirect("add_individual_savings", id=id)
           
            try:
                global_interest = InterestAmount.objects.latest("date_created").amount
            except InterestAmount.DoesNotExist:
                messages.error(request, " No Subscription amount has been set. Please set one first.")
                return redirect("add_individual_savings", id=id)

            #  Check if saving amount is valid
            if month_saving <= global_interest:
                messages.error( request, f" Savings must be greater than the interest amount (₦{global_interest:,.2f})." )
                return redirect("add_individual_savings", id=id)

            #  Calculate amount after interest
            amount_after_interest = month_saving - global_interest
            half_amount = amount_after_interest / 2

            #  Create Savings record with NET savings
            savings_record = Savings.objects.create(
                member=member,
                month=month,
                month_saving=amount_after_interest,   
                original_amount=month_saving         
            )

            #  Deduct Interest (only once)
            Interest.objects.create( member=member, month=month,amount_deducted=global_interest)

            #  Distribute Remaining Amount into Loanable and Investment (if not already distributed)
            if not Loanable.objects.filter(member=member, month=month).exists():
                current_loanable_total = Loanable.objects.filter(member=member).aggregate(
                    total=Sum("amount")
                )["total"] or Decimal("0.00")

                Loanable.objects.create( member=member, month=month, amount=half_amount,
                    total_amount=current_loanable_total + half_amount )

            if not Investment.objects.filter(member=member, month=month).exists():
                current_investment_total = Investment.objects.filter(member=member).aggregate(
                    total=Sum("amount")
                )["total"] or Decimal("0.00")

                Investment.objects.create( member=member, month=month, amount=half_amount,
                    total_amount=current_investment_total + half_amount)

            #  Success Message
            messages.success(
                request,
                f" Savings of ₦{month_saving:,.2f} added for **{member.member}** "
                f"({month.strftime('%B %Y')}). subscription deducted: ₦{global_interest:,.2f}. "
                f"Loanable: ₦{half_amount:,.2f}, Investment: ₦{half_amount:,.2f}."
            )

            return redirect("add_individual_savings", id=id)

        except (ValueError, DecimalException):
            messages.error(request, " Invalid date format or saving amount.")
        except Exception as e:
            messages.error(request, f" Unexpected error: {e}")

    context = {"member": member}
    return render(request, "saving/add_individual_savings.html", context)

@login_required
@group_required(['admin','staff'])
@transaction.atomic
def upload_savings(request):
    if request.method == "POST" and request.FILES.get("file"):
        try:
            # Get selected month
            selected_month = request.POST.get("month")
            if not selected_month:
                messages.error(request, "Please select a month.")
                return redirect("upload_savings")

            month = pd.to_datetime(selected_month).date().replace(day=1)
            file = request.FILES["file"]

            # Read Excel or CSV
            if file.name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file, dtype={'IPPIS': str})
            else:
                df = pd.read_csv(file, dtype={'IPPIS': str})

            # Validate required columns
            required_columns = ["IPPIS", "Amount"]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                messages.error(
                    request,
                    f"Missing columns: {', '.join(missing_columns)}. "
                    f"Found columns: {', '.join(df.columns.tolist())}"
                )
                return redirect("upload_savings")

            # Clean data
            df = df.dropna(subset=['IPPIS', 'Amount'])
            df["IPPIS"] = df["IPPIS"].astype(str).str.strip()
            df["Amount"] = pd.to_numeric(df["Amount"], errors='coerce')
            df = df[(df["Amount"] > 0) & df["Amount"].notna()]

            if len(df) == 0:
                messages.error(request, "No valid data found in file.")
                return redirect("upload_savings")
            
            file_ippis = df["IPPIS"].tolist()
            members_qs = Member.objects.filter(ippis__in=file_ippis).only("id", "ippis", "member")
            members_dict = {str(m.ippis).strip(): m for m in members_qs}

            existing_savings_ids = set(
                Savings.objects.filter(month=month, member__ippis__in=file_ippis)
                .values_list("member_id", flat=True)
            )

            try:
                global_interest = InterestAmount.objects.latest("date_created").amount
            except InterestAmount.DoesNotExist:
                messages.error(request, "No Subscription amount has been set. Please add one first.")
                return redirect("upload_savings")

            # Prepare bulk insert lists
            savings_to_create = []
            interests_to_create = []
            loanables_to_create = []
            investments_to_create = []
            savings_member_ids_added = set()
            interest_member_ids_added = set()
            skipped_members_report = []

            records = df.to_dict("records")
            for row in records:
                ippis = str(row["IPPIS"]).strip()
                amount = Decimal(str(row["Amount"]))
           
                # Member lookup
                member = members_dict.get(ippis)
                if not member:
                    skipped_members_report.append({
                        "IPPIS": ippis,
                        "Member Name": "N/A",
                        "Reason": "IPPIS not found"
                    })
                    continue
                
                # ✅ Skip if linked user is missing or deactivated
                if not member.member or not member.member.is_active:
                    skipped_members_report.append({
                        "IPPIS": ippis,
                        "Member Name": str(member.member) if member.member else "No linked user",
                        "Reason": "User account is deactivated"
                    })
                    continue

                # Skip if already saved
                if member.id in existing_savings_ids:
                    skipped_members_report.append({
                        "IPPIS": ippis,
                        "Member Name": member.member,
                        "Reason": "Savings already exist"
                    })
                    continue

                # Deduct interest only once per member
                final_amount = max(amount - global_interest, Decimal("0.00"))
                if member.id not in interest_member_ids_added:
                    interests_to_create.append(
                        Interest(member=member, month=month, amount_deducted=global_interest)
                    )
                    interest_member_ids_added.add(member.id)

                # Prepare savings record
                savings_to_create.append(
                    Savings(member=member, month=month, month_saving=final_amount, original_amount=amount)
                )
                savings_member_ids_added.add(member.id)

                # Split into loanable & investment
                half_amount = (final_amount / 2).quantize(Decimal("0.01"))

                # Get current totals before adding new records
                current_loanable_total = Loanable.objects.filter(member=member).aggregate(
                    total=Sum("amount")
                )["total"] or Decimal("0.00")

                current_investment_total = Investment.objects.filter(member=member).aggregate(
                    total=Sum("amount")
                )["total"] or Decimal("0.00")

                # Create updated loanable record
                loanables_to_create.append(
                    Loanable(
                        member=member,
                        month=month,
                        amount=half_amount,
                        total_amount=current_loanable_total + half_amount
                    )
                )

                # Create updated investment record
                investments_to_create.append(
                    Investment(
                        member=member,
                        month=month,
                        amount=half_amount,
                        total_amount=current_investment_total + half_amount
                    )
                )

            # Bulk insert
            with transaction.atomic():
                if interests_to_create:
                    Interest.objects.bulk_create(interests_to_create, batch_size=1000)

                if savings_to_create:
                    Savings.objects.bulk_create(savings_to_create, batch_size=1000)

                    #  Efficiently update totals using bulk_update
                    totals = (
                        Savings.objects.filter(member__in=savings_member_ids_added)
                        .values("member")
                        .annotate(total=Sum("month_saving"))
                    )

                    members_to_update = [
                        Member(id=row["member"], total_savings=row["total"])
                        for row in totals
                    ]
                    if members_to_update:
                        Member.objects.bulk_update(members_to_update, ["total_savings"])

                if loanables_to_create:
                    Loanable.objects.bulk_create(loanables_to_create, batch_size=1000)
                if investments_to_create:
                    Investment.objects.bulk_create(investments_to_create, batch_size=1000)

            created_count = len(savings_to_create)

            # If skipped members exist, generate downloadable Excel
            if skipped_members_report:
                skipped_df = pd.DataFrame(skipped_members_report)
                response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                response['Content-Disposition'] = f'attachment; filename="skipped_members_{month}.xlsx"'
                skipped_df.to_excel(response, index=False)
                messages.warning(request, f"⚠️ {len(skipped_members_report)} members skipped. Download report for details.")
                return response

            total_uploaded_amount = sum([s.month_saving for s in savings_to_create])
            total_interest_deducted = len(interests_to_create) * global_interest
            total_remaining_balance = total_uploaded_amount - total_interest_deducted
            half_amount = (total_remaining_balance / 2).quantize(Decimal("0.01"))

            messages.success(
                request,
                f"✅ Uploaded {created_count} records!<br>"
                f"💰 Total Uploaded: <b>₦{total_uploaded_amount:,.2f}</b><br>"
                f"📌 Subscription Deducted: <b>₦{total_interest_deducted:,.2f}</b><br>"
                f"🏦 Loanable: <b>₦{half_amount:,.2f}</b> | 📈 Investment: <b>₦{half_amount:,.2f}</b>"
            )
            return redirect("upload_savings")

        except Exception as e:
            transaction.set_rollback(True)
            messages.error(request, f"Error processing file: {str(e)}")
            return redirect("upload_savings")

    return render(request, "saving/upload_savings.html")

@login_required
@group_required(['admin','staff'])
@transaction.atomic
def edit_saving(request, saving_id):
    saving = get_object_or_404(Savings, id=saving_id)
    member = saving.member
    month = saving.month

    if request.method == "POST":
        try:
            # Get the new amount from the form
            new_amount = Decimal(request.POST.get("month_saving", saving.month_saving))

            # Get global interest amount
            try:
                global_interest = InterestAmount.objects.latest("date_created").amount
            except InterestAmount.DoesNotExist:
                messages.error(request, "No Subscription amount has been set. Please add one first.")
                return redirect("upload_savings")

            # Check if interest was already deducted for this member/month
            interest_record = Interest.objects.filter(member=member, month=month).first()
            if not interest_record:
                # Deduct interest for the first time
                amount_after_interest = max(new_amount - global_interest, Decimal("0.00"))
                Interest.objects.create(
                    member=member,
                    month=month,
                    amount_deducted=global_interest
                )
            else:
                # If already deducted, don't deduct again
                amount_after_interest = new_amount

            # Update savings amount
            saving.month_saving = amount_after_interest
            saving.original_amount = new_amount
            saving.save()

            # Update loanable and investment amounts
            loanable_amount = (amount_after_interest / 2).quantize(Decimal("0.01"))
            investment_amount = (amount_after_interest / 2).quantize(Decimal("0.01"))

            # Update or create loanable record
            Loanable.objects.update_or_create(
                member=member, month=month,
                defaults={"amount": loanable_amount, "total_amount": loanable_amount}
            )

            # Update or create investment record
            Investment.objects.update_or_create(
                member=member, month=month,
                defaults={"amount": investment_amount, "total_amount": investment_amount}
            )

            messages.success(
                request,
                f"Savings updated successfully for {member.member.first_name} {member.member.last_name} ({member.ippis})!"
            )

            return redirect("list_savings")

        except Exception as e:
            transaction.set_rollback(True)
            messages.error(request, f"Error updating savings: {str(e)}")
            return redirect("list_savings")

    return render(request, "saving/edit_saving.html", {"saving": saving})

from django.db.models import OuterRef, Subquery, DecimalField
from django.db.models.functions import Coalesce

@login_required
@group_required(['admin','staff'])
def member_savings_summary(request):
    groups = UserGroup.objects.all()
    search_name = request.GET.get("name", "").strip()
    search_ippis = request.GET.get("ippis", "").strip()
    search_group = request.GET.get("group", "").strip()
    selected_month = request.GET.get("month", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()
    per_page = request.GET.get("per_page", "100")
    export = request.GET.get("export", "")

    try:
        per_page = int(per_page)
        if per_page not in [100, 150, 200, 250]:
            per_page = 100
    except (ValueError, TypeError):
        per_page = 100

    # Base queryset
    members = Member.objects.select_related("member")

    # Filters
    if search_name:
        members = members.filter(
            Q(member__first_name__icontains=search_name) |
            Q(member__last_name__icontains=search_name)
        )
    if search_ippis:
        members = members.filter(ippis__icontains=search_ippis)
    if search_group.isdigit():
        members = members.filter(member__group_id=search_group)

    # ✅ Parse month filter
    month_filter = {}
    month_label = ""
    if selected_month:
        try:
            year, month_num = selected_month.split("-")
            month_filter = {"year": int(year), "month": int(month_num)}
            from calendar import month_name
            month_label = f"{month_name[int(month_num)]} {year}"
        except (ValueError, IndexError):
            pass

    # ✅ Parse date range filter
    date_range = {}
    date_range_label = ""
    start_date = None
    end_date = None
    if date_from:
        start_date = parse_date(date_from)
        if start_date:
            date_range["start"] = start_date
    if date_to:
        end_date = parse_date(date_to)
        if end_date:
            date_range["end"] = end_date

    if start_date and end_date:
        date_range_label = f"{start_date.strftime('%d %b %Y')} – {end_date.strftime('%d %b %Y')}"
    elif start_date:
        date_range_label = f"From {start_date.strftime('%d %b %Y')}"
    elif end_date:
        date_range_label = f"Up to {end_date.strftime('%d %b %Y')}"

    # ✅ Helper to build date filter kwargs for a given field name
    def build_date_filter(field, use_month=False, use_range=False):
        kwargs = {}
        if use_month and month_filter:
            kwargs[f"{field}__year"] = month_filter["year"]
            kwargs[f"{field}__month"] = month_filter["month"]
        elif use_range and date_range:
            if "start" in date_range and "end" in date_range:
                kwargs[f"{field}__range"] = [date_range["start"], date_range["end"]]
            elif "start" in date_range:
                kwargs[f"{field}__gte"] = date_range["start"]
            elif "end" in date_range:
                kwargs[f"{field}__lte"] = date_range["end"]
        return kwargs

    # ✅ All-time subqueries
    savings_subquery = Savings.objects.filter(
        member=OuterRef("pk")
    ).values("member").annotate(total=Sum("month_saving")).values("total")

    loanable_subquery = Loanable.objects.filter(
        member=OuterRef("pk")
    ).values("member").annotate(total=Sum("amount")).values("total")

    investment_subquery = Investment.objects.filter(
        member=OuterRef("pk")
    ).values("member").annotate(total=Sum("amount")).values("total")

    # ✅ Period subqueries — month filter takes priority over date range
    use_month = bool(month_filter)
    use_range = bool(date_range) and not use_month

    period_savings_sq = period_loanable_sq = period_investment_sq = None

    if use_month or use_range:
        period_savings_sq = Savings.objects.filter(
            member=OuterRef("pk"),
            **build_date_filter("month", use_month=use_month, use_range=use_range)
        ).values("member").annotate(total=Sum("month_saving")).values("total")

        period_loanable_sq = Loanable.objects.filter(
            member=OuterRef("pk"),
            **build_date_filter("month", use_month=use_month, use_range=use_range)
        ).values("member").annotate(total=Sum("amount")).values("total")

        period_investment_sq = Investment.objects.filter(
            member=OuterRef("pk"),
            **build_date_filter("month", use_month=use_month, use_range=use_range)
        ).values("member").annotate(total=Sum("amount")).values("total")

    # Build annotation dict
    annotation_kwargs = {
        "agg_savings": Coalesce(
            Subquery(savings_subquery, output_field=DecimalField()), Decimal("0.00")
        ),
        "agg_loanable": Coalesce(
            Subquery(loanable_subquery, output_field=DecimalField()), Decimal("0.00")
        ),
        "agg_investment": Coalesce(
            Subquery(investment_subquery, output_field=DecimalField()), Decimal("0.00")
        ),
    }

    if period_savings_sq is not None:
        annotation_kwargs["period_savings"] = Coalesce(
            Subquery(period_savings_sq, output_field=DecimalField()), Decimal("0.00")
        )
        annotation_kwargs["period_loanable"] = Coalesce(
            Subquery(period_loanable_sq, output_field=DecimalField()), Decimal("0.00")
        )
        annotation_kwargs["period_investment"] = Coalesce(
            Subquery(period_investment_sq, output_field=DecimalField()), Decimal("0.00")
        )

    members = members.annotate(**annotation_kwargs).order_by("member__first_name")

    # Grand totals
    grand_agg = {
        "grand_savings": Coalesce(Sum("agg_savings"), Decimal("0.00"), output_field=DecimalField()),
        "grand_loanable": Coalesce(Sum("agg_loanable"), Decimal("0.00"), output_field=DecimalField()),
        "grand_investment": Coalesce(Sum("agg_investment"), Decimal("0.00"), output_field=DecimalField()),
    }
    if period_savings_sq is not None:
        grand_agg["grand_period_savings"] = Coalesce(Sum("period_savings"), Decimal("0.00"), output_field=DecimalField())
        grand_agg["grand_period_loanable"] = Coalesce(Sum("period_loanable"), Decimal("0.00"), output_field=DecimalField())
        grand_agg["grand_period_investment"] = Coalesce(Sum("period_investment"), Decimal("0.00"), output_field=DecimalField())

    grand_totals = members.aggregate(**grand_agg)

    show_period = period_savings_sq is not None
    period_label = month_label or date_range_label

    # ✅ Excel export
    if export == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Savings Summary"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
        total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        center = Alignment(horizontal="center")
        right = Alignment(horizontal="right")

        title = f"Member Savings Summary{' — ' + period_label if period_label else ''}"
        ws.merge_cells("A1:I1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].alignment = center

        headers = ["#", "Member", "IPPIS", "Total Savings", "Total Loanable", "Total Investment"]
        if show_period:
            headers += [
                f"Savings ({period_label})",
                f"Loanable ({period_label})",
                f"Investment ({period_label})",
            ]

        header_row = 3
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        all_members = members.all()
        for idx, member in enumerate(all_members, start=1):
            row = header_row + idx
            row_data = [
                idx,
                f"{member.member.first_name} {member.member.last_name}",
                member.ippis,
                float(member.agg_savings),
                float(member.agg_loanable),
                float(member.agg_investment),
            ]
            if show_period:
                row_data += [
                    float(member.period_savings),
                    float(member.period_loanable),
                    float(member.period_investment),
                ]
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                if col >= 4:
                    cell.number_format = '#,##0.00'
                    cell.alignment = right

        total_row = header_row + len(list(all_members)) + 1
        ws.cell(row=total_row, column=1, value="GRAND TOTALS").font = Font(bold=True)
        ws.merge_cells(f"A{total_row}:C{total_row}")

        totals = [
            float(grand_totals["grand_savings"]),
            float(grand_totals["grand_loanable"]),
            float(grand_totals["grand_investment"]),
        ]
        if show_period:
            totals += [
                float(grand_totals["grand_period_savings"]),
                float(grand_totals["grand_period_loanable"]),
                float(grand_totals["grand_period_investment"]),
            ]
        for col, value in enumerate(totals, start=4):
            cell = ws.cell(row=total_row, column=col, value=value)
            cell.font = Font(bold=True)
            cell.fill = total_fill
            cell.number_format = '#,##0.00'
            cell.alignment = right

        col_widths = [5, 30, 15, 18, 18, 18, 25, 25, 25]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        filename = f"savings_summary{'_' + selected_month if selected_month else ''}{'_' + date_from + '_' + date_to if date_from or date_to else ''}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    # Pagination
    paginator = Paginator(members, per_page)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "groups": groups,
        "search_name": search_name,
        "search_ippis": search_ippis,
        "search_group": search_group,
        "selected_month": selected_month,
        "date_from": date_from,
        "date_to": date_to,
        "month_label": month_label,
        "period_label": period_label,
        "date_range_label": date_range_label,
        "per_page": per_page,
        "total_records": paginator.count,
        "grand_totals": grand_totals,
        "show_period": show_period,
        "pagination_info": {
            "current_page": page_obj.number,
            "total_pages": paginator.num_pages,
            "has_previous": page_obj.has_previous(),
            "has_next": page_obj.has_next(),
            "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
            "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
        },
    }

    return render(request, "saving/member_savings_summary.html", context)

@login_required
@group_required(['admin','staff'])
def list_savings(request):
    groups = UserGroup.objects.all()
    selected_month = request.GET.get("month")
    search_name = request.GET.get("name", "").strip()
    search_ippis = request.GET.get("ippis", "").strip()
    search_group = request.GET.get("group", "").strip()
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")
    per_page = request.GET.get("per_page", "25")

    # Validate per_page value
    try:
        per_page = int(per_page)
        if per_page not in [100, 150, 200, 250]:
            per_page = 100
    except (ValueError, TypeError):
        per_page = 100

    # Base queryset
    savings = Savings.objects.select_related("member__member")

    # Parse month filter
    month_filter = {}
    if selected_month:
        try:
            year, month_num = selected_month.split("-")
            month_filter = {"month__year": year, "month__month": month_num}
            savings = savings.filter(**month_filter)
        except (ValueError, IndexError):
            pass

    # Apply date range filter
    if date_from and date_to:
        start_date = parse_date(date_from)
        end_date = parse_date(date_to)
        if start_date and end_date:
            savings = savings.filter(month__range=[start_date, end_date])
    elif date_from:
        start_date = parse_date(date_from)
        if start_date:
            savings = savings.filter(month__gte=start_date)
    elif date_to:
        end_date = parse_date(date_to)
        if end_date:
            savings = savings.filter(month__lte=end_date)

    # Filter by member name
    if search_name:
        savings = savings.filter(
            Q(member__member__first_name__icontains=search_name) |
            Q(member__member__last_name__icontains=search_name)
        )

    # Filter by IPPIS
    if search_ippis:
        savings = savings.filter(member__ippis__icontains=search_ippis)

    # Filter by group
    if search_group.isdigit():
        savings = savings.filter(member__member__group_id=search_group)

    # Compute total savings for current search/filter
    total_savings_amount = (
        savings.aggregate(total=Sum("month_saving"))["total"] or Decimal("0.00")
    )

    # ✅ Compute total savings per member (all-time, unfiltered by date)
    # Get the distinct member IDs from the current filtered queryset
    filtered_member_ids = savings.values_list("member_id", flat=True).distinct()

    member_total_savings = (
        Savings.objects.filter(member_id__in=filtered_member_ids)
        .values("member_id")
        .annotate(total_savings=Sum("month_saving"))
    )

    # Build a dict: { member_id: total_savings }
    member_total_savings_dict = {
        entry["member_id"]: entry["total_savings"] or Decimal("0.00")
        for entry in member_total_savings
    }

    # Order results
    savings = savings.order_by("-id", "member__member__first_name")

    # Pagination
    paginator = Paginator(savings, per_page)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Get member IDs from current page only
    current_page_member_ids = [saving.member_id for saving in page_obj.object_list]

    # Build dictionaries for loanable & investment amounts
    loanable_dict = {}
    investment_dict = {}

    if current_page_member_ids:
        loanable_filter = {"member_id__in": current_page_member_ids}
        investment_filter = {"member_id__in": current_page_member_ids}

        # Add date filters
        if month_filter:
            loanable_filter.update(month_filter)
            investment_filter.update(month_filter)
        elif date_from and date_to:
            loanable_filter["month__range"] = [start_date, end_date]
            investment_filter["month__range"] = [start_date, end_date]
        elif date_from:
            loanable_filter["month__gte"] = start_date
            investment_filter["month__gte"] = start_date
        elif date_to:
            loanable_filter["month__lte"] = end_date
            investment_filter["month__lte"] = end_date

        loanables = Loanable.objects.filter(**loanable_filter).values("member_id", "amount", "month")
        for loanable in loanables:
            loanable_dict[(loanable["member_id"], loanable["month"])] = loanable["amount"]

        investments = Investment.objects.filter(**investment_filter).values("member_id", "amount", "month")
        for investment in investments:
            investment_dict[(investment["member_id"], investment["month"])] = investment["amount"]

    # Assign amounts
    for saving in page_obj.object_list:
        lookup_key = (saving.member_id, saving.month)
        saving.loanable_amount = loanable_dict.get(lookup_key, Decimal("0.00"))
        saving.investment_amount = investment_dict.get(lookup_key, Decimal("0.00"))
        # ✅ Attach each member's all-time total savings directly to the saving object
        saving.member_total_savings = member_total_savings_dict.get(saving.member_id, Decimal("0.00"))

    context = {
        "page_obj": page_obj,
        "groups": groups,
        "selected_month": selected_month,
        "search_name": search_name,
        "search_ippis": search_ippis,
        "search_group": search_group,
        "per_page": per_page,
        "date_from": date_from,
        "date_to": date_to,
        "total_records": paginator.count,
        "total_savings_amount": total_savings_amount,
        "pagination_info": {
            "current_page": page_obj.number,
            "total_pages": paginator.num_pages,
            "has_previous": page_obj.has_previous(),
            "has_next": page_obj.has_next(),
            "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
            "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
        },
    }

    return render(request, "saving/list_savings.html", context)
# def list_savings(request):
#     groups = UserGroup.objects.all()
#     selected_month = request.GET.get("month")
#     search_name = request.GET.get("name", "").strip()
#     search_ippis = request.GET.get("ippis", "").strip()
#     search_group = request.GET.get("group", "").strip()
#     date_from = request.GET.get("date_from")
#     date_to = request.GET.get("date_to")
#     per_page = request.GET.get("per_page", "25")

#     # Validate per_page value
#     try:
#         per_page = int(per_page)
#         if per_page not in [10, 25, 50, 100]:
#             per_page = 25
#     except (ValueError, TypeError):
#         per_page = 25

#     # Base queryset
#     savings = Savings.objects.select_related("member__member")

#     # Parse month filter
#     month_filter = {}
#     if selected_month:
#         try:
#             year, month_num = selected_month.split("-")
#             month_filter = {"month__year": year, "month__month": month_num}
#             savings = savings.filter(**month_filter)
#         except (ValueError, IndexError):
#             pass

#     # Apply date range filter
#     if date_from and date_to:
#         start_date = parse_date(date_from)
#         end_date = parse_date(date_to)
#         if start_date and end_date:
#             savings = savings.filter(month__range=[start_date, end_date])
#     elif date_from:
#         start_date = parse_date(date_from)
#         if start_date:
#             savings = savings.filter(month__gte=start_date)
#     elif date_to:
#         end_date = parse_date(date_to)
#         if end_date:
#             savings = savings.filter(month__lte=end_date)

#     # Filter by member name
#     if search_name:
#         savings = savings.filter(
#             Q(member__member__first_name__icontains=search_name) |
#             Q(member__member__last_name__icontains=search_name)
#         )

#     # Filter by IPPIS
#     if search_ippis:
#         savings = savings.filter(member__ippis__icontains=search_ippis)

#     # Filter by group
#     if search_group.isdigit():
#         savings = savings.filter(member__member__group_id=search_group)

#     # ✅ Compute total savings for current search/filter
#     total_savings_amount = (
#         savings.aggregate(total=Sum("month_saving"))["total"] or Decimal("0.00")
#     )
    
#     # Order results
#     savings = savings.order_by("-id", "member__member__first_name")

#     # Pagination
#     paginator = Paginator(savings, per_page)
#     page_number = request.GET.get("page")
#     page_obj = paginator.get_page(page_number)

#     # Get member IDs from current page only
#     current_page_member_ids = [saving.member_id for saving in page_obj.object_list]

#     # Build dictionaries for loanable & investment amounts
#     loanable_dict = {}
#     investment_dict = {}

#     if current_page_member_ids:
#         loanable_filter = {"member_id__in": current_page_member_ids}
#         investment_filter = {"member_id__in": current_page_member_ids}

#         # Add date filters
#         if month_filter:
#             loanable_filter.update(month_filter)
#             investment_filter.update(month_filter)
#         elif date_from and date_to:
#             loanable_filter["month__range"] = [start_date, end_date]
#             investment_filter["month__range"] = [start_date, end_date]
#         elif date_from:
#             loanable_filter["month__gte"] = start_date
#             investment_filter["month__gte"] = start_date
#         elif date_to:
#             loanable_filter["month__lte"] = end_date
#             investment_filter["month__lte"] = end_date

#         loanables = Loanable.objects.filter(**loanable_filter).values("member_id", "amount", "month")
#         for loanable in loanables:
#             loanable_dict[(loanable["member_id"], loanable["month"])] = loanable["amount"]

#         investments = Investment.objects.filter(**investment_filter).values("member_id", "amount", "month")
#         for investment in investments:
#             investment_dict[(investment["member_id"], investment["month"])] = investment["amount"]

#     # Assign amounts
#     for saving in page_obj.object_list:
#         lookup_key = (saving.member_id, saving.month)
#         saving.loanable_amount = loanable_dict.get(lookup_key, Decimal("0.00"))
#         saving.investment_amount = investment_dict.get(lookup_key, Decimal("0.00"))

#     context = {
#         "page_obj": page_obj,
#         "groups": groups,
#         "selected_month": selected_month,
#         "search_name": search_name,
#         "search_ippis": search_ippis,
#         "search_group": search_group,
#         "per_page": per_page,
#         "date_from": date_from,
#         "date_to": date_to,
#         "total_records": paginator.count,
#         "total_savings_amount": total_savings_amount, 
#         "pagination_info": {
#             "current_page": page_obj.number,
#             "total_pages": paginator.num_pages,
#             "has_previous": page_obj.has_previous(),
#             "has_next": page_obj.has_next(),
#             "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
#             "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
#         },
#     }

#     return render(request, "saving/list_savings.html", context)

@login_required
@group_required(['admin'])
def subscription_fee(request):
    if request.method == 'POST':
        amount = request.POST.get("amount")
        InterestAmount.objects.create(amount=amount)
        messages.success(request, "Subscription Fee Added Successfully")
        return redirect('subscription_fee')
    subscription = InterestAmount.objects.all()
    return render(request,"saving/subscription_fee.html",{'subscription':subscription})

@login_required
@group_required(['admin'])
def edit_subscription_fee(request,id):
    subscription_fee = InterestAmount.objects.get(id=id)
    if request.method == 'POST':
        amount = request.POST.get("amount")
        InterestAmount.objects.filter(id=id).update(amount=amount)
        messages.success(request, "Subscription Fee Edited Successfully")
        return redirect('subscription_fee')
    context = {'subscription_fee':subscription_fee}
    return render(request,"saving/edit_subscription_fee.html",context)


@login_required
@group_required(['admin','staff'])
def monthly_savings_uploads(request):
    # Get monthly savings uploads

    monthly_savings = ( Savings.objects.annotate(month_only=TruncMonth("month")).values("month_only")
        .annotate(
            total_uploaded=Sum("month_saving"),
            total_members=Count("member", distinct=True),
        )
        .order_by("-month_only")
    )

    # Get monthly loanable totals
    month_original_amount = (Savings.objects.annotate(month_only=TruncMonth("month")).values("month_only")
        .annotate(total_original_amount=Sum("original_amount"))
    )
    original_amount_dict = {item["month_only"]: item["total_original_amount"] for item in month_original_amount}


    # Get monthly loanable totals
    monthly_loanable = (Loanable.objects.annotate(month_only=TruncMonth("month")).values("month_only")
        .annotate(total_loanable=Sum("amount"))
    )
    loanable_dict = {item["month_only"]: item["total_loanable"] for item in monthly_loanable}

    # Get monthly investment totals
    monthly_investment = (Investment.objects .annotate(month_only=TruncMonth("month"))
        .values("month_only")
        .annotate(total_investment=Sum("amount"))
    )
    investment_dict = {item["month_only"]: item["total_investment"] for item in monthly_investment}

    # Get monthly interest totals
    monthly_interest = (
        Interest.objects
        .annotate(month_only=TruncMonth("month"))
        .values("month_only")
        .annotate(total_interest=Sum("amount_deducted"))
    )
    interest_dict = {item["month_only"]: item["total_interest"] for item in monthly_interest}

    # Merge results into one list
    monthly_uploads = []
    for item in monthly_savings:
        month = item["month_only"]
        monthly_uploads.append({
            "month_only": month,
            "total_uploaded": item["total_uploaded"] or 0,
            "total_members": item["total_members"] or 0,
            "total_original_amount":original_amount_dict.get(month, 0),
            "total_loanable": loanable_dict.get(month, 0),
            "total_investment": investment_dict.get(month, 0),
            "total_interest": interest_dict.get(month, 0),
        })
    context = { "monthly_uploads": monthly_uploads,}
    return render(request, "saving/monthly_uploads.html", context)

@login_required
@group_required(['admin','staff'])
def view_monthly_savings(request, month):
    # Convert "YYYY-MM" into first day of the month
    try:
        month_start = datetime.strptime(month, "%Y-%m").date()
    except ValueError:
        messages.error(request, "Invalid month format.")
        return redirect("monthly_savings_uploads")

    # Get all savings for the month
    savings_qs = (
        Savings.objects
        .filter(month=month_start)
        .select_related("member")
        .order_by("member__ippis")
    )

    # Pre-fetch all related data for performance
    member_ids = [s.member_id for s in savings_qs]

    loanables = {
        l.member_id: l.amount for l in Loanable.objects.filter(member_id__in=member_ids, month=month_start)
    }
    investments = {
        i.member_id: i.amount for i in Investment.objects.filter(member_id__in=member_ids, month=month_start)
    }
    interests = {
        it.member_id: it.amount_deducted for it in Interest.objects.filter(member_id__in=member_ids, month=month_start)
    }

    # Build structured savings data
    savings_data = []
    for s in savings_qs:
        member = s.member
        savings_data.append({
            "member": member,
            "savings": s.month_saving,
            "interest": interests.get(member.id, 0),
            "investment": investments.get(member.id, 0),
            "loanable": loanables.get(member.id, 0),
        })

    # Paginate results
    paginator = Paginator(savings_data, 100)  
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Grand totals for all data (not per page)
    total_uploaded = sum(item["savings"] for item in savings_data)
    total_interest = sum(item["interest"] for item in savings_data)
    total_investment = sum(item["investment"] for item in savings_data)
    total_loanable = sum(item["loanable"] for item in savings_data)
    total_saving_intresr = total_uploaded + total_interest
    print(total_saving_intresr)
    context = {
        'total_saving_intresr':total_saving_intresr,
        "month": month_start,
        "page_obj": page_obj,
        "total_uploaded": total_uploaded,
        "total_interest": total_interest,
        "total_investment": total_investment,
        "total_loanable": total_loanable,
    }

    return render(request, "saving/monthly_savings_details.html", context)



@login_required
@group_required(['admin'])
def delete_monthly_savings(request):
    selected_month = None
    
    # Set the number of records per page
    per_page = 100

    if request.method == "POST":
        selected_month = request.POST.get("month")
        confirm_delete = request.POST.get("confirm_delete")
        download_excel = request.POST.get("download_excel")

        if not selected_month:
            messages.error(request, "Please select a month.")
            return redirect("delete_monthly_savings")

        try:
            month_date = datetime.strptime(selected_month, "%Y-%m").date()
        except ValueError:
            messages.error(request, "Invalid month format.")
            return redirect("delete_monthly_savings")

        # Handle "Download Excel" and "Confirm Delete" first, as they don't need pagination.
        if download_excel == "yes":
            # Fetch all records for the selected month to include in the Excel file
            savings_list = list(Savings.objects.filter(
                month__month=month_date.month, month__year=month_date.year
            ).select_related("member"))
            
            loanable_list = list(Loanable.objects.filter(
                month__month=month_date.month, month__year=month_date.year
            ).select_related("member"))

            investment_list = list(Investment.objects.filter(
                month__month=month_date.month, month__year=month_date.year
            ).select_related("member"))

            interest_list = list(Interest.objects.filter(
                month__month=month_date.month, month__year=month_date.year
            ).select_related("member"))

            # Create DataFrames
            df_savings = pd.DataFrame([
                {"Member": s.member.member.get_full_name() if s.member and s.member.member else "",
                 "IPPIS": s.member.ippis if s.member else "",
                 "Savings Amount": float(s.month_saving)} for s in savings_list
            ])

            df_loanable = pd.DataFrame([
                {"Member": l.member.member.get_full_name() if l.member and l.member.member else "",
                 "IPPIS": l.member.ippis if l.member else "",
                 "Loanable Amount": float(l.amount)} for l in loanable_list
            ])

            df_investment = pd.DataFrame([
                {"Member": i.member.member.get_full_name() if i.member and i.member.member else "",
                 "IPPIS": i.member.ippis if i.member else "",
                 "Investment Amount": float(i.amount)} for i in investment_list
            ])

            df_interest = pd.DataFrame([
                {"Member": i.member.member.get_full_name() if i.member and i.member.member else "",
                 "IPPIS": i.member.ippis if i.member else "",
                 "Interest Deducted": float(i.amount_deducted)} for i in interest_list
            ])

            # Create Excel with multiple sheets
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = (
                f'attachment; filename="monthly_savings_preview_{month_date.strftime("%Y_%m")}.xlsx"'
            )

            with pd.ExcelWriter(response, engine="openpyxl") as writer:
                if not df_savings.empty:
                    df_savings.to_excel(writer, sheet_name="Savings", index=False)
                if not df_loanable.empty:
                    df_loanable.to_excel(writer, sheet_name="Loanable", index=False)
                if not df_investment.empty:
                    df_investment.to_excel(writer, sheet_name="Investment", index=False)
                if not df_interest.empty:
                    df_interest.to_excel(writer, sheet_name="Interest", index=False)
            
            return response

        if confirm_delete == "yes":
            try:
                with transaction.atomic():
                    savings_deleted, _ = Savings.objects.filter(
                        month__month=month_date.month, month__year=month_date.year
                    ).delete()
                    loanable_deleted, _ = Loanable.objects.filter(
                        month__month=month_date.month, month__year=month_date.year
                    ).delete()
                    investment_deleted, _ = Investment.objects.filter(
                        month__month=month_date.month, month__year=month_date.year
                    ).delete()
                    interest_deleted, _ = Interest.objects.filter(
                        month__month=month_date.month, month__year=month_date.year
                    ).delete()

                    total_deleted = savings_deleted + loanable_deleted + investment_deleted + interest_deleted
                    # ✅ Create delete log
                    DeleteLog.objects.create(
                        user=request.user,
                        action="monthly_savings_delete",
                        month=month_date,
                        records_deleted=total_deleted,
                        remarks=f"Deleted Savings, Loanable, Investment, and Interest for {month_date.strftime('%B %Y')}."
                    )
                    if total_deleted > 0:
                        messages.success(request, f"Successfully deleted all records for {month_date.strftime('%B %Y')}.")
                    else:
                        messages.warning(request, f"No records found for {month_date.strftime('%B %Y')} to delete.")
            except Exception as e:
                messages.error(request, f"Error deleting records: {str(e)}")
            
            return redirect("delete_monthly_savings")

    # This block handles both POST (after a preview) and GET (pagination clicks or initial load)
    selected_month = request.POST.get("month") or request.GET.get("month")
    
    if selected_month:
        try:
            month_date = datetime.strptime(selected_month, "%Y-%m").date()
            # Use querysets here for lazy loading
            savings_queryset = Savings.objects.filter(
                month__month=month_date.month, month__year=month_date.year
            ).select_related("member")
            
            loanable_queryset = Loanable.objects.filter(
                month__month=month_date.month, month__year=month_date.year
            ).select_related("member")
            
            investment_queryset = Investment.objects.filter(
                month__month=month_date.month, month__year=month_date.year
            ).select_related("member")
            
            interest_queryset = Interest.objects.filter(
                month__month=month_date.month, month__year=month_date.year
            ).select_related("member")

        except (ValueError, TypeError):
            messages.error(request, "Invalid month format. Please select a valid month.")
            savings_queryset = Savings.objects.none()
            loanable_queryset = Loanable.objects.none()
            investment_queryset = Investment.objects.none()
            interest_queryset = Interest.objects.none()
    else:
        # Default for initial page load
        savings_queryset = Savings.objects.none()
        loanable_queryset = Loanable.objects.none()
        investment_queryset = Investment.objects.none()
        interest_queryset = Interest.objects.none()

    # Apply pagination to the querysets
    savings_paginator = Paginator(savings_queryset, per_page)
    savings_page = request.GET.get('savings_page', 1)
    savings_list = savings_paginator.get_page(savings_page)

    loanable_paginator = Paginator(loanable_queryset, per_page)
    loanable_page = request.GET.get('loanable_page', 1)
    loanable_list = loanable_paginator.get_page(loanable_page)

    investment_paginator = Paginator(investment_queryset, per_page)
    investment_page = request.GET.get('investment_page', 1)
    investment_list = investment_paginator.get_page(investment_page)

    interest_paginator = Paginator(interest_queryset, per_page)
    interest_page = request.GET.get('interest_page', 1)
    interest_list = interest_paginator.get_page(interest_page)

    return render( request,  "saving/delete_monthly_savings.html",{
            "savings_list": savings_list,"loanable_list": loanable_list,
            "investment_list": investment_list,"interest_list": interest_list,"selected_month": selected_month,})



@login_required
@group_required(['admin','staff'])
def report_view(request):
    start_month_str = request.GET.get('start_month')  # e.g. "2026-01"
    end_month_str = request.GET.get('end_month')       # e.g. "2026-03"
    export_format = request.GET.get('format')

    start_date = None
    end_date = None
    per_page = 100

    # ✅ Parse month inputs and convert to first/last day of month
    if start_month_str and end_month_str:
        try:
            start_date = datetime.strptime(start_month_str, '%Y-%m').date().replace(day=1)
            
            # Get last day of end month
            end_parsed = datetime.strptime(end_month_str, '%Y-%m')
            last_day = calendar.monthrange(end_parsed.year, end_parsed.month)[1]
            end_date = end_parsed.date().replace(day=last_day)

            # Swap if end is before start
            if end_date < start_date:
                start_date, end_date = end_date, start_date

        except (ValueError, TypeError):
            messages.error(request, "Invalid month format. Please use YYYY-MM.")
            return redirect(request.path)

    # 1. Base querysets
    savings_queryset = Savings.objects.select_related('member__member')
    interest_queryset = Interest.objects.select_related('member__member')
    loanable_queryset = Loanable.objects.select_related('member__member')
    investment_queryset = Investment.objects.select_related('member__member')

    # 2. Apply month filters
    if start_date and end_date:
        savings_queryset = savings_queryset.filter(month__range=[start_date, end_date])
        interest_queryset = interest_queryset.filter(month__range=[start_date, end_date])
        loanable_queryset = loanable_queryset.filter(month__range=[start_date, end_date])
        investment_queryset = investment_queryset.filter(month__range=[start_date, end_date])

    # 3. Handle Excel export
    if export_format == 'excel':
        savings_data = list(savings_queryset.order_by('month').values(
            'month', 'month_saving', 'member__member__first_name', 'member__member__last_name', 'member__ippis'
        ))
        interest_data = list(interest_queryset.order_by('month').values(
            'month', 'amount_deducted', 'member__member__first_name', 'member__member__last_name', 'member__ippis'
        ))
        loanable_data = list(loanable_queryset.order_by('month').values(
            'month', 'amount', 'member__member__first_name', 'member__member__last_name', 'member__ippis'
        ))
        investment_data = list(investment_queryset.order_by('month').values(
            'month', 'amount', 'member__member__first_name', 'member__member__last_name', 'member__ippis'
        ))

        df_savings = pd.DataFrame(savings_data).rename(columns={
            'member__member__first_name': 'Member First Name',
            'member__member__last_name': 'Member Last Name',
            'member__ippis': 'IPPIS',
            'month_saving': 'Amount'
        })
        df_interest = pd.DataFrame(interest_data).rename(columns={
            'member__member__first_name': 'Member First Name',
            'member__member__last_name': 'Member Last Name',
            'member__ippis': 'IPPIS',
            'amount_deducted': 'Amount Deducted'
        })
        df_loanable = pd.DataFrame(loanable_data).rename(columns={
            'member__member__first_name': 'Member First Name',
            'member__member__last_name': 'Member Last Name',
            'member__ippis': 'IPPIS',
        })
        df_investment = pd.DataFrame(investment_data).rename(columns={
            'member__member__first_name': 'Member First Name',
            'member__member__last_name': 'Member Last Name',
            'member__ippis': 'IPPIS',
        })

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=financial_report.xlsx'

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            if not df_savings.empty:
                df_savings.to_excel(writer, sheet_name='Savings', index=False)
            if not df_interest.empty:
                df_interest.to_excel(writer, sheet_name='Interest', index=False)
            if not df_loanable.empty:
                df_loanable.to_excel(writer, sheet_name='Loanable', index=False)
            if not df_investment.empty:
                df_investment.to_excel(writer, sheet_name='Investment', index=False)

        return response

    # 4. Calculate totals
    total_savings = savings_queryset.aggregate(Sum('month_saving'))['month_saving__sum'] or 0
    total_interest = interest_queryset.aggregate(Sum('amount_deducted'))['amount_deducted__sum'] or 0
    total_loanable = loanable_queryset.aggregate(Sum('amount'))['amount__sum'] or 0
    total_investment = investment_queryset.aggregate(Sum('amount'))['amount__sum'] or 0

    # 5. Pagination
    savings_paginator = Paginator(savings_queryset.order_by('month'), per_page)
    interest_paginator = Paginator(interest_queryset.order_by('month'), per_page)
    loanable_paginator = Paginator(loanable_queryset.order_by('month'), per_page)
    investment_paginator = Paginator(investment_queryset.order_by('month'), per_page)

    savings_details = savings_paginator.get_page(request.GET.get('savings_page', 1))
    interest_details = interest_paginator.get_page(request.GET.get('interest_page', 1))
    loanable_details = loanable_paginator.get_page(request.GET.get('loanable_page', 1))
    investment_details = investment_paginator.get_page(request.GET.get('investment_page', 1))

    context = {
        'total_savings': total_savings,
        'total_interest': total_interest,
        'total_loanable': total_loanable,
        'total_investment': total_investment,
        'start_month': start_month_str,  
        'end_month': end_month_str,        
        'start_date': start_date,         
        'end_date': end_date,              
        'savings_details': savings_details,
        'interest_details': interest_details,
        'loanable_details': loanable_details,
        'investment_details': investment_details,
    }
    return render(request, 'saving/report.html', context)

# =============member and non member ===============
@login_required
@group_required(['admin','staff'])
def all_member_savings(request):
    group_title = request.GET.get('group')

    users = User.objects.filter(member__isnull=False).order_by('group__title', 'first_name')
    if group_title:
        users = users.filter(group__title=group_title)

    # Prefetch savings
    member_ids = [u.member.id for u in users]
    savings_qs = Savings.objects.filter(member__id__in=member_ids).values('member').annotate(total=Sum('month_saving'))
    savings_map = {entry['member']: entry['total'] or Decimal('0.00') for entry in savings_qs}

    users_data = []
    for user in users:
        total_savings = savings_map.get(user.member.id, Decimal('0.00'))
        users_data.append({
            'user': user,
            'group': user.group.title,
            'total_savings': total_savings,
        })

    # Pagination
    page_number = request.GET.get('page', 1)
    paginator = Paginator(users_data, 50)  # Show 10 users per page
    page_obj = paginator.get_page(page_number)

    all_groups = UserGroup.objects.all()
    group_totals = {}
    for group in all_groups:
        total = Savings.objects.filter(member__member__group=group).aggregate(total=Sum('month_saving'))['total'] or Decimal('0.00')
        group_totals[group.title] = total
    overall_total = sum(group_totals.values(), Decimal('0.00'))

    context = {
        'page_obj': page_obj,
        'all_groups': all_groups,
        'selected_group': group_title,
        'group_totals': group_totals,
        'overall_total': overall_total,
    }
    return render(request, "saving/all_users_savings.html", context)