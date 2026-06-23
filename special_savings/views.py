import openpyxl
from django.shortcuts import render, redirect,get_object_or_404
from decimal import Decimal, InvalidOperation
from datetime import date
from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator,PageNotAnInteger,EmptyPage
from accounts.models import Member
from .models import *


@login_required
# @group_required(['admin','staff'])
def add_forms_type(request):
    forms_types = SavingType.objects.all()

    if request.method == 'POST':
        title = request.POST.get('title')
        request_fee = request.POST.get('request_fee')
        form_type_id = request.POST.get('form_type_id')
        action = request.POST.get('action')  # 'toggle' or 'edit'

        if form_type_id:
            form_type = get_object_or_404(SavingType, id=form_type_id)

            if action == 'toggle':
                form_type.available = not form_type.available
                form_type.save()
                messages.success(request, 'Form type availability updated successfully.')
                return redirect('add_forms_type')

            elif action == 'edit':
                form_type.title = title
                form_type.request_fee = request_fee
               
                form_type.save()
                messages.success(request, 'Form type updated successfully.')
                return redirect('add_forms_type')

        else:
            SavingType.objects.create(title=title, request_fee=request_fee, available=True,)
            messages.success(request, 'Form type created successfully.')
            return redirect('add_forms_type')

    context = {'forms_types': forms_types}
    return render(request, 'special_savings/add_forms_type.html', context)


# Create your views here.
def savings_request_form_payment(request):
    member_info = None
    saving_types = SavingType.objects.filter(available=True)
    
    selected_saving_type_id = request.GET.get("savings_type")

    if request.method == "POST":
        if "search_member" in request.POST:  # step 1: search
            ippis = request.POST.get("ippis")
            try:
                member = Member.objects.get(ippis=ippis)
                member_info = {
                    "id": member.id,
                    "name": f"{member.member.first_name} {member.member.last_name}",
                    "ippis": member.ippis,
                }
            except Member.DoesNotExist:
                messages.error(request, f"No member found with IPPIS {ippis}.")

        elif "make_payment" in request.POST:  
            member_id = request.POST.get("member_id")
            savings_type_id = request.POST.get("savings_type")
            duration_in_months = request.POST.get("duration_in_months")

            try:
                amount = Decimal(request.POST.get("amount"))
            except (ValueError, TypeError, InvalidOperation):
                amount = Decimal("0.00")

            if not savings_type_id:
                messages.error(request, "Please select a Savings Type.")
                return redirect("savings_request_form_payment")

            member = get_object_or_404(Member, id=member_id)
            savings_type = get_object_or_404(SavingType, id=savings_type_id)

            if SpecialSavingsTergetSavingsRequestForm.objects.filter(
                member=member, 
                savings_type=savings_type,
                status = "paid"
                
            ).exists():
                messages.warning(
                    request, 
                    f"{member} has already paid the request fee for {savings_type.title}."
                )
                return redirect("savings_request_form_payment")
            if duration_in_months == "":
                duration_in_months = None
            else:
                duration_in_months = int(duration_in_months)
                
            SpecialSavingsTergetSavingsRequestForm.objects.create(
                member=member,
                savings_type=savings_type,
                form_fee=savings_type.request_fee,
                amount=amount,
                duration_in_months= duration_in_months,
                created_by=request.user,
                status = "used"
            )

            messages.success(
                request, 
                f"Savings request fee of ₦{savings_type.request_fee:,.2f} recorded for {member}."
            )
            return redirect("savings_request_form_payment")

    # Filter records
    members = SpecialSavingsTergetSavingsRequestForm.objects.select_related(
        'member__member', 'savings_type', 'created_by'
    ).order_by('-date_created')

    if selected_saving_type_id:
        members = members.filter(savings_type_id=selected_saving_type_id)

    # Aggregates
    saving = members.aggregate(total=Sum('amount'))['total'] or Decimal("0.00")
    fee = members.aggregate(total=Sum('form_fee'))['total'] or Decimal("0.00")
    saving_req_form = members.count()

    # Pagination
    page_number = request.GET.get('page')
    paginator = Paginator(members, 100)
    page_obj = paginator.get_page(page_number)

    context = {
        "fee": fee,
        "saving_types": saving_types,
        "saving": saving,
        "saving_req_form": saving_req_form,
        "page_obj": page_obj,
        "paginator": paginator,  # Added for template
        "selected_saving_type_id": selected_saving_type_id,
        "member_info": member_info,
    }

    return render(request, "special_savings/savings_request_form_payment.html", context)




@login_required
def upload_special_savings(request):
    if request.method == "POST" and request.FILES.get("file"):
        selected_month = request.POST.get("month")
        if selected_month:
            year, m = map(int, selected_month.split("-"))
            month = date(year, m, 1)
        else:
            month = date.today().replace(day=1)

        wb = openpyxl.load_workbook(request.FILES["file"])
        ws = wb.active

        created = skipped = invalid = not_found = duplicate = 0

        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            ippis, amount = row

            # Validate ippis & amount
            if not ippis or not amount:
                invalid += 1
                continue

            try:
                member = Member.objects.get(ippis=int(ippis))
            except Member.DoesNotExist:
                not_found += 1
                continue

            try:
                amount = Decimal(amount)
                if amount <= 0:
                    raise InvalidOperation
            except:
                invalid += 1
                continue

            # Check duplicate month
            if SpecialSavings.objects.filter(member=member, month=month).exists():
                duplicate += 1
                continue

            SpecialSavings.objects.create(
                member=member,
                month=month,
                month_savings=amount,
                created_by=request.user
            )
            created += 1

        # Messages
        if created:
            messages.success(request, f"{created} special savings added successfully.")

        if not_found:
            messages.warning(request, f"{not_found} records skipped — IPPIS not found.")

        if duplicate:
            messages.warning(request, f"{duplicate} records skipped — already exists for this month.")

        if invalid:
            messages.error(request, f"{invalid} records skipped — invalid or missing amount/IPPIS.")

        return redirect("upload_special_savings")

    return render(request, "special_savings/upload_special_savings.html")


def admin_create_special_savings(request):

    if request.method == "POST":
        ippis = request.POST.get("ippis")
        amount = request.POST.get("amount")
        month = request.POST.get("month")

        if not ippis or not amount or not month:
            messages.error(request, "All fields are required.")
            return redirect("create_special_savings")

        try:
            member = Member.objects.get(ippis=ippis)
        except Member.DoesNotExist:
            messages.error(request, "Member not found.")
            return redirect("create_special_savings")


        try:
            year, m = map(int, month.split("-"))
            month = date(year, m, 1)
            amount = Decimal(amount)
            if amount <= 0:
                raise ValueError
        except:
            messages.error(request, "Invalid amount or month.")
            return redirect("create_special_savings")

        # Prevent duplicate
        if SpecialSavings.objects.filter(member=member, month=month).exists():
            messages.warning(request, "Special savings already exists for this month.")
            return redirect("create_special_savings")

        # Create savings
        SpecialSavings.objects.create(
            member=member,
            month=month,
            month_savings=amount,
            created_by=request.user
        )

        

        messages.success(request, "Special savings created successfully.")
        return redirect("create_special_savings")

    return render(request, "special_savings/admin_create_special_savings.html")


@login_required
def special_savings_list(request):
    
    monthly_data = SpecialSavings.objects.values('month').annotate(
     total_amount=Sum('month_savings'),
     member_count=Count('member', distinct=True)
    ).order_by('-month')
    
     # Calculate overall totals
    overall_stats = SpecialSavings.objects.aggregate(
        total_members=Count('member', distinct=True),
        total_amount=Sum('month_savings')
    )
   
    paginator = Paginator(monthly_data, 50)
    page_number = request.GET.get('page')
    special_sav = paginator.get_page(page_number)
    context = {
        'special_sav':special_sav,
        'monthly_data': monthly_data,
        'overall_stats': overall_stats,
    }
    
    return render(request, "special_savings/special_savings_list.html", context)


def monthly_special_savings_detail(request, year, month):
    """
    Display detailed savings for a specific month showing individual member contributions
    """
    from datetime import date
    
    # Create date object for the specified month
    month_date = date(year, month, 1)
    
    # Get all savings for this month
    savings_list = (
        SpecialSavings.objects
        .filter(month=month_date)
        .select_related('member', 'created_by')
        .order_by('-month_savings')
    )
    
    # Calculate month statistics
    month_stats = savings_list.aggregate(
        total_members=Count('id'),
        total_amount=Sum('month_savings')
    )
    paginator = Paginator(savings_list, 100)
    page_number = request.GET.get('page')
    details_savings_list = paginator.get_page(page_number)
    context = {
        'details_savings_list': details_savings_list,
        'month_date': month_date,
        'savings_list': savings_list,
        'month_stats': month_stats,
    }
    
    return render(request, 'special_savings/special_savings_detail.html', context)


def delete_monthly_savings(request, year, month):
    """
    Delete all savings for a specific month
    """
    from django.contrib import messages
    from django.shortcuts import redirect
    from datetime import date
    
    if request.method == 'POST':
        # Create date object for the specified month
        month_date = date(year, month, 1)
        
        # Get all savings for this month
        savings_to_delete = SpecialSavings.objects.filter(month=month_date)
        
        # Count before deletion
        count = savings_to_delete.count()
        
        if count > 0:
            # Delete all savings for this month
            savings_to_delete.delete()
            messages.success(
                request, 
                f'Successfully deleted {count} savings record(s) for {month_date.strftime("%B %Y")}'
            )
        else:
            messages.info(request, f'No savings found for {month_date.strftime("%B %Y")}')
        
        return redirect('special_savings_list')
    
    # If not POST, redirect back
    return redirect('special_savings_list')


# Admin views all pending withdrawal requests
@login_required
def admin_special_savings_withdrawals(request):
    withdrawals = SpecialSavingsWithdrawal.objects.select_related("member__member").order_by("-requested_at")

    # Total savings deposited
    total_savings = SpecialSavings.objects.aggregate(total=Sum("month_savings") )["total"] or 0

    # Total approved withdrawals
    total_withdrawals = SpecialSavingsWithdrawal.objects.filter(
        status="approved").aggregate(total=Sum("amount"))["total"] or 0

    # Available balance
    total_available = total_savings - total_withdrawals

    context = {
        "withdrawals": withdrawals,
        "total_savings": total_savings,
        "total_withdrawals": total_withdrawals,
        "total_available": total_available,
    }

    return render(request,"special_savings/admin_special_withdrawals.html",context,)

# Admin approves or rejects
@login_required
def review_special_savings_withdrawal(request, pk):
    withdrawal = get_object_or_404(SpecialSavingsWithdrawal, pk=pk)

    if request.method == "POST":
        action = request.POST.get("action")
        try:
            if action == "approve":
                withdrawal.approve(reviewed_by=request.user)
                messages.success(request, f"Withdrawal of ₦{withdrawal.amount} approved.")
            elif action == "reject":
                withdrawal.reject(reviewed_by=request.user)
                messages.warning(request, "Withdrawal request rejected.")
        except ValueError as e:
            messages.error(request, str(e))

        return redirect("admin_special_savings_withdrawals")

    return render(request, "special_savings/approve_special_withdrawal.html", {"withdrawal": withdrawal})


# ============== Target Savings Views can be added below similarly ================

@login_required
def upload_target_savings(request):
    if request.method == "POST" and request.FILES.get("file"):

        # Get selected month
        selected_month = request.POST.get("month")
        if selected_month:
            year, m = map(int, selected_month.split("-"))
            month = date(year, m, 1)
        else:
            month = date.today().replace(day=1)

        wb = openpyxl.load_workbook(request.FILES["file"])
        ws = wb.active

        created = 0
        invalid = []
        not_found = []
        duplicate = []

        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            ippis, amount = row

            # Validate raw data
            if not ippis or not amount:
                invalid.append(str(ippis))
                continue

            try:
                member = Member.objects.get(ippis=int(ippis))
            except Member.DoesNotExist:
                not_found.append(str(ippis))
                continue

            try:
                amount = Decimal(amount)
                if amount <= 0:
                    raise InvalidOperation
            except:
                invalid.append(str(ippis))
                continue

            # Duplicate month check
            if TargetSavings.objects.filter(member=member, month=month).exists():
                duplicate.append(str(ippis))
                continue

            # Save record
            TargetSavings.objects.create(
                member=member,
                month=month,
                month_savings=amount,
                created_by=request.user
            )
            created += 1

        # User feedback
        if created:
            messages.success(request, f"{created} target savings uploaded successfully.")

        if not_found:
            messages.warning(request, f"IPPIS not found: {', '.join(not_found)}")

        if duplicate:
            messages.warning(request, f"Duplicate month skipped: {', '.join(duplicate)}")

        if invalid:
            messages.error(request, f"Invalid or missing data: {', '.join(invalid)}")

        return redirect("upload_target_savings")

    return render(request, "special_savings/upload_target_savings.html")


def create_target_savings(request):

    if request.method == "POST":
        ippis = request.POST.get("ippis")
        amount = request.POST.get("amount")
        month = request.POST.get("month")

        if not ippis or not amount or not month:
            messages.error(request, "All fields are required.")
            return redirect("create_target_savings")

        try:
            member = Member.objects.get(ippis=ippis)
        except Member.DoesNotExist:
            messages.error(request, "Member not found.")
            return redirect("create_target_savings")


        try:
            year, m = map(int, month.split("-"))
            month = date(year, m, 1)
            amount = Decimal(amount)
            if amount <= 0:
                raise ValueError
        except:
            messages.error(request, "Invalid amount or month.")
            return redirect("create_target_savings")

        # Prevent duplicate
        if TargetSavings.objects.filter(member=member, month=month).exists():
            messages.warning(request, "Target savings already exists for this month.")
            return redirect("create_target_savings")

        # Create savings
        TargetSavings.objects.create(
            member=member,
            month=month,
            month_savings=amount,
            created_by=request.user
        )

        

        messages.success(request, "Special savings created successfully.")
        return redirect("create_target_savings")

    return render(request, "special_savings/create_target_savings.html")



@login_required
def target_savings_list(request):
    monthly_data = TargetSavings.objects.values('month').annotate(
     total_amount=Sum('month_savings'),
     member_count=Count('member', distinct=True)
    ).order_by('-month')
    
     # Calculate overall totals
    overall_stats = TargetSavings.objects.aggregate(
        total_members=Count('member', distinct=True),
        total_amount=Sum('month_savings')
    )
   
    paginator = Paginator(monthly_data, 50)
    page_number = request.GET.get('page')
    special_sav = paginator.get_page(page_number)
    context = {
        'special_sav':special_sav,
        'monthly_data': monthly_data,
        'overall_stats': overall_stats,
    }

    return render(request, "special_savings/target_savings_list.html", context)


def monthly_target_savings_detail(request, year, month):
    """
    Display detailed savings for a specific month showing individual member contributions
    """
    from datetime import date
    
    # Create date object for the specified month
    month_date = date(year, month, 1)
    
    # Get all savings for this month
    savings_list = (
        TargetSavings.objects
        .filter(month=month_date)
        .select_related('member', 'created_by')
        .order_by('-month_savings')
    )
    
    # Calculate month statistics
    month_stats = savings_list.aggregate(
        total_members=Count('id'),
        total_amount=Sum('month_savings')
    )
    paginator = Paginator(savings_list, 100)
    page_number = request.GET.get('page')
    details_savings_list = paginator.get_page(page_number)
    context = {
        'details_savings_list': details_savings_list,
        'month_date': month_date,
        'savings_list': savings_list,
        'month_stats': month_stats,
    }

    return render(request, 'special_savings/target_savings_detail.html', context)


def delete_monthly_target_savings(request, year, month):
    """
    Delete all savings for a specific month
    """
    # from django.contrib import messages
    # from django.shortcuts import redirect
    # from datetime import date
    
    if request.method == 'POST':
        # Create date object for the specified month
        month_date = date(year, month, 1)
        
        # Get all savings for this month
        savings_to_delete = TargetSavings.objects.filter(month=month_date)
        
        # Count before deletion
        count = savings_to_delete.count()
        
        if count > 0:
            # Delete all savings for this month
            savings_to_delete.delete()
            messages.success(
                request, 
                f'Successfully deleted {count} savings record(s) for {month_date.strftime("%B %Y")}'
            )
        else:
            messages.info(request, f'No savings found for {month_date.strftime("%B %Y")}')
        
        return redirect('target_savings_list')
    
    # If not POST, redirect back
    return redirect('target_savings_list')



# Admin views all pending withdrawal requests
@login_required
def admin_target_savings_withdrawals(request):
    withdrawals = TargetSavingsWithdrawal.objects.select_related("member__member").order_by("-requested_at")

    # Total savings deposited
    total_savings = TargetSavings.objects.aggregate(total=Sum("month_savings") )["total"] or 0

    # Total approved withdrawals
    total_withdrawals = TargetSavingsWithdrawal.objects.filter(
        status="approved").aggregate(total=Sum("amount"))["total"] or 0

    # Available balance
    total_available = total_savings - total_withdrawals

    context = {
        "withdrawals": withdrawals,
        "total_savings": total_savings,
        "total_withdrawals": total_withdrawals,
        "total_available": total_available,
    }

    return render(request,"special_savings/admin_target_withdrawals.html",context,)

# Admin approves or rejects
@login_required
def review_target_savings_withdrawal(request, pk):
    withdrawal = get_object_or_404(TargetSavingsWithdrawal, pk=pk)

    if request.method == "POST":
        action = request.POST.get("action")
        try:
            if action == "approve":
                withdrawal.approve(reviewed_by=request.user)
                messages.success(request, f"Withdrawal of ₦{withdrawal.amount} approved.")
            elif action == "reject":
                withdrawal.reject(reviewed_by=request.user)
                messages.warning(request, "Withdrawal request rejected.")
        except ValueError as e:
            messages.error(request, str(e))

        return redirect("admin_target_savings_withdrawals")

    return render(request, "special_savings/approve_target_withdrawal.html", {"withdrawal": withdrawal})
